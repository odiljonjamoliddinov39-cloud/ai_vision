import asyncio
import io
import json
from pathlib import Path

import cv2
import numpy as np
import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from starlette.datastructures import Headers

from api import server
from database.catalog_db import CatalogDB
from recognition.embedding import image_embedding
from recognition.product_recognizer import ProductRecognizer


ROOT = Path(__file__).resolve().parents[1]


def _reference_image(color: tuple[int, int, int]) -> np.ndarray:
    image = np.zeros((80, 120, 3), dtype=np.uint8)
    image[:, :] = color
    cv2.rectangle(image, (18, 15), (100, 68), (255, 255, 255), 3)
    return image


def _upload(image: np.ndarray, name: str) -> UploadFile:
    ok, encoded = cv2.imencode(".jpg", image)
    assert ok
    return UploadFile(
        filename=name,
        file=io.BytesIO(encoded.tobytes()),
        headers=Headers({"content-type": "image/jpeg"}),
    )


def _catalog_with_item(tmp_path: Path, name: str = "Blue crate") -> tuple[CatalogDB, dict]:
    db = CatalogDB(str(tmp_path / "catalog.db"))
    item = db.create_item("warehouse-a", name)
    for index, color in enumerate(((220, 60, 20), (210, 70, 25)), start=1):
        image = _reference_image(color)
        db.add_image(
            item_id=item["id"],
            filename=f"reference_{index}.jpg",
            url=f"/snapshots/catalog/reference_{index}.jpg",
            embedding=image_embedding(image),
            width_px=image.shape[1],
            height_px=image.shape[0],
        )
    return db, item


def test_catalog_keeps_multiple_reference_images_and_candidates(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path)

    stored = db.get_item(item["id"])
    candidates = db.reference_candidates()

    assert stored["image_count"] == 2
    assert len(stored["images"]) == 2
    assert [candidate["name"] for candidate in candidates] == ["Blue crate", "Blue crate"]
    assert all(candidate["embedding"] for candidate in candidates)


def test_catalog_upload_requires_and_persists_multiple_images(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db = CatalogDB(str(tmp_path / "catalog.db"))
    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "CATALOG_IMAGE_DIR", tmp_path / "images")
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", tmp_path / "missing-health.json")
    first = _upload(_reference_image((220, 60, 20)), "front.jpg")

    with pytest.raises(HTTPException, match="at least two") as error:
        asyncio.run(server.create_catalog_item("warehouse-a", "Blue crate", [first]))
    assert error.value.status_code == 400

    files = [
        _upload(_reference_image((220, 60, 20)), "front.jpg"),
        _upload(_reference_image((210, 70, 25)), "side.jpg"),
    ]
    payload = asyncio.run(server.create_catalog_item("warehouse-a", "Blue crate", files))

    assert payload["item"]["image_count"] == 2
    assert len(list((tmp_path / "images" / "warehouse-a" / payload["item"]["id"]).glob("*.jpg"))) == 2
    assert payload["recognition"]["run"]["status"] == "completed"


def test_catalog_only_recognizer_never_calls_provider_for_enrolled_item(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, _ = _catalog_with_item(tmp_path)

    class FailingProvider:
        def recognize(self, image):  # pragma: no cover - proves it is never called
            raise AssertionError("catalog-only recognition must not call the general provider")

    recognizer = ProductRecognizer(
        provider=FailingProvider(),
        catalog_only=True,
        catalog_db=db,
        similarity_threshold=0.90,
    )
    result = recognizer.recognize(_reference_image((220, 60, 20)))

    assert result.name == "Blue crate"
    assert result.source == "catalog_reference"
    recognizer.close()


def test_recognition_run_counts_only_catalog_item_and_records_3d_measurement(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path)
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [],
                "last_spatial_objects_by_camera": {
                    "Camera 1": [
                        {
                            "inventory_name": "Blue crate",
                            "quantity": 7,
                            "width_m": 0.42,
                            "height_m": 0.31,
                            "depth_m": 0.28,
                            "method": "monocular_ground_plane",
                        },
                        {"inventory_name": "Ignored unknown item", "quantity": 99},
                    ]
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)
    monkeypatch.setenv("CATALOG_RECOGNITION_INTERVAL_HOURS", "12")

    payload = server._run_catalog_recognition("warehouse-a")
    result = payload["results"][0]

    assert result["item_id"] == item["id"]
    assert result["item_name"] == "Blue crate"
    assert result["quantity"] == 7
    assert result["confidence"] == 1.0
    assert result["width_m"] == 0.42
    assert result["measurement_method"] == "monocular_ground_plane"
    assert payload["schedule"]["interval_hours"] == 12


def test_excel_export_is_formatted_and_immediately_readable(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path)
    run_id = db.start_run("warehouse-a", interval_hours=12, camera_count=10)
    db.add_result(
        run_id,
        item["id"],
        "Blue crate",
        quantity=12,
        confidence=0.97,
        dimensions_m=(0.42, 0.31, 0.28),
        measurement_method="monocular_ground_plane",
    )
    db.complete_run(run_id)
    monkeypatch.setattr(server, "_catalog_db", db)

    content = server._catalog_export_workbook("warehouse-a")
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["Detected Items"]

    assert sheet["A1"].value == "AI Vision — Detected Item Count"
    assert [sheet.cell(5, column).value for column in range(1, 8)] == [
        "Item",
        "Count",
        "Confidence",
        "Width (cm)",
        "Height (cm)",
        "Depth (cm)",
        "3D method",
    ]
    assert sheet["A6"].value == "Blue crate"
    assert sheet["B6"].value == 12
    assert sheet["C6"].number_format == "0.0%"
    assert sheet.freeze_panes == "A6"
    assert sheet.column_dimensions["A"].width >= 25
    assert "DetectedItems" in sheet.tables


def test_dashboard_exposes_multi_image_catalog_results_excel_and_3d_views():
    source = (ROOT / "dashboard-v2" / "app.js").read_text(encoding="utf-8")
    styles = (ROOT / "dashboard-v2" / "styles.css").read_text(encoding="utf-8")
    compose = (ROOT / "docker-compose.yml").read_text(encoding="utf-8")

    assert 'name="images" type="file" accept="image/*" multiple required' in source
    assert "at least two reference images" in source
    assert "Export to Excel" in source
    assert 'class="detected-list-head"' in source
    assert "renderCatalogDimensions" in source
    assert "Waiting for video" in source
    assert "setFeedBadgeLive(image, true)" in source
    assert "not recording continuous video" in source
    assert ".detected-list-head" in styles
    assert ".dimension-visual" in styles
    assert '.catalog-form input:not([type="file"])' in styles
    assert "appearance: none;" in styles
    assert 'class="catalog-upload-help" data-image-count' in source
    assert ".catalog-form > button[type=\"submit\"]" in styles
    assert "grid-row: 1;" in styles
    assert "function currentOperationalAlerts()" in source
    assert "No active alerts" in source
    assert "Camera Offline" not in source
    assert "NVR Disconnected" not in source
    assert "Low Production Rate" not in source
    assert 'CATALOG_RECOGNITION_INTERVAL_HOURS: "12"' in compose


def test_live_catalog_recognition_samples_repeatedly_and_persists_a_completed_run(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path, name="Blue crate")
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [],
                "last_spatial_objects_by_camera": {
                    "Camera 1": [{"inventory_name": "Blue crate", "quantity": 3}]
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)
    monkeypatch.setattr(server, "CATALOG_LIVE_RUN_SAMPLE_INTERVAL_SECONDS", 0.05)
    server._live_catalog_runs.clear()

    started_at = server.datetime.now(server.timezone.utc)
    ends_at = started_at + server.timedelta(seconds=0.2)
    server._live_catalog_runs["warehouse-a"] = {
        "status": "running",
        "started_at": started_at.isoformat(),
        "ends_at": ends_at.isoformat(),
        "items": {},
    }

    asyncio.run(server._run_live_catalog_recognition("warehouse-a", ends_at))

    state = server._live_catalog_runs["warehouse-a"]
    assert state["status"] == "completed"
    assert state["items"][item["id"]]["quantity"] == 3

    run = db.latest_run("warehouse-a")
    assert run is not None
    assert run["status"] == "completed"
    results = db.latest_results("warehouse-a")
    assert results[0]["item_name"] == "Blue crate"
    assert results[0]["quantity"] == 3


def test_live_catalog_recognition_only_matches_catalog_items(tmp_path, monkeypatch):
    # Confirms the requirement that a live run only ever recognizes items
    # enrolled via AI Check-in - it reuses _catalog_match_current_frame,
    # which is scoped to db.list_items(), not raw detector output.
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, _ = _catalog_with_item(tmp_path, name="Blue crate")
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [],
                "last_spatial_objects_by_camera": {
                    "Camera 1": [
                        {"inventory_name": "Blue crate", "quantity": 1},
                        {"inventory_name": "Random forklift", "quantity": 5},
                    ]
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)

    matches = server._catalog_match_current_frame("warehouse-a")

    assert [match["item_name"] for match in matches] == ["Blue crate"]


def test_catalog_recognition_matches_checked_in_item_from_yolo_crop(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path, name="Baget Box")
    snapshot_dir = tmp_path / "snapshots"
    snapshot_dir.mkdir()
    frame = np.zeros((180, 240, 3), dtype=np.uint8)
    frame[40:120, 50:170] = _reference_image((220, 60, 20))
    cv2.imwrite(str(snapshot_dir / "latest_stream_slot_1.jpg"), frame)
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [{"name": "NVR Camera 2", "slot_number": 1}],
                "last_spatial_objects_by_camera": {"NVR Camera 2": []},
                "last_detections_by_camera": {
                    "NVR Camera 2": [
                        {
                            "class_name": "cardboard box",
                            "quantity": 4,
                            "width_m": 0.91,
                            "height_m": 0.79,
                            "depth_m": 0.5,
                            "method": "monocular_ground_plane",
                            "bbox": {"x1": 50, "y1": 40, "x2": 170, "y2": 120},
                        }
                    ]
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "SNAPSHOT_DIR", snapshot_dir)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)

    matches = server._catalog_match_current_frame("warehouse-a")

    assert matches == [
        {
            "item_id": str(item["id"]),
            "item_name": "Baget Box",
            "quantity": 4,
            "confidence": matches[0]["confidence"],
            "dimensions_m": (0.91, 0.79, 0.5),
            "measurement_method": "monocular_ground_plane",
        }
    ]
    assert matches[0]["confidence"] >= 0.9


def test_catalog_recognition_runs_fresh_yolo_when_cached_detections_are_empty(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path, name="Baget Box")
    snapshot_dir = tmp_path / "snapshots"
    snapshot_dir.mkdir()
    frame = np.zeros((180, 240, 3), dtype=np.uint8)
    frame[40:120, 50:170] = _reference_image((220, 60, 20))
    cv2.imwrite(str(snapshot_dir / "latest_stream_slot_1.jpg"), frame)
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [{"name": "NVR Camera 2", "slot_number": 1}],
                "last_spatial_objects_by_camera": {"NVR Camera 2": []},
                "last_detections_by_camera": {"NVR Camera 2": []},
            }
        ),
        encoding="utf-8",
    )

    class FakeDetector:
        def __init__(self, **kwargs):
            self.prompts = kwargs["class_prompts"]

        def detect(self, _frame):
            detection = type(
                "Detection",
                (),
                {
                    "class_name": "box",
                    "confidence": 0.82,
                    "box": (50, 40, 170, 120),
                    "quantity": 1,
                },
            )()
            return [detection]

    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "SNAPSHOT_DIR", snapshot_dir)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)
    monkeypatch.setattr(server, "Detector", FakeDetector)
    monkeypatch.setattr(server, "_catalog_yolo_detector", None)
    monkeypatch.setattr(server, "_catalog_yolo_detector_key", None)

    matches = server._catalog_match_current_frame("warehouse-a")

    assert matches[0]["item_id"] == item["id"]
    assert matches[0]["item_name"] == "Baget Box"
    assert matches[0]["quantity"] == 1
    assert matches[0]["confidence"] == pytest.approx(0.82)


def test_single_catalog_item_counts_current_yolo_box_even_when_reference_similarity_is_low(tmp_path, monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path, name="Baget Box")
    snapshot_dir = tmp_path / "snapshots"
    snapshot_dir.mkdir()
    frame = np.zeros((180, 240, 3), dtype=np.uint8)
    frame[40:120, 50:170] = _reference_image((35, 160, 220))
    cv2.imwrite(str(snapshot_dir / "latest_stream_slot_1.jpg"), frame)
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [{"name": "NVR Camera 2", "slot_number": 1}],
                "last_spatial_objects_by_camera": {"NVR Camera 2": []},
                "last_detections_by_camera": {"NVR Camera 2": []},
            }
        ),
        encoding="utf-8",
    )

    class FakeDetector:
        def __init__(self, **kwargs):
            self.prompts = kwargs["class_prompts"]

        def detect(self, _frame):
            detection = type(
                "Detection",
                (),
                {
                    "class_name": "cardboard box",
                    "confidence": 0.56,
                    "box": (50, 40, 170, 120),
                    "quantity": 3,
                    "width_m": 0.8,
                    "height_m": 0.6,
                    "depth_m": 0.4,
                    "method": "monocular_ground_plane",
                },
            )()
            return [detection]

    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "SNAPSHOT_DIR", snapshot_dir)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)
    monkeypatch.setattr(server, "Detector", FakeDetector)
    monkeypatch.setattr(server, "_catalog_yolo_detector", None)
    monkeypatch.setattr(server, "_catalog_yolo_detector_key", None)
    monkeypatch.setattr(server, "_read_yaml", lambda _path: {"detection": {}, "spatial_analysis": {"enabled": False}})

    matches = server._catalog_match_current_frame("warehouse-a")

    assert matches[0] == {
        "item_id": str(item["id"]),
        "item_name": "Baget Box",
        "quantity": 3,
        "confidence": pytest.approx(0.75),
        "dimensions_m": (0.8, 0.6, 0.4),
        "measurement_method": "monocular_ground_plane",
    }


def test_live_catalog_recognition_http_endpoints_report_progress(tmp_path, monkeypatch):
    from fastapi.testclient import TestClient

    monkeypatch.delenv("DATABASE_URL", raising=False)
    db, item = _catalog_with_item(tmp_path, name="Blue crate")
    health_path = tmp_path / "detection_health.json"
    health_path.write_text(
        json.dumps(
            {
                "cameras": [],
                "last_spatial_objects_by_camera": {
                    "Camera 1": [{"inventory_name": "Blue crate", "quantity": 2}]
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(server, "_catalog_db", db)
    monkeypatch.setattr(server, "DETECTION_HEALTH_PATH", health_path)
    monkeypatch.setattr(server, "CATALOG_LIVE_RUN_DURATION_SECONDS", 0.2)
    monkeypatch.setattr(server, "CATALOG_LIVE_RUN_SAMPLE_INTERVAL_SECONDS", 0.05)
    server._live_catalog_runs.clear()

    with TestClient(server.app) as client:
        start = client.post("/api/catalog/recognition/run-live", params={"scope_id": "warehouse-a"})
        assert start.status_code == 200
        assert start.json()["running"] is True

        again = client.post("/api/catalog/recognition/run-live", params={"scope_id": "warehouse-a"})
        assert again.status_code == 409

        import time

        for _ in range(50):
            status = client.get(
                "/api/catalog/recognition/run-live/status", params={"scope_id": "warehouse-a"}
            ).json()
            if not status["running"]:
                break
            time.sleep(0.05)

        assert status["running"] is False
        assert status["results"][0]["item_name"] == "Blue crate"
        assert status["results"][0]["quantity"] == 2

        results = client.get("/api/catalog/results", params={"scope_id": "warehouse-a"}).json()
        assert results["results"][0]["item_id"] == item["id"]


def test_dashboard_run_recognition_button_uses_immediate_catalog_pass():
    source = (ROOT / "dashboard-v2" / "app.js").read_text(encoding="utf-8")

    assert 'data-run-live-recognition' in source
    assert '"Run recognition now"' in source
    assert 'catalogApiPath("/api/catalog/recognition/run")' in source
    assert 'catalogApiPath("/api/catalog/recognition/run-live/status")' not in source
    assert "s left" not in source
    assert "Recognition complete." in source


def test_catalog_results_panel_does_not_show_cached_warehouse_movements():
    source = (ROOT / "dashboard-v2" / "app.js").read_text(encoding="utf-8")

    assert "function liveAiCheckInTableHtml(movements)" not in source
    assert 'accountsApi("/api/warehouse/movements?limit=50")' not in source
    assert "catalogResultsTableHtml(payload.results)" in source
    assert "No checked-in AI item was recognized" in source
