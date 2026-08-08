"""
FastAPI control server for the AI Vision Assistant dashboard.

Run:
    uvicorn api.server:app --reload --host 0.0.0.0 --port 8000
"""

from __future__ import annotations

import io
import ipaddress
import json
import os
import re
import secrets
import signal
import struct
import subprocess
import sys
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlsplit

import yaml
from fastapi import (
    FastAPI,
    File,
    Form,
    HTTPException,
    Request,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
import asyncio
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from webauthn import (
    generate_authentication_options,
    generate_registration_options,
    options_to_json,
    verify_authentication_response,
    verify_registration_response,
)
from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
from webauthn.helpers.structs import (
    AuthenticatorAttachment,
    AuthenticatorSelectionCriteria,
    PublicKeyCredentialDescriptor,
    ResidentKeyRequirement,
    UserVerificationRequirement,
)

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from discovery import discover_device  # noqa: E402
from discovery.portscan import DiscoveryHostError, resolve_and_guard  # noqa: E402
from discovery.providers import StreamCredentials, enumerate_streams  # noqa: E402
from database.access_control_db import AccessControlDB  # noqa: E402
from database.camera_db import CameraDB  # noqa: E402
from database.vision_config_db import VisionConfigDB  # noqa: E402
from database.device_db import DeviceDB  # noqa: E402
from database.accounts_db import AccountsDB  # noqa: E402
from database.catalog_db import CatalogDB  # noqa: E402
from database.security_audit_db import SecurityAuditDB  # noqa: E402
from database.tracking_db import TrackingDB  # noqa: E402
from database.vision_db import VisionDB  # noqa: E402
from database.warehouse_db import WarehouseDB  # noqa: E402
from dataset_builder import DatasetBuilder, DatasetError, DatasetTrainingManager  # noqa: E402
from detection.detector import Detector  # noqa: E402
from detection.spatial import SpatialAnalyzer  # noqa: E402
from inventory import InventoryCandidate, VisibleInventoryCounter  # noqa: E402
from streams import StreamManager, StreamSessionConfig  # noqa: E402
from api.vision_routes import router as vision_router  # noqa: E402
from warehouse_engine.database import EngineDatabase  # noqa: E402
from warehouse_engine.rules import parse_task_prompt  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "config.yaml"
LOG_PATH = ROOT / "logs" / "events.log"
SNAPSHOT_DIR = ROOT / "snapshots"
INVENTORY_PATH = ROOT / "logs" / "inventory.json"
INVENTORY_IMAGE_DIR = SNAPSHOT_DIR / "inventory"
CATALOG_IMAGE_DIR = SNAPSHOT_DIR / "catalog"
DASHBOARD_V2_DIR = ROOT / "dashboard-v2"
TRACKING_DB_PATH = ROOT / "database" / "tracking.db"
WAREHOUSE_DB_PATH = ROOT / "database" / "warehouse.db"
CAMERA_DB_PATH = ROOT / "database" / "cameras.db"
DEVICE_DB_PATH = ROOT / "database" / "devices.db"
SECURITY_AUDIT_DB_PATH = ROOT / "database" / "security_audit.db"
ACCESS_CONTROL_DB_PATH = ROOT / "database" / "access_control.db"
CATALOG_DB_PATH = ROOT / "database" / "catalog.db"
CATALOG_PROMPTS_PATH = ROOT / "logs" / "catalog_prompts.json"
PRODUCT_FINGERPRINTS_PATH = ROOT / "logs" / "product_fingerprints.json"
ACTIVE_MODELS_PATH = ROOT / "models" / "active_models.json"
WAREHOUSE_ENGINE_DB_PATH = ROOT / "database" / "warehouse_engine.db"
ACCOUNTS_DB_PATH = ROOT / "database" / "accounts.db"
DETECTION_STDOUT_PATH = ROOT / "logs" / "detection_stdout.log"
DETECTION_STDERR_PATH = ROOT / "logs" / "detection_stderr.log"
DETECTION_HEALTH_PATH = ROOT / "logs" / "detection_health.json"
DETECTION_PID_PATH = ROOT / "logs" / "detection.pid"
MAX_CAMERA_SLOTS = 100
DEFAULT_ALLOWED_ORIGINS = [
    "https://ai-vision-dashboard-phi.vercel.app",
    "http://localhost:8000",
    "http://127.0.0.1:8000",
]

app = FastAPI(title="AI Vision Control API", version="0.1.0")
app.include_router(vision_router)
app.routes[:] = [route for route in app.routes if hasattr(route, "path")]


def _env_list(name: str, default: list[str]) -> list[str]:
    raw = os.getenv(name, "")
    values = [value.strip().rstrip("/") for value in raw.split(",") if value.strip()]
    return values or default


_tracking_db: TrackingDB | None = None
_warehouse_db: WarehouseDB | None = None
_camera_db: CameraDB | None = None
_device_db: DeviceDB | None = None
_stream_manager: StreamManager | None = None
_security_audit_db: SecurityAuditDB | None = None
_access_control_db: AccessControlDB | None = None
_catalog_db: CatalogDB | None = None
_catalog_yolo_detector: Detector | None = None
_catalog_yolo_detector_key: tuple[Any, ...] | None = None
_catalog_yolo_last_scan: dict[str, dict[str, Any]] = {}
_accounts_db: AccountsDB | None = None
_rate_limits: dict[tuple[str, str, int], int] = {}
_watchdog_task: asyncio.Task | None = None
_catalog_recognition_task: asyncio.Task | None = None
_catalog_run_lock: asyncio.Lock | None = None
_manual_stop_requested = False
_watchdog_last_start_attempt = 0.0
_dataset_builder_obj: DatasetBuilder | None = None
_dataset_training_obj: DatasetTrainingManager | None = None
_dataset_capture_jobs: dict[str, dict[str, Any]] = {}
_active_product_detectors: dict[tuple[str, str], Detector] = {}


def _dataset_builder() -> DatasetBuilder:
    global _dataset_builder_obj
    if _dataset_builder_obj is None:
        database = VisionDB(os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db")))
        _dataset_builder_obj = DatasetBuilder(ROOT, database)
    return _dataset_builder_obj


def _dataset_training() -> DatasetTrainingManager:
    global _dataset_training_obj
    if _dataset_training_obj is None:
        _dataset_training_obj = DatasetTrainingManager(_dataset_builder())
    return _dataset_training_obj

# On-demand "run recognition for a few minutes" mode (POST
# /api/catalog/recognition/run-live): in-memory only, keyed by scope_id.
# Deliberately not persisted - if the server restarts mid-run the run is
# simply gone, which is fine for an ephemeral progress indicator.
_live_catalog_runs: dict[str, dict[str, Any]] = {}
_live_catalog_tasks: dict[str, asyncio.Task] = {}
_product_learning_sessions: dict[str, dict[str, Any]] = {}
_product_learning_tasks: dict[str, asyncio.Task] = {}
CATALOG_LIVE_RUN_DURATION_SECONDS = 60
CATALOG_LIVE_RUN_SAMPLE_INTERVAL_SECONDS = 8

ROLE_PERMISSIONS: dict[str, set[str]] = {
    "super_admin": {
        "view_dashboard",
        "view_organizations",
        "manage_organizations",
        "view_users",
        "manage_users",
        "view_permissions",
        "manage_permissions",
        "view_controllers",
        "configure_cameras",
        "view_cameras",
        "view_live_monitoring",
        "view_products",
        "manage_products",
        "configure_ai",
        "view_counts",
        "correct_counts",
        "view_alerts",
        "manage_alerts",
        "view_analytics",
        "view_reports",
        "export_reports",
        "view_system_health",
        "configure_system",
        "view_audit_logs",
        "manage_integrations",
        "view_settings",
    },
    "company_admin": {
        "view_dashboard",
        "view_users",
        "manage_users",
        "view_permissions",
        "view_controllers",
        "configure_cameras",
        "view_cameras",
        "view_live_monitoring",
        "view_products",
        "manage_products",
        "view_counts",
        "correct_counts",
        "view_alerts",
        "manage_alerts",
        "view_analytics",
        "view_reports",
        "export_reports",
        "view_system_health",
        "view_audit_logs",
        "view_settings",
    },
    "factory_manager": {
        "view_dashboard",
        "view_cameras",
        "view_live_monitoring",
        "view_products",
        "view_counts",
        "correct_counts",
        "view_alerts",
        "view_analytics",
        "view_reports",
        "export_reports",
        "view_system_health",
    },
    "warehouse_manager": {
        "view_dashboard",
        "view_cameras",
        "view_live_monitoring",
        "view_products",
        "view_counts",
        "correct_counts",
        "view_alerts",
        "view_reports",
        "export_reports",
    },
    "operator": {
        "view_dashboard",
        "view_cameras",
        "view_live_monitoring",
        "view_counts",
        "correct_counts",
        "view_alerts",
        "view_reports",
    },
    "viewer": {
        "view_dashboard",
        "view_cameras",
        "view_live_monitoring",
        "view_counts",
        "view_alerts",
        "view_reports",
    },
    "technician": {
        "view_dashboard",
        "view_controllers",
        "configure_cameras",
        "view_cameras",
        "view_live_monitoring",
        "view_system_health",
        "view_settings",
    },
}

DASHBOARD_V2_MODULES: dict[str, list[dict[str, str]]] = {
    "head": [
        {"id": "overview", "label": "Dashboard Overview", "permission": "view_dashboard"},
        {"id": "organizations", "label": "Organizations", "permission": "view_organizations"},
        {"id": "users", "label": "Users & Roles", "permission": "view_users"},
        {"id": "permissions", "label": "Permissions", "permission": "view_permissions"},
        {"id": "controllers", "label": "Controllers / NVR", "permission": "view_controllers"},
        {"id": "cameras", "label": "Cameras", "permission": "view_cameras"},
        {"id": "live", "label": "Live Monitoring", "permission": "view_live_monitoring"},
        {"id": "products", "label": "Products", "permission": "view_products"},
        {"id": "ai", "label": "AI Management", "permission": "configure_ai"},
        {"id": "counting", "label": "Counting Management", "permission": "view_counts"},
        {"id": "alerts", "label": "Alerts Center", "permission": "view_alerts"},
        {"id": "analytics", "label": "Analytics", "permission": "view_analytics"},
        {"id": "reports", "label": "Reports", "permission": "view_reports"},
        {"id": "health", "label": "System Health", "permission": "view_system_health"},
        {"id": "audit", "label": "Audit Logs", "permission": "view_audit_logs"},
        {"id": "integrations", "label": "Integrations", "permission": "manage_integrations"},
        {"id": "settings", "label": "Settings", "permission": "view_settings"},
    ],
    "user": [
        {"id": "home", "label": "Home", "permission": "view_dashboard"},
        {"id": "live", "label": "Live Monitoring", "permission": "view_live_monitoring"},
        {"id": "counting", "label": "Counting", "permission": "view_counts"},
        {"id": "shift", "label": "Current Shift", "permission": "view_counts"},
        {"id": "verification", "label": "Verification Tasks", "permission": "correct_counts"},
        {"id": "alerts", "label": "Alerts", "permission": "view_alerts"},
        {"id": "reports", "label": "Reports", "permission": "view_reports"},
        {"id": "activity", "label": "Activity History", "permission": "view_reports"},
        {"id": "profile", "label": "Profile", "permission": "view_dashboard"},
    ],
}


def _env_bool(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() not in {"0", "false", "no", "off"}


def _get_tracking_db() -> TrackingDB:
    global _tracking_db
    if _tracking_db is None:
        _tracking_db = TrackingDB(db_path=str(TRACKING_DB_PATH))
    return _tracking_db


def _get_warehouse_db() -> WarehouseDB:
    global _warehouse_db
    if _warehouse_db is None:
        _warehouse_db = WarehouseDB(db_path=str(WAREHOUSE_DB_PATH))
    return _warehouse_db


def _get_catalog_db() -> CatalogDB:
    global _catalog_db
    if _catalog_db is None:
        _catalog_db = CatalogDB(db_path=str(CATALOG_DB_PATH))
    return _catalog_db


def _get_accounts_db() -> AccountsDB:
    global _accounts_db
    if _accounts_db is None:
        _accounts_db = AccountsDB(db_path=str(ACCOUNTS_DB_PATH))
    return _accounts_db


def _get_camera_db() -> CameraDB:
    global _camera_db
    if _camera_db is None:
        _camera_db = CameraDB(db_path=str(CAMERA_DB_PATH))
        _seed_cameras_from_environment(_camera_db)
        config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
        first_camera = (config.get("cameras") or [{"name": "Camera 1", "source": 0}])[0]
        _camera_db.ensure_default_camera(
            name=str(first_camera.get("name", "Camera 1")),
            stream_url=str(first_camera.get("source", 0)),
        )
    return _camera_db


def _get_device_db() -> DeviceDB:
    global _device_db
    if _device_db is None:
        _device_db = DeviceDB(db_path=str(DEVICE_DB_PATH))
    return _device_db


def _get_stream_manager() -> StreamManager:
    global _stream_manager
    if _stream_manager is None:
        _stream_manager = StreamManager(snapshot_dir=SNAPSHOT_DIR)
    return _stream_manager


def _ensure_streams_from_active_cameras() -> dict[str, Any]:
    db = _get_camera_db()
    if db is None or not hasattr(db, "list_active_cameras"):
        return {"streams": []}
    active = db.list_active_cameras(include_secret=True)
    return _get_stream_manager().ensure_from_cameras(active)


def _start_stream_for_camera(camera: dict[str, Any]) -> dict[str, Any]:
    return _get_stream_manager().start(
        StreamSessionConfig(
            channel_id=str(camera["id"]),
            name=str(camera["name"]),
            source=str(camera["stream_url"]),
            slot_number=camera.get("slot_number"),
            snapshot_dir=SNAPSHOT_DIR,
            width=_bounded_stream_int("STREAM_FRAME_WIDTH", 1280, 240, 1280),
            jpeg_quality=_bounded_stream_int("STREAM_JPEG_QUALITY", 85, 20, 90),
            preview_fps=_bounded_stream_float("STREAM_PREVIEW_FPS", 12.0, 0.5, 30.0),
        )
    )


def _bounded_stream_int(name: str, default: int, minimum: int, maximum: int) -> int:
    try:
        value = int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(value, maximum))


def _bounded_stream_float(name: str, default: float, minimum: float, maximum: float) -> float:
    try:
        value = float(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(value, maximum))


def _seed_cameras_from_environment(db: CameraDB) -> None:
    """Optional boot-time camera seeding for stateless cloud deployments.

    DigitalOcean App Platform files can reset on rebuild. These env vars let the
    backend recreate controller channels on startup so the dashboard does not
    fall back to only the checked-in demo camera.
    """

    host = os.getenv("CAMERA_CONTROLLER_HOST", "").strip()
    if not host:
        return

    protocol = os.getenv("CAMERA_CONTROLLER_PROTOCOL", "rtsp").strip().lower()
    if protocol not in STREAM_DEFAULT_PORTS:
        protocol = "rtsp"

    try:
        port = int(os.getenv("CAMERA_CONTROLLER_PORT", str(STREAM_DEFAULT_PORTS[protocol])))
        channel_count = int(os.getenv("CAMERA_CONTROLLER_CHANNEL_COUNT", "10"))
        channel_start = int(os.getenv("CAMERA_CONTROLLER_CHANNEL_START", "1"))
        start_slot = int(os.getenv("CAMERA_CONTROLLER_START_SLOT", "1"))
        # Optional explicit, non-contiguous channel list, e.g. "2,5,6,16,20,23,25".
        channels_env = os.getenv("CAMERA_CONTROLLER_CHANNELS", "").strip()
        channels = (
            [int(part) for part in re.split(r"[,\s]+", channels_env) if part.strip()]
            if channels_env
            else None
        )
    except ValueError:
        return

    legacy_channels = {2, 5, 6, 16, 20, 23, 25}
    expand_legacy_seed = (
        os.getenv("CAMERA_CONTROLLER_EXPAND_ALL", "true").strip().lower()
        in {"1", "true", "yes", "on"}
        and channels is not None
        and set(channels) == legacy_channels
    )
    existing_active = db.list_active_cameras(include_secret=False)
    only_dummy = (
        len(existing_active) == 1
        and str(existing_active[0].get("masked_stream_url", "")).strip().lower() == "dummy"
    )
    if existing_active and not only_dummy and not expand_legacy_seed:
        return

    stream_template = os.getenv(
        "CAMERA_CONTROLLER_STREAM_TEMPLATE",
        "/Streaming/Channels/{channel}01",
    )
    if expand_legacy_seed:
        channels = list(range(1, 27))
        stream_template = re.sub(r"01$", "02", stream_template)
        os.environ.setdefault("STREAM_FRAME_WIDTH", "640")
        os.environ.setdefault("STREAM_PREVIEW_FPS", "6")
        os.environ.setdefault("STREAM_JPEG_QUALITY", "70")
        for camera in db.list_cameras(include_secret=True):
            stream_url = str(camera.get("stream_url") or "")
            endpoint = urlsplit(stream_url)
            if endpoint.hostname == host and re.search(r"/Streaming/Channels/\d+01$", endpoint.path):
                db.delete_camera(int(camera["id"]))

    controller = CameraControllerCreate(
        name=os.getenv("CAMERA_CONTROLLER_NAME", "Warehouse NVR Substream"),
        host=host,
        protocol=protocol,
        port=port,
        username=os.getenv("CAMERA_CONTROLLER_USERNAME") or None,
        password=os.getenv("CAMERA_CONTROLLER_PASSWORD") or None,
        channel_count=max(1, min(len(channels) if channels else channel_count, MAX_CAMERA_SLOTS)),
        channel_start=max(1, channel_start),
        channels=channels,
        start_slot=max(1, min(start_slot, MAX_CAMERA_SLOTS)),
        stream_path_template=stream_template,
        camera_name_template=os.getenv(
            "CAMERA_CONTROLLER_CAMERA_NAME_TEMPLATE",
            "{controller} Camera {channel}",
        ),
        make_active=True,
        # Unlike the dashboard's Add NVR flow, this seed path used to skip
        # every connectivity check and activate channels unconditionally -
        # stale or wrong credentials baked into old environment variables
        # could silently occupy real slots forever without ever actually
        # working, since nothing ever tested them.
        test_controller=True,
        test_streams=True,
        require_public=False,
    )

    try:
        _register_controller_channels(controller, db)
    except HTTPException as exc:
        print(f"WARNING: environment camera seed skipped: {exc.detail}")

    _sync_config_active_cameras(db)


def _get_security_audit_db() -> SecurityAuditDB:
    global _security_audit_db
    if _security_audit_db is None:
        _security_audit_db = SecurityAuditDB(db_path=str(SECURITY_AUDIT_DB_PATH))
    return _security_audit_db


def _get_access_control_db() -> AccessControlDB:
    global _access_control_db
    if _access_control_db is None:
        _access_control_db = AccessControlDB(db_path=str(ACCESS_CONTROL_DB_PATH))
    return _access_control_db


def _admin_api_key() -> str:
    return os.getenv("ADMIN_API_KEY", "").strip()


def _security_enabled() -> bool:
    return bool(_admin_api_key())


def _request_actor(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip()
    return request.client.host if request.client else "unknown"


def _v2_user_email(request: Request) -> str:
    user = _v2_session_user(request)
    if user:
        return str(user["email"]).strip().lower()
    return (
        request.query_params.get("user_email")
        or request.headers.get("x-ai-user-email")
        or "admin@ai-vision.local"
    ).strip().lower()


def _v2_bearer_token(request: Request) -> str:
    authorization = request.headers.get("authorization", "")
    if authorization.lower().startswith("bearer "):
        return authorization.split(" ", 1)[1].strip()
    return ""


def _v2_session_user(request: Request) -> dict[str, Any] | None:
    token = _v2_bearer_token(request)
    if not token:
        return None
    return _get_access_control_db().get_user_by_session_token(token)


def _v2_rp_id(request: Request) -> str:
    origin = request.headers.get("origin", "")
    if origin:
        host = urlsplit(origin).hostname
        if host:
            return host
    return request.url.hostname or "localhost"


def _v2_expected_origins(request: Request) -> list[str]:
    current = f"{request.url.scheme}://{request.url.netloc}"
    browser_origin = request.headers.get("origin", "").rstrip("/")
    configured = _env_list("WEBAUTHN_ALLOWED_ORIGINS", DEFAULT_ALLOWED_ORIGINS)
    origins = {current, browser_origin, *configured, "http://localhost:8000", "http://127.0.0.1:8000"}
    return sorted(origin.rstrip("/") for origin in origins if origin)


def _v2_public_key_options(options: Any) -> dict[str, Any]:
    return json.loads(options_to_json(options))


def _v2_dashboard(request: Request) -> dict[str, Any]:
    ac = _get_access_control_db()
    session_user = _v2_session_user(request)
    if session_user:
        return ac.resolve_dashboard(user_id=int(session_user["id"]))
    email = (
        request.query_params.get("user_email")
        or request.headers.get("x-ai-user-email")
        or "admin@ai-vision.local"
    ).strip().lower()
    user = ac.get_user_by_email(email)
    if user and user.get("has_password"):
        raise HTTPException(status_code=401, detail="Login required for this account.")
    return ac.resolve_dashboard(email=email)


def _v2_auth_response(user: dict[str, Any], token: dict[str, Any]) -> dict[str, Any]:
    dashboard = _get_access_control_db().resolve_dashboard(user_id=int(user["id"]))
    return {"user": dashboard["user"], "modules": dashboard["modules"], **token}



def _v2_require_permission(request: Request, permission: str) -> dict[str, Any]:
    dashboard = _v2_dashboard(request)
    if permission not in set(dashboard.get("permissions", [])):
        raise HTTPException(status_code=403, detail=f"Permission required: {permission}")
    return dashboard


def _v2_require_module(request: Request, module_code: str, permission: str | None = None) -> dict[str, Any]:
    dashboard = _v2_dashboard(request)
    modules = {module["code"] for module in dashboard.get("modules", [])}
    if module_code not in modules:
        raise HTTPException(status_code=403, detail=f"Module not assigned: {module_code}")
    if permission and permission not in set(dashboard.get("permissions", [])):
        raise HTTPException(status_code=403, detail=f"Permission required: {permission}")
    return dashboard


def _is_public_path(path: str) -> bool:
    return (
        path == "/"
        or path == "/api/status"
        or path.startswith("/assets/")
        or path in {"/favicon.ico", "/robots.txt"}
    )


def _valid_api_key(request: Request) -> bool:
    expected = _admin_api_key()
    if not expected:
        return True
    provided = request.headers.get("x-api-key") or request.query_params.get("api_key") or ""
    return secrets.compare_digest(provided, expected)


def _normalize_role(role: str | None) -> str:
    value = (role or "super_admin").strip().lower().replace(" ", "_").replace("-", "_")
    return value if value in ROLE_PERMISSIONS else "viewer"


def _parse_csv_header(value: str | None) -> set[str]:
    if not value:
        return set()
    return {item.strip() for item in value.split(",") if item.strip()}


def _rbac_context(request: Request) -> dict[str, Any]:
    role = _normalize_role(request.headers.get("x-ai-role"))
    base_permissions = set(ROLE_PERMISSIONS.get(role, set()))
    explicit_permissions = _parse_csv_header(request.headers.get("x-ai-permissions"))
    denied_permissions = _parse_csv_header(request.headers.get("x-ai-deny-permissions"))
    permissions = sorted((base_permissions | explicit_permissions) - denied_permissions)
    return {
        "user": {
            "id": request.headers.get("x-ai-user-id", "demo-super-admin"),
            "name": request.headers.get("x-ai-user-name", "Demo Super Admin"),
            "email": request.headers.get("x-ai-user-email", "admin@ai-vision.local"),
        },
        "role": role,
        "role_label": role.replace("_", " ").title(),
        "scope": {
            "company": request.headers.get("x-ai-company", "All Companies"),
            "factory": request.headers.get("x-ai-factory", "All Factories"),
            "warehouse": request.headers.get("x-ai-warehouse", "All Warehouses"),
            "production_line": request.headers.get("x-ai-production-line", "All Lines"),
            "camera": request.headers.get("x-ai-camera", "All Cameras"),
        },
        "permissions": permissions,
    }


def _authorized_modules(surface: str, permissions: set[str]) -> list[dict[str, str]]:
    modules = DASHBOARD_V2_MODULES.get(surface, [])
    return [module for module in modules if module["permission"] in permissions]


def _require_permission(request: Request, permission: str) -> dict[str, Any]:
    context = _rbac_context(request)
    if permission not in set(context["permissions"]):
        raise HTTPException(
            status_code=403,
            detail=f"Permission required: {permission}",
        )
    return context


def _rate_limit(request: Request) -> JSONResponse | None:
    rate_path = request.url.path
    if rate_path == "/api/live_frame":
        # Live snapshots are intentionally requested more frequently than the
        # control API. This remains isolated per camera below, so one feed
        # cannot consume another feed's allowance.
        limit = int(os.getenv("LIVE_FRAME_RATE_LIMIT_PER_MINUTE", "300"))
    else:
        limit = int(os.getenv("SECURITY_RATE_LIMIT_PER_MINUTE", "120"))
    if limit <= 0:
        return None
    actor = _request_actor(request)
    window = int(time.time() // 60)
    # Live camera frames are polled independently per slot. Keep the existing
    # per-route limit, but isolate each feed so ten healthy cameras do not
    # exhaust one shared request bucket.
    if rate_path == "/api/live_frame":
        slot = request.query_params.get("slot", "")
        camera = request.query_params.get("camera", "")
        rate_path = f"{rate_path}?slot={slot}&camera={camera}"
    key = (actor, rate_path, window)
    _rate_limits[key] = _rate_limits.get(key, 0) + 1
    if len(_rate_limits) > 5000:
        stale_windows = {window - 2, window - 1, window}
        for old_key in list(_rate_limits):
            if old_key[2] not in stale_windows:
                _rate_limits.pop(old_key, None)
    if _rate_limits[key] > limit:
        return JSONResponse(
            status_code=429,
            content={"detail": "Rate limit exceeded. Try again shortly."},
            headers={"X-RateLimit-Limit": str(limit), "X-RateLimit-Remaining": "0"},
        )
    return None


def _audit(action: str, payload: dict[str, Any], actor: str = "system") -> None:
    try:
        _get_security_audit_db().append(actor=actor, action=action, payload=payload)
    except Exception:
        # Audit logging should never take the control API down.
        pass


@app.middleware("http")
async def security_middleware(request: Request, call_next):
    path = request.url.path
    if path.startswith("/api/") or path.startswith("/snapshots/"):
        limited = _rate_limit(request)
        if limited is not None:
            return limited

    if _security_enabled() and not _is_public_path(path):
        if request.method != "OPTIONS" and (path.startswith("/api/") or path.startswith("/snapshots/")):
            if not _valid_api_key(request):
                actor = _request_actor(request)
                _audit(
                    "auth.denied",
                    {"method": request.method, "path": path},
                    actor=actor,
                )
                return JSONResponse(status_code=401, content={"detail": "Valid API key required."})

    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store"
    elif path.startswith("/dashboard-v2/"):
        # StaticFiles never sets an explicit Cache-Control header for
        # dashboard-v2/assets/*, which lets the Vercel edge (proxying via
        # rewrites) apply its own default caching for asset-looking paths —
        # independent of the ?v= query-string bump. That let a stale
        # pre-migration app.js keep running in production and silently
        # create client-only "accounts" that never reached the server. Force
        # revalidation on every dashboard-v2 asset so a fresh deploy is
        # always picked up.
        response.headers["Cache-Control"] = "no-cache"

    if path.startswith("/api/") and request.method not in {"GET", "HEAD", "OPTIONS"}:
        actor = _request_actor(request)
        _audit(
            "api.mutation",
            {"method": request.method, "path": path, "status_code": response.status_code},
            actor=actor,
        )
    return response


# Registered after security_middleware (not alongside the other
# app.add_middleware calls near the top of the file) so that CORSMiddleware
# ends up as the OUTERMOST layer — Starlette wraps in the order middleware is
# added, last-added-wraps-outermost. With CORS registered before
# security_middleware, an early 401 returned by security_middleware (e.g.
# when ADMIN_API_KEY is set and no key is supplied) never reached
# CORSMiddleware at all, so the response carried no
# Access-Control-Allow-Origin header — the browser reported it as a CORS
# failure instead of surfacing the real 401, which would silently break
# every cross-origin caller (including the Vercel-hosted dashboard) the
# moment API-key auth was turned on.
app.add_middleware(
    CORSMiddleware,
    allow_origins=_env_list("ALLOWED_ORIGINS", DEFAULT_ALLOWED_ORIGINS),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=[
        "Authorization",
        "Content-Type",
        "X-API-Key",
        "X-AI-Company",
        "X-AI-Role",
        "X-AI-User-Email",
        "X-AI-User-Name",
        "X-Requested-With",
    ],
)

_process: subprocess.Popen | None = None
_started_at: float | None = None
_last_exit_code: int | None = None
_stdout_handle = None
_stderr_handle = None


class StartRequest(BaseModel):
    no_display: bool = True
    config_path: str = "config/config.yaml"


class ConfigPatch(BaseModel):
    model_path: str | None = Field(default=None, min_length=1)
    fallback_model_path: str | None = Field(default=None, min_length=1)
    confidence_threshold: float | None = Field(default=None, ge=0.0, le=1.0)
    iou_threshold: float | None = Field(default=None, ge=0.0, le=1.0)
    max_detections: int | None = Field(default=None, ge=1, le=3000)
    image_size: int | None = Field(default=None, ge=320, le=1920)
    device: str | None = None
    target_fps: float | None = Field(default=None, ge=0.1, le=60.0)
    stale_after_ms: int | None = Field(default=None, ge=250, le=60000)
    max_concurrent_cameras: int | None = Field(default=None, ge=0, le=MAX_CAMERA_SLOTS)
    classes: list[str] | None = None
    class_prompts: list[str] | None = None
    class_agnostic_nms: bool | None = None
    show_fps: bool | None = None
    live_feed_enabled: bool | None = None
    live_frame_width: int | None = Field(default=None, ge=160, le=3840)
    live_frame_jpeg_quality: int | None = Field(default=None, ge=30, le=100)
    spatial_enabled: bool | None = None
    horizontal_fov_degrees: float | None = Field(default=None, ge=20.0, le=160.0)
    camera_height_m: float | None = Field(default=None, ge=0.1, le=20.0)
    horizon_y_ratio: float | None = Field(default=None, ge=0.0, le=1.0)
    min_distance_m: float | None = Field(default=None, ge=0.1, le=100.0)
    max_distance_m: float | None = Field(default=None, ge=0.1, le=500.0)
    estimate_depth_layers: bool | None = None
    max_units_per_detection: int | None = Field(default=None, ge=1, le=10000)
    tracking_enabled: bool | None = None
    tracking_grace_period_seconds: float | None = Field(default=None, ge=0.0, le=600.0)
    warehouse_counting_enabled: bool | None = None
    warehouse_confidence_threshold: float | None = Field(default=None, ge=0.0, le=1.0)
    count_low_confidence_as_unknown: bool | None = None
    snapshots_enabled: bool | None = None
    snapshot_trigger_classes: list[str] | None = None
    snapshot_cooldown_seconds: int | None = Field(default=None, ge=0)
    logging_enabled: bool | None = None
    recognition_enabled: bool | None = None
    recognition_provider: str | None = Field(default=None, min_length=1)
    recognition_model: str | None = Field(default=None, min_length=1)
    recognition_confidence_threshold: float | None = Field(default=None, ge=0.0, le=1.0)
    recognition_similarity_threshold: float | None = Field(default=None, ge=0.0, le=1.0)
    recognition_cache_enabled: bool | None = None
    recognition_cache_expiration: int | None = Field(default=None, ge=0)
    recognition_timeout: int | None = Field(default=None, ge=1, le=300)
    recognition_retries: int | None = Field(default=None, ge=0, le=10)
    recognition_max_workers: int | None = Field(default=None, ge=1, le=32)
    recognition_catalog_only: bool | None = None


class CompanyCreate(BaseModel):
    name: str = Field(min_length=1, max_length=120)


class CompanyRename(BaseModel):
    name: str = Field(min_length=1, max_length=120)


class CameraConfigUpdate(BaseModel):
    cameraConfig: dict[str, Any]  # noqa: N815 - matches the JS wire shape


class RoleCreate(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    login: str = Field(min_length=1, max_length=80)
    password: str = Field(min_length=1, max_length=200)
    access_camera: bool = False
    access_analytics: bool = False


class RoleUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=80)
    login: str | None = Field(default=None, min_length=1, max_length=80)
    password: str | None = Field(default=None, min_length=1, max_length=200)
    access_camera: bool | None = None
    access_analytics: bool | None = None


class ProfileUpdate(BaseModel):
    login: str | None = Field(default=None, min_length=1, max_length=80)
    password: str | None = Field(default=None, min_length=1, max_length=200)
    avatar: str | None = None
    remove_avatar: bool = False


class ItemCreate(BaseModel):
    item_id: str
    name: str
    item_type: str | None = None


class InventoryAction(BaseModel):
    item_id: str
    quantity: int = Field(default=1, ge=1)
    note: str | None = None


class CatalogPromptUpdate(BaseModel):
    prompts: list[str] = Field(default_factory=list, max_length=20)


class CatalogCorrection(BaseModel):
    correct_name: str = Field(min_length=1, max_length=60)
    prompt: str | None = Field(default=None, max_length=500)
    crop_url: str = Field(min_length=1, max_length=500)
    predicted_name: str | None = Field(default=None, max_length=120)


class ProductLearningStart(BaseModel):
    duration_seconds: int = Field(default=12, ge=10, le=20)
    camera_name: str = Field(min_length=1, max_length=200)


class ProductLearningSave(BaseModel):
    session_id: str = Field(min_length=8, max_length=120)
    product_name: str = Field(min_length=1, max_length=60)
    view_indices: list[int] = Field(min_length=1, max_length=8)
    existing_item_id: str | None = Field(default=None, max_length=120)


class WarehouseTaskRequest(BaseModel):
    prompt: str = Field(min_length=3, max_length=500)


class CameraCreate(BaseModel):
    name: str = Field(min_length=1)
    stream_url: str = Field(min_length=1)
    make_active: bool = True
    test_connection: bool = True
    slot_number: int | None = Field(default=None, ge=1, le=MAX_CAMERA_SLOTS)


class CameraTestRequest(BaseModel):
    stream_url: str = Field(min_length=1)


class CameraOperationsUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=200)
    stream_url: str | None = Field(default=None, min_length=1)
    block_id: int | None = None
    block_name: str | None = Field(default=None, max_length=200)


class CameraSlotRequest(BaseModel):
    slot_number: int = Field(default=1, ge=1, le=MAX_CAMERA_SLOTS)


class CameraCleanupRequest(BaseModel):
    name_prefix: str = Field(min_length=1)


class DiscoveryScanRequest(BaseModel):
    host: str = Field(min_length=1, max_length=255)


class DiscoveryConnectRequest(BaseModel):
    host: str = Field(min_length=1, max_length=255)
    protocol: str = Field(default="rtsp", pattern="^(rtsp|http|https)$")
    port: int | None = Field(default=None, ge=1, le=65535)
    username: str | None = None
    password: str | None = None
    vendor: str | None = None  # fingerprint hint carried from the scan step
    channel_count: int = Field(default=1, ge=1, le=MAX_CAMERA_SLOTS)
    name: str = Field(default="Camera", min_length=1, max_length=60)
    make_active: bool = True
    test_streams: bool = True


class V2DeviceDiscoverRequest(BaseModel):
    host: str = Field(min_length=1, max_length=255)
    name: str | None = Field(default=None, max_length=80)


class V2DeviceAuthenticateRequest(BaseModel):
    protocol: str = Field(default="rtsp", pattern="^(rtsp|http|https)$")
    port: int | None = Field(default=None, ge=1, le=65535)
    username: str | None = None
    password: str | None = None
    channel_count: int = Field(default=1, ge=1, le=MAX_CAMERA_SLOTS)
    make_active: bool = True
    test_streams: bool = False


def _v2_stream_vendor_hint(device: dict[str, Any], request: V2DeviceAuthenticateRequest) -> str | None:
    vendor = str(device.get("vendor") or "").strip().lower()
    # "generic-embedded" only means the web server banner looked embedded
    # (for example Boa/lighttpd). It is not a real RTSP path provider, and
    # choosing generic RTSP would save rtsp://host:554/ which most NVRs do not
    # use for video. Treat it as unknown so NVR/IP-camera RTSP falls through
    # to the safer Hikvision-style channel template below.
    if vendor and vendor not in {"generic-embedded", "generic-rtsp", "generic", "unknown"}:
        return vendor

    # Some public NVR forwards expose only RTSP/554 and do not return a brand
    # banner to the unauthenticated OPTIONS probe. The RTSP root almost never
    # carries usable video for those devices; the common Hikvision channel
    # profile is a better automatic first template while keeping vendor/path
    # choices out of the operator UI.
    if request.protocol.lower() == "rtsp":
        device_type = str(device.get("device_type") or "")
        if device_type in {"nvr_or_dvr", "ip_camera"} or request.channel_count > 1:
            return "hikvision"
    return None


class V2StreamStartRequest(BaseModel):
    slot_number: int | None = Field(default=None, ge=1, le=MAX_CAMERA_SLOTS)


class CameraControllerCreate(BaseModel):
    name: str = Field(default="Camera Controller", min_length=1)
    host: str = Field(min_length=1)
    protocol: str = Field(default="rtsp", pattern="^(rtsp|http|https)$")
    port: int | None = Field(default=None, ge=1, le=65535)
    username: str | None = None
    password: str | None = None
    channel_count: int = Field(default=4, ge=1, le=MAX_CAMERA_SLOTS)
    channel_start: int = Field(default=1, ge=1)
    # Optional explicit, possibly non-contiguous channel numbers (e.g. Block A/B
    # = 2,5,6,16,...). When set it overrides channel_start/channel_count so a
    # controller can register only the specific channels we want.
    channels: list[int] | None = Field(default=None)
    start_slot: int = Field(default=1, ge=1, le=MAX_CAMERA_SLOTS)
    stream_path_template: str = Field(default="/Streaming/Channels/{channel}01", min_length=1)
    camera_name_template: str = Field(default="{controller} Camera {channel}", min_length=1)
    make_active: bool = True
    test_controller: bool = True
    test_streams: bool = False
    require_public: bool = True


class V2UserCreate(BaseModel):
    name: str = Field(min_length=1)
    email: str = Field(min_length=3)


class V2LoginRequest(BaseModel):
    email: str = Field(min_length=3)
    password: str = Field(min_length=1)


class V2PasswordSet(BaseModel):
    password: str = Field(min_length=8)


class V2AuthPreferenceSet(BaseModel):
    preferred_auth_method: str = Field(pattern="^(biometric_first|password_first|password_and_biometric)$")


class V2PasskeyRegisterStart(BaseModel):
    name: str | None = None


class V2PasskeyRegisterFinish(BaseModel):
    challenge_id: int
    credential: dict[str, Any]
    name: str | None = None


class V2PasskeyLoginFinish(BaseModel):
    email: str = Field(min_length=3)
    challenge_id: int
    credential: dict[str, Any]


class V2PasskeyLoginStart(BaseModel):
    email: str = Field(min_length=3)


class V2RoleCreate(BaseModel):
    name: str = Field(min_length=1)
    code: str = Field(min_length=1)


class V2RoleAssignment(BaseModel):
    role_code: str = Field(min_length=1)


class V2ModuleAssignment(BaseModel):
    module_code: str = Field(min_length=1)
    effect: str = Field(pattern="^(allow|deny)$")
    display_order: int | None = Field(default=None, ge=1)


class V2PermissionAssignment(BaseModel):
    permission_code: str = Field(min_length=1)
    effect: str = Field(pattern="^(allow|deny)$")


class V2ScopeAssignment(BaseModel):
    scope_type: str = Field(pattern="^(company|factory|warehouse|production_line|zone|camera)$")
    scope_ids: list[str] = Field(default_factory=list)
    effect: str = Field(default="allow", pattern="^(allow|deny)$")


STREAM_DEFAULT_PORTS = {
    "rtsp": 554,
    "http": 80,
    "https": 443,
}

SECRET_URL_RE = re.compile(r"\b(?P<scheme>rtsp|https?)://(?P<username>[^:/\s]+):(?P<password>[^@\s]+)@")


def _read_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _write_yaml(path: Path, data: dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, sort_keys=False)


def _camera_source_from_text(stream_url: str):
    value = stream_url.strip()
    if value.isdigit():
        return int(value)
    return value


def _redact_sensitive_text(text: str) -> str:
    return SECRET_URL_RE.sub(
        lambda match: f"{match.group('scheme')}://{match.group('username')}:****@",
        text,
    )


def _is_local_capture_source(value: str) -> bool:
    if value.isdigit() or value.lower() == "dummy":
        return True
    try:
        return Path(value).exists()
    except (OSError, ValueError):
        return False


def _camera_stream_endpoint(stream_url: str) -> tuple[dict[str, Any] | None, str | None]:
    value = stream_url.strip()
    if _is_local_capture_source(value):
        return None, None

    try:
        parsed = urlsplit(value)
    except ValueError as exc:
        return None, f"Invalid camera stream URL: {exc}"

    scheme = parsed.scheme.lower()
    if scheme not in STREAM_DEFAULT_PORTS:
        return (
            None,
            "Use a full camera stream URL starting with rtsp://, http://, or https://, "
            "or use a local webcam index like 0.",
        )

    if not parsed.hostname:
        return None, "Camera stream URL is missing a host or IP address."

    try:
        port = parsed.port or STREAM_DEFAULT_PORTS[scheme]
    except ValueError as exc:
        return None, f"Invalid camera stream port: {exc}"

    return {"scheme": scheme, "host": parsed.hostname, "port": port}, None


def _normalize_controller_host(host: str) -> str:
    value = host.strip()
    if "://" in value:
        parsed = urlsplit(value)
        if parsed.hostname:
            return parsed.hostname
    return value.strip("/")


def _private_controller_host_message(host: str) -> str | None:
    try:
        address = ipaddress.ip_address(host)
    except ValueError:
        return None

    if not address.is_global:
        return (
            f"Controller host {host} is not publicly reachable from the internet. "
            "Use the controller/router public IP address or a DNS/DDNS hostname, and forward the "
            "RTSP/HTTP stream port to the controller. Private LAN addresses like 192.168.x.x, "
            "10.x.x.x, 172.16-31.x.x, and 127.x.x.x only work from the same local network."
        )
    return None


def _controller_endpoint(controller: CameraControllerCreate) -> dict[str, Any]:
    protocol = controller.protocol.lower()
    return {
        "scheme": protocol,
        "host": _normalize_controller_host(controller.host),
        "port": controller.port or STREAM_DEFAULT_PORTS[protocol],
    }


def _controller_stream_url(controller: CameraControllerCreate, channel: int) -> str:
    protocol = controller.protocol.lower()
    host = _normalize_controller_host(controller.host)
    port = controller.port or STREAM_DEFAULT_PORTS[protocol]
    path = controller.stream_path_template.format(channel=channel)
    if not path.startswith("/"):
        path = f"/{path}"

    credentials = ""
    if controller.username:
        credentials = quote(controller.username, safe="")
        if controller.password:
            credentials += f":{quote(controller.password, safe='')}"
        credentials += "@"

    return f"{protocol}://{credentials}{host}:{port}{path}"


def _controller_camera_name(controller: CameraControllerCreate, channel: int, slot: int) -> str:
    return controller.camera_name_template.format(
        controller=controller.name.strip(),
        channel=channel,
        slot=slot,
    )


def _set_config_active_cameras(cameras: list[dict[str, Any]]) -> dict[str, Any]:
    data = _read_yaml(CONFIG_PATH)
    data["cameras"] = [
        {
            "name": camera["name"],
            "source": _camera_source_from_text(camera["stream_url"]),
            "slot_number": camera.get("slot_number") or index,
        }
        for index, camera in enumerate(cameras, start=1)
    ]
    _write_yaml(CONFIG_PATH, data)
    return data


def _sync_config_active_cameras(db: CameraDB) -> dict[str, Any]:
    return _set_config_active_cameras(db.list_active_cameras(include_secret=True))


def _next_available_slot(cameras: list[dict[str, Any]]) -> int:
    used_slots = {
        int(camera["slot_number"])
        for camera in cameras
        if camera.get("is_active") and camera.get("slot_number") is not None
    }
    slot_number = 1
    while slot_number in used_slots:
        slot_number += 1
    return slot_number


def _activate_stream_managed_camera(
    db: CameraDB,
    camera_id: int,
    next_slot: int,
    used_slots: set[int],
    *,
    reuse_existing_slot: bool = True,
) -> tuple[dict[str, Any] | None, int | None, int]:
    current = db.get_camera(camera_id, include_secret=True)
    if (
        reuse_existing_slot
        and current
        and current.get("is_active")
        and current.get("slot_number") is not None
    ):
        assigned_slot = int(current["slot_number"])
        used_slots.add(assigned_slot)
        return current, assigned_slot, next_slot

    while next_slot in used_slots and next_slot <= MAX_CAMERA_SLOTS:
        next_slot += 1
    if next_slot > MAX_CAMERA_SLOTS:
        return None, None, next_slot

    active = db.assign_slot(camera_id, next_slot)
    used_slots.add(next_slot)
    assigned_slot = next_slot
    return active, assigned_slot, next_slot + 1


def _delete_duplicate_stream_url_cameras(db: CameraDB, stream_url: str, keep_id: int) -> None:
    for stale in db.list_cameras(include_secret=True):
        stale_id = int(stale["id"])
        if stale_id == keep_id or stale.get("stream_url") != stream_url:
            continue
        if stale.get("is_active"):
            _get_stream_manager().stop(str(stale_id))
        db.delete_camera(stale_id)


def _alternate_hikvision_stream_profile(stream_url: str) -> str:
    main = re.sub(
        r"(/Streaming/Channels/\d+)02(?=($|[/?#]))", r"\g<1>01", stream_url,
        count=1, flags=re.IGNORECASE,
    )
    if main != stream_url:
        return main
    return re.sub(
        r"(/Streaming/Channels/\d+)01(?=($|[/?#]))", r"\g<1>02", stream_url,
        count=1, flags=re.IGNORECASE,
    )


def _test_camera_stream(
    stream_url: str,
    timeout_seconds: int = 10,
    *,
    allow_profile_fallback: bool = True,
) -> dict[str, Any]:
    endpoint, validation_error = _camera_stream_endpoint(stream_url)
    if validation_error:
        return {"status": "failed", "message": validation_error}

    if stream_url.strip().lower() == "dummy":
        return {"status": "connected", "message": "Demo camera source is available."}

    probe_id = f"probe-{int(time.time() * 1000)}"
    manager = _get_stream_manager()
    manager.start(
        StreamSessionConfig(
            channel_id=probe_id,
            name="Camera probe",
            source=stream_url,
            snapshot_dir=SNAPSHOT_DIR,
        )
    )
    deadline = time.time() + max(1, int(timeout_seconds))
    status: dict[str, Any] = {"status": "starting"}
    try:
        while time.time() < deadline:
            status = manager.status(probe_id)
            frame = manager.latest_frame_bytes(channel_id=probe_id)
            if status.get("status") == "online" and frame:
                response = {
                    "status": "connected",
                    "message": "Stream Manager opened the camera stream and returned a frame.",
                }
                if endpoint is not None:
                    response["details"] = {
                        "host": endpoint["host"],
                        "port": endpoint["port"],
                        "scheme": endpoint["scheme"],
                        "stream_manager": True,
                        "frame_read": True,
                    }
                return response
            if status.get("last_error") and status.get("status") == "reconnecting":
                break
            time.sleep(0.1)
    finally:
        manager.stop(probe_id)

    last_error = status.get("last_error")
    if last_error:
        fallback_url = _alternate_hikvision_stream_profile(stream_url)
        if allow_profile_fallback and fallback_url != stream_url:
            fallback = _test_camera_stream(
                fallback_url,
                timeout_seconds,
                allow_profile_fallback=False,
            )
            if fallback.get("status") == "connected":
                fallback["message"] = (
                    "Stream Manager opened the camera through its alternate "
                    "Hikvision stream profile."
                )
                return fallback
        response = {
            "status": "failed",
            "message": _redact_sensitive_text(
                f"Stream Manager could not read the camera stream: {last_error}"
            ),
        }
        if endpoint is not None:
            response["details"] = {
                "host": endpoint["host"],
                "port": endpoint["port"],
                "scheme": endpoint["scheme"],
                "stream_manager": True,
                "frame_read": False,
            }
        return response

    response = {
        "status": "connected",
        "message": (
            "Stream Manager is connected or warming up; it will keep waiting "
            "for the first video keyframe."
        ),
    }
    if endpoint is not None:
        response["details"] = {
            "host": endpoint["host"],
            "port": endpoint["port"],
            "scheme": endpoint["scheme"],
            "stream_manager": True,
            "waiting_for_frame": True,
        }
    return response


def _load_inventory() -> dict[str, Any]:
    INVENTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not INVENTORY_PATH.exists():
        _save_inventory({"items": [], "history": []})

    with INVENTORY_PATH.open("r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {"items": [], "history": []}


def _save_inventory(data: dict[str, Any]) -> None:
    with INVENTORY_PATH.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def _now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())


def _ensure_inventory() -> dict[str, Any]:
    data = _load_inventory()
    data.setdefault("items", [])
    data.setdefault("history", [])
    return data


def _find_item(data: dict[str, Any], item_id: str) -> dict[str, Any] | None:
    return next((item for item in data["items"] if item["item_id"] == item_id), None)


def _record_inventory_event(data: dict[str, Any], action: str, item_id: str, quantity: int, note: str | None) -> None:
    data["history"].insert(0, {
        "timestamp": _now_iso(),
        "item_id": item_id,
        "action": action,
        "quantity": quantity,
        "note": note,
    })


def _catalog_interval_hours() -> int:
    return max(1, int(os.getenv("CATALOG_RECOGNITION_INTERVAL_HOURS", "12")))


def _catalog_scope(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip()).strip("-")
    if not cleaned:
        raise HTTPException(status_code=400, detail="A valid catalog scope is required.")
    return cleaned[:80]


def _catalog_safe_name(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9._-]+", "_", value).strip("._")
    return cleaned[:100] or "reference.jpg"


def _catalog_prompt_store() -> dict[str, Any]:
    data = _read_json(CATALOG_PROMPTS_PATH) or {}
    scopes = data.get("scopes")
    return {"scopes": scopes if isinstance(scopes, dict) else {}}


def _catalog_clean_prompts(values: list[str]) -> list[str]:
    prompts: list[str] = []
    seen: set[str] = set()
    for raw in values:
        prompt = " ".join(str(raw).split()).strip()[:80]
        normalized = _catalog_normalize_name(prompt)
        if prompt and normalized and normalized not in seen:
            seen.add(normalized)
            prompts.append(prompt)
    return prompts[:20]


def _catalog_item_prompts(scope_id: str, item_id: str) -> list[str]:
    scope_prompts = _catalog_prompt_store()["scopes"].get(scope_id) or {}
    values = scope_prompts.get(str(item_id)) if isinstance(scope_prompts, dict) else []
    return _catalog_clean_prompts(values if isinstance(values, list) else [])


def _catalog_save_item_prompts(
    scope_id: str, item_id: str, prompts: list[str]
) -> list[str]:
    data = _catalog_prompt_store()
    scopes = data["scopes"]
    scope_prompts = scopes.setdefault(scope_id, {})
    cleaned = _catalog_clean_prompts(prompts)
    if cleaned:
        scope_prompts[str(item_id)] = cleaned
    else:
        scope_prompts.pop(str(item_id), None)
    CATALOG_PROMPTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    temporary = CATALOG_PROMPTS_PATH.with_suffix(".tmp")
    temporary.write_text(json.dumps(data, indent=2), encoding="utf-8")
    temporary.replace(CATALOG_PROMPTS_PATH)
    return cleaned


def _catalog_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        try:
            parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        except ValueError:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _catalog_schedule(scope_id: str) -> dict[str, Any]:
    latest = _get_catalog_db().latest_run(scope_id)
    interval = _catalog_interval_hours()
    completed = _catalog_datetime(latest.get("completed_at")) if latest else None
    next_run = completed + timedelta(hours=interval) if completed else datetime.now(timezone.utc)
    return {
        "interval_hours": interval,
        "last_run_at": completed.isoformat() if completed else None,
        "next_run_at": next_run.isoformat(),
    }


def _catalog_dimensions(obj: dict[str, Any] | None) -> tuple[float, float, float] | None:
    if not obj:
        return None
    try:
        values = (float(obj["width_m"]), float(obj["height_m"]), float(obj["depth_m"]))
    except (KeyError, TypeError, ValueError):
        return None
    return values if all(value > 0 for value in values) else None


def _catalog_normalize_name(value: Any) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).split())


def _catalog_frame_embeddings(health: dict[str, Any]) -> dict[str, list[float]]:
    try:
        import cv2
        from recognition.embedding import image_embedding
    except ImportError:
        return {}

    embeddings: dict[str, list[float]] = {}
    for camera in health.get("cameras") or []:
        slot = camera.get("slot_number")
        name = str(camera.get("name") or f"slot-{slot}")
        if not slot:
            continue
        frame = _catalog_live_frame_image(slot=int(slot), camera=name)
        if frame is not None:
            embeddings[name] = image_embedding(frame)
    return embeddings


def _catalog_live_frame_image(slot: int | None = None, camera: str | None = None):
    """Decode the current frame owned by Stream Manager for one camera."""
    try:
        import cv2
        import numpy as np

        data = _get_stream_manager().latest_frame_bytes(
            slot_number=slot, name=camera
        )
        if data:
            frame = cv2.imdecode(np.frombuffer(data, dtype=np.uint8), cv2.IMREAD_COLOR)
            if frame is not None:
                return frame
    except Exception:
        pass

    try:
        import cv2
    except ImportError:
        return None
    for path in _live_feed_paths(slot=slot, camera=camera):
        if not path.exists():
            continue
        frame = cv2.imread(str(path))
        if frame is not None:
            return frame
    return None


def _catalog_health_snapshot() -> dict[str, Any]:
    health = _read_json(DETECTION_HEALTH_PATH) or {}
    if (
        health.get("cameras")
        or health.get("last_spatial_objects_by_camera")
        or health.get("last_spatial_objects")
    ):
        return health

    streams = (health.get("stream_manager") or {}).get("streams") or []
    if not streams:
        try:
            streams = _get_stream_manager().status().get("streams", [])
        except Exception:  # noqa: BLE001 - recognition can still return an empty run
            streams = []

    cameras = [
        {"name": str(stream.get("name") or f"slot-{stream.get('slot_number')}"), "slot_number": stream.get("slot_number")}
        for stream in streams
        if stream.get("slot_number") is not None
        and str(stream.get("status") or "").lower() not in {"offline", "stopped"}
    ]
    return {**health, "cameras": cameras}


def _catalog_crop_candidates(health: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        import cv2
        from recognition.embedding import image_embedding
    except ImportError:
        return []

    slots_by_camera = {
        str(camera.get("name") or f"slot-{camera.get('slot_number')}"): camera.get("slot_number")
        for camera in health.get("cameras") or []
    }
    candidates: list[dict[str, Any]] = []
    by_camera = health.get("last_detections_by_camera") or {}
    for camera_name, detections in by_camera.items():
        slot = slots_by_camera.get(str(camera_name))
        frame = _catalog_live_frame_image(
            slot=int(slot) if slot else None, camera=str(camera_name)
        )
        if frame is None:
            continue

        for detection in detections or []:
            crop = _catalog_detection_crop(frame, detection.get("bbox") or {})
            if crop is None:
                continue
            candidates.append(
                {
                    "camera_name": str(camera_name),
                    "detection": detection,
                    "frame": frame,
                    "crop": crop,
                    "embedding": image_embedding(crop),
                }
            )
    return candidates


def _catalog_detection_crop(frame, bbox: dict[str, Any]):
    try:
        x1 = int(float(bbox["x1"]))
        y1 = int(float(bbox["y1"]))
        x2 = int(float(bbox["x2"]))
        y2 = int(float(bbox["y2"]))
    except (KeyError, TypeError, ValueError):
        return None

    height, width = frame.shape[:2]
    box_width = max(1, x2 - x1)
    box_height = max(1, y2 - y1)
    pad_x = max(4, int(box_width * 0.08))
    pad_y = max(4, int(box_height * 0.08))
    x1 = max(0, x1 - pad_x)
    y1 = max(0, y1 - pad_y)
    x2 = min(width, x2 + pad_x)
    y2 = min(height, y2 + pad_y)
    if x2 <= x1 or y2 <= y1:
        return None
    return frame[y1:y2, x1:x2]


def _catalog_live_frames(health: dict[str, Any], max_frames: int | None = None) -> list[dict[str, Any]]:
    frames: list[dict[str, Any]] = []
    for camera in health.get("cameras") or []:
        if max_frames is not None and len(frames) >= max_frames:
            break
        slot = camera.get("slot_number")
        camera_name = str(camera.get("name") or f"slot-{slot}")
        frame = _catalog_live_frame_image(
            slot=int(slot) if slot else None, camera=camera_name
        )
        if frame is not None:
            frames.append({"camera_name": camera_name, "slot": slot, "frame": frame})
    return frames


def _catalog_detection_payload(detection) -> dict[str, Any]:
    x1, y1, x2, y2 = getattr(detection, "box", (0, 0, 0, 0))
    payload = {
        "class_name": getattr(detection, "class_name", "object"),
        "confidence": float(getattr(detection, "confidence", 0.0)),
        "quantity": max(1, int(getattr(detection, "quantity", 1))),
        "bbox": {"x1": float(x1), "y1": float(y1), "x2": float(x2), "y2": float(y2)},
    }
    for field in ("width_m", "height_m", "depth_m", "distance_m", "method", "object_type", "quantity_grid"):
        if hasattr(detection, field):
            value = getattr(detection, field)
            payload[field] = list(value) if field == "quantity_grid" else value
    return payload


def _catalog_snapshot_path(url: str) -> Path | None:
    """Resolve a /snapshots/... URL to a file under SNAPSHOT_DIR, safely.

    Guards against path traversal so a crafted crop_url cannot read files
    outside the snapshots directory.
    """
    prefix = "/snapshots/"
    if not url.startswith(prefix):
        return None
    relative = unquote(url[len(prefix):]).lstrip("/")
    if not relative:
        return None
    candidate = (SNAPSHOT_DIR / relative).resolve()
    root = SNAPSHOT_DIR.resolve()
    if candidate != root and not candidate.is_relative_to(root):
        return None
    return candidate


def _catalog_detection_prompts(
    items: list[dict[str, Any]], scope_id: str
) -> list[str]:
    prompts = [str(item["name"]) for item in items]
    for item in items:
        prompts.extend(_catalog_item_prompts(scope_id, str(item["id"])))
    # NOTE: no generic "box"/"carton"/"package" fallback prompts. They made the
    # detector hunt for any rectangular package, and every such box then matched
    # a specific catalog item like "Baget Box". Exact inventory recognition is
    # driven by the item's own name/prompts plus reference-image confirmation.
    seen: set[str] = set()
    unique = []
    for prompt in prompts:
        normalized = _catalog_normalize_name(prompt)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique.append(prompt)
    return unique


def _catalog_detection_matches_item_prompt(
    detection: dict[str, Any],
    item_name: str,
    aliases: list[str] | None = None,
) -> bool:
    targets = [
        _catalog_normalize_name(value)
        for value in [item_name, *(aliases or [])]
        if _catalog_normalize_name(value)
    ]
    labels = [
        _catalog_normalize_name(str(detection.get("inventory_name") or "")),
        _catalog_normalize_name(str(detection.get("class_name") or "")),
        _catalog_normalize_name(str(detection.get("object_type") or "")),
    ]
    labels = [label for label in labels if label]
    if any(
        label == target or label in target or target in label
        for target in targets
        for label in labels
    ):
        return True

    if any("box" in target for target in targets):
        return any(
            any(term in label for term in ("box", "carton", "package", "cardboard"))
            for label in labels
        )

    if any(
        any(term in target for term in ("sack", "bag"))
        for target in targets
    ):
        return any(any(term in label for term in ("sack", "bag")) for label in labels)

    return False


def _catalog_detection_confidence(detection: dict[str, Any], fallback: float = 0.9) -> float:
    try:
        confidence = float(detection.get("confidence"))
    except (TypeError, ValueError):
        return fallback
    return confidence if confidence > 0 else fallback


def _catalog_camera_label(value: Any) -> str:
    return str(value or "Camera").strip() or "Camera"


def _catalog_visual_slug(value: Any) -> str:
    slug = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "")).strip("._-")
    return slug[:80] or "visual"


def _catalog_read_camera_frame(health: dict[str, Any], camera_name: str):
    camera_label = _catalog_camera_label(camera_name)
    slots_by_camera = {
        _catalog_camera_label(camera.get("name") or f"slot-{camera.get('slot_number')}"): camera.get("slot_number")
        for camera in health.get("cameras") or []
    }
    slot = slots_by_camera.get(camera_label)
    return _catalog_live_frame_image(
        slot=int(slot) if slot else None, camera=camera_label
    )


def _catalog_spatial_visuals(
    health: dict[str, Any],
    entries: list[tuple[str, dict[str, Any]]],
) -> dict[str, dict[str, Any]]:
    visuals: dict[str, dict[str, Any]] = {}
    frames: dict[str, Any] = {}
    for camera_name, detection in entries:
        label = _catalog_camera_label(camera_name)
        if label not in frames:
            frames[label] = _catalog_read_camera_frame(health, label)
        frame = frames[label]
        crop = _catalog_detection_crop(frame, detection.get("bbox") or {}) if frame is not None else None
        visuals.setdefault(label, {"frame": frame, "crop": crop, "detection": detection})
    return visuals


def _catalog_candidate_visuals(candidates: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    visuals: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        label = _catalog_camera_label(candidate.get("camera_name"))
        visuals.setdefault(
            label,
            {
                "frame": candidate.get("frame"),
                "crop": candidate.get("crop"),
                "detection": candidate.get("detection") or {},
            },
        )
    return visuals


def _catalog_save_visual_image(
    scope_id: str,
    run_id: str,
    filename: str,
    image: Any,
) -> str | None:
    if image is None or getattr(image, "size", 0) == 0:
        return None
    try:
        import cv2
    except ImportError:
        return None

    scope_slug = _catalog_visual_slug(scope_id)
    run_slug = _catalog_visual_slug(run_id)
    directory = SNAPSHOT_DIR / "catalog-recognition" / scope_slug / run_slug
    directory.mkdir(parents=True, exist_ok=True)
    file_slug = _catalog_visual_slug(filename)
    path = directory / f"{file_slug}.jpg"
    if not cv2.imwrite(str(path), image):
        return None
    return f"/snapshots/catalog-recognition/{quote(scope_slug)}/{quote(run_slug)}/{quote(path.name)}"


def _catalog_persist_match_visuals(scope_id: str, run_id: str, match: dict[str, Any]) -> dict[str, Any]:
    evidence = match.get("_visual_evidence") or {}
    persisted = {key: value for key, value in match.items() if not key.startswith("_")}
    if not evidence:
        return persisted

    camera_counts = []
    item_slug = _catalog_visual_slug(persisted.get("item_name"))
    for entry in persisted.get("camera_counts") or []:
        enriched = dict(entry)
        enriched.setdefault("detected_at", _now_iso())
        camera_name = _catalog_camera_label(enriched.get("camera_name"))
        visual = evidence.get(camera_name)
        if visual:
            prefix = f"{item_slug}_{_catalog_visual_slug(camera_name)}"
            frame_url = _catalog_save_visual_image(scope_id, run_id, f"{prefix}_frame", visual.get("frame"))
            crop_url = _catalog_save_visual_image(scope_id, run_id, f"{prefix}_object", visual.get("crop"))
            detection = visual.get("detection") or {}
            if frame_url:
                enriched["frame_url"] = frame_url
            if crop_url:
                enriched["crop_url"] = crop_url
            if detection.get("bbox"):
                enriched["bbox"] = detection.get("bbox")
            if detection.get("class_name") or detection.get("object_type"):
                enriched["class_name"] = detection.get("class_name") or detection.get("object_type")
        camera_counts.append(enriched)
    persisted["camera_counts"] = camera_counts
    return persisted


def _catalog_camera_counts_payload(counts: dict[str, int]) -> list[dict[str, Any]]:
    return [
        {"camera_name": camera_name, "quantity": int(quantity)}
        for camera_name, quantity in sorted(counts.items())
        if int(quantity) > 0
    ]


def _catalog_count_objects_by_camera(
    entries: list[tuple[str, dict[str, Any]]],
) -> tuple[int, dict[str, int]]:
    counts: dict[str, int] = {}
    total = 0
    for camera_name, obj in entries:
        quantity = max(1, int(obj.get("quantity") or 1))
        label = _catalog_camera_label(camera_name)
        counts[label] = counts.get(label, 0) + quantity
        total += quantity
    return total, counts


def _catalog_recognition_sample_count() -> int:
    try:
        count = int(os.getenv("CATALOG_RECOGNITION_SAMPLES", "3"))
    except ValueError:
        count = 3
    return max(1, min(count, 8))


def _catalog_recognition_sample_interval_seconds() -> float:
    try:
        interval = float(os.getenv("CATALOG_RECOGNITION_SAMPLE_INTERVAL_SECONDS", "0.2"))
    except ValueError:
        interval = 0.2
    return max(0.0, min(interval, 2.0))


def _catalog_match_rank(match: dict[str, Any]) -> tuple[int, float]:
    quantity = max(0, int(match.get("quantity") or 0))
    confidence = max(0.0, min(1.0, float(match.get("confidence") or 0.0)))
    return quantity, confidence


def _catalog_merge_match_samples(samples: list[list[dict[str, Any]]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    camera_counts_by_item: dict[str, dict[str, int]] = {}
    evidence_by_item: dict[str, dict[str, dict[str, Any]]] = {}

    for sample in samples:
        for match in sample:
            item_id = str(match.get("item_id") or "")
            if not item_id:
                continue
            current = merged.get(item_id)
            if current is None or _catalog_match_rank(match) > _catalog_match_rank(current):
                merged[item_id] = {key: value for key, value in match.items() if not key.startswith("_")}
            item_counts = camera_counts_by_item.setdefault(item_id, {})
            item_evidence = evidence_by_item.setdefault(item_id, {})
            visual_evidence = match.get("_visual_evidence") or {}
            for entry in match.get("camera_counts") or []:
                quantity = max(0, int(entry.get("quantity") or 0))
                if quantity <= 0:
                    continue
                camera_name = _catalog_camera_label(entry.get("camera_name"))
                if quantity > item_counts.get(camera_name, 0):
                    item_counts[camera_name] = quantity
                    if visual_evidence.get(camera_name):
                        item_evidence[camera_name] = visual_evidence[camera_name]

    results: list[dict[str, Any]] = []
    for item_id, match in merged.items():
        item_counts = camera_counts_by_item.get(item_id, {})
        if item_counts:
            match["quantity"] = sum(item_counts.values())
            match["camera_counts"] = _catalog_camera_counts_payload(item_counts)
        if evidence_by_item.get(item_id):
            match["_visual_evidence"] = evidence_by_item[item_id]
        results.append(match)
    return sorted(results, key=lambda result: (-int(result.get("quantity") or 0), str(result.get("item_name") or "")))


def _catalog_sample_current_frame(scope_id: str, include_visuals: bool = False) -> list[dict[str, Any]]:
    sample_count = _catalog_recognition_sample_count()
    interval = _catalog_recognition_sample_interval_seconds()
    samples: list[list[dict[str, Any]]] = []
    for index in range(sample_count):
        samples.append(_catalog_match_current_frame(scope_id, include_visuals=include_visuals))
        if interval > 0 and index < sample_count - 1:
            time.sleep(interval)
    return _catalog_merge_match_samples(samples)


def _catalog_yolo_for_prompts(prompts: list[str]) -> Detector | None:
    global _catalog_yolo_detector, _catalog_yolo_detector_key
    if not prompts:
        return None

    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    det_cfg = config.get("detection", {})
    key = (
        det_cfg.get("model_path", "yolov8s-world.pt"),
        tuple(prompts),
        float(os.getenv("CATALOG_YOLO_CONFIDENCE_THRESHOLD", "0.01")),
        det_cfg.get("device", "cpu"),
        int(os.getenv("CATALOG_YOLO_IMAGE_SIZE", "1280")),
    )
    if _catalog_yolo_detector is not None and _catalog_yolo_detector_key == key:
        return _catalog_yolo_detector

    try:
        _catalog_yolo_detector = Detector(
            model_path=str(key[0]),
            confidence_threshold=float(key[2]),
            device=str(key[3]),
            class_prompts=prompts,
            image_size=int(key[4]),
            class_agnostic_nms=True,
        )
        _catalog_yolo_detector_key = key
        return _catalog_yolo_detector
    except Exception as exc:  # noqa: BLE001 - catalog recognition must degrade gracefully
        _audit("catalog_yolo_detector_failed", {"error": str(exc)})
        _catalog_yolo_detector = None
        _catalog_yolo_detector_key = None
        return None


def _catalog_box_iou(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> float:
    x1 = max(left[0], right[0])
    y1 = max(left[1], right[1])
    x2 = min(left[2], right[2])
    y2 = min(left[3], right[3])
    intersection = max(0, x2 - x1) * max(0, y2 - y1)
    if intersection <= 0:
        return 0.0
    left_area = max(1, (left[2] - left[0]) * (left[3] - left[1]))
    right_area = max(1, (right[2] - right[0]) * (right[3] - right[1]))
    return intersection / max(1, left_area + right_area - intersection)


# Product families used to reject cross-family candidates during matching.
# Box terms are checked first because "baget"/"baguette" contain the substring
# "bag", which would otherwise be misread as a sack.
_CATALOG_BOX_TERMS = ("box", "carton", "cardboard", "crate", "baget", "baguette", "package")
_CATALOG_SACK_TERMS = ("sack", "bag")


def _catalog_family(text: str) -> str | None:
    normalized = _catalog_normalize_name(text)
    if any(term in normalized for term in _CATALOG_BOX_TERMS):
        return "box"
    if any(term in normalized for term in _CATALOG_SACK_TERMS):
        return "sack"
    return None


def _catalog_item_category(item_name: str, aliases: list[str] | None = None) -> str | None:
    """Coarse product family ('box' or 'sack') for an enrolled item."""
    return _catalog_family(" ".join([str(item_name), *(str(a) for a in aliases or [])]))


def _catalog_detection_category(detection: dict[str, Any]) -> str | None:
    return _catalog_family(
        " ".join(
            str(detection.get(field) or "")
            for field in ("class_name", "object_type", "inventory_name")
        )
    )


def _catalog_category_conflicts(item_category: str | None, detection: dict[str, Any]) -> bool:
    """True when the candidate's family clearly differs from the enrolled item.

    A sack/bag detection can never be a box catalog item (and vice versa), so it
    is rejected before the visual check. Candidates with no recognisable family
    (e.g. class-agnostic edge proposals) are allowed through for the reference
    comparison to decide.
    """
    if not item_category:
        return False
    detection_category = _catalog_detection_category(detection)
    if not detection_category:
        return False
    return detection_category != item_category


def _catalog_detection_bbox(detection: dict[str, Any]) -> tuple[int, int, int, int] | None:
    bbox = detection.get("bbox") or {}
    try:
        return (
            int(float(bbox["x1"])),
            int(float(bbox["y1"])),
            int(float(bbox["x2"])),
            int(float(bbox["y2"])),
        )
    except (KeyError, TypeError, ValueError):
        return None


def _catalog_boxes_overlap(
    left: tuple[int, int, int, int],
    right: tuple[int, int, int, int],
    ratio: float = 0.6,
) -> bool:
    x1 = max(left[0], right[0])
    y1 = max(left[1], right[1])
    x2 = min(left[2], right[2])
    y2 = min(left[3], right[3])
    intersection = max(0, x2 - x1) * max(0, y2 - y1)
    if intersection <= 0:
        return False
    left_area = max(1, (left[2] - left[0]) * (left[3] - left[1]))
    right_area = max(1, (right[2] - right[0]) * (right[3] - right[1]))
    return intersection / min(left_area, right_area) >= ratio


def _catalog_dedupe_candidate_matches(
    matches: list[tuple[float, dict[str, Any]]]
) -> list[tuple[float, dict[str, Any]]]:
    """Collapse candidates that cover the same physical object.

    A class-agnostic edge proposal and the detector box for the same object can
    both pass the reference match, which would count one object twice. Prefer
    real detections (they carry the 3D quantity estimate) over edge proposals,
    then higher similarity, and drop later candidates that overlap a kept one.
    """

    def sort_key(entry: tuple[float, dict[str, Any]]):
        score, candidate = entry
        method = str((candidate.get("detection") or {}).get("method") or "")
        is_proposal = method == "class-agnostic-edge-proposal"
        return (1 if is_proposal else 0, -score)

    kept: list[tuple[float, dict[str, Any]]] = []
    for score, candidate in sorted(matches, key=sort_key):
        box = _catalog_detection_bbox(candidate.get("detection") or {})
        camera = candidate.get("camera_name")
        duplicate = False
        if box is not None:
            for _, kept_candidate in kept:
                if kept_candidate.get("camera_name") != camera:
                    continue
                kept_box = _catalog_detection_bbox(kept_candidate.get("detection") or {})
                if kept_box is not None and _catalog_boxes_overlap(box, kept_box):
                    duplicate = True
                    break
        if not duplicate:
            kept.append((score, candidate))
    return kept


def _catalog_class_agnostic_boxes(frame) -> list[tuple[int, int, int, int]]:
    """Return object-like regions without requiring a trained class label.

    This intentionally uses inexpensive edge/contour proposals so it can run as
    a fallback over many warehouse feeds. Product identity is decided later by
    reference matching; these boxes only answer "where might an object be?".
    """
    try:
        import cv2
    except ImportError:
        return []

    height, width = frame.shape[:2]
    if height < 32 or width < 32:
        return []
    max_dimension = max(height, width)
    scale = min(1.0, 960.0 / max_dimension)
    working = (
        cv2.resize(frame, (max(1, int(width * scale)), max(1, int(height * scale))))
        if scale < 1.0
        else frame
    )
    gray = cv2.cvtColor(working, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(gray, 35, 110)
    edges = cv2.morphologyEx(
        edges,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (9, 9)),
        iterations=2,
    )
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    work_height, work_width = gray.shape[:2]
    frame_area = float(work_height * work_width)
    proposals: list[tuple[float, tuple[int, int, int, int]]] = []
    for contour in contours:
        x, y, box_width, box_height = cv2.boundingRect(contour)
        area = float(box_width * box_height)
        area_ratio = area / frame_area
        aspect = box_width / max(1.0, float(box_height))
        fill = float(cv2.contourArea(contour)) / max(1.0, area)
        if (
            area_ratio < 0.0015
            or area_ratio > 0.65
            or box_width < 24
            or box_height < 24
            or aspect < 0.15
            or aspect > 6.5
            or fill < 0.08
        ):
            continue
        inverse_scale = 1.0 / scale
        pad_x = max(4, int(box_width * 0.04))
        pad_y = max(4, int(box_height * 0.04))
        box = (
            max(0, int((x - pad_x) * inverse_scale)),
            max(0, int((y - pad_y) * inverse_scale)),
            min(width, int((x + box_width + pad_x) * inverse_scale)),
            min(height, int((y + box_height + pad_y) * inverse_scale)),
        )
        proposals.append((area_ratio * (0.5 + fill), box))

    selected: list[tuple[int, int, int, int]] = []
    max_per_camera = max(
        1, min(int(os.getenv("CATALOG_PROPOSAL_MAX_PER_CAMERA", "24")), 60)
    )
    for _, box in sorted(proposals, key=lambda value: value[0], reverse=True):
        if any(_catalog_box_iou(box, existing) >= 0.40 for existing in selected):
            continue
        selected.append(box)
        if len(selected) >= max_per_camera:
            break

    # Plain cartons and wrapped products can have too few internal edges for a
    # contour detector. Add a small deterministic set of overlapping regions
    # so reference matching still receives candidates on low-texture frames.
    minimum_proposals = max(
        1, min(int(os.getenv("CATALOG_PROPOSAL_MIN_PER_CAMERA", "8")), 16)
    )
    grid_boxes: list[tuple[int, int, int, int]] = []
    for rows, columns in ((2, 2), (3, 3)):
        cell_width = width / columns
        cell_height = height / rows
        overlap_x = int(cell_width * 0.12)
        overlap_y = int(cell_height * 0.12)
        for row in range(rows):
            for column in range(columns):
                grid_boxes.append(
                    (
                        max(0, int(column * cell_width) - overlap_x),
                        max(0, int(row * cell_height) - overlap_y),
                        min(width, int((column + 1) * cell_width) + overlap_x),
                        min(height, int((row + 1) * cell_height) + overlap_y),
                    )
                )
    grid_boxes.insert(
        0,
        (
            int(width * 0.15),
            int(height * 0.15),
            int(width * 0.85),
            int(height * 0.85),
        ),
    )
    for box in grid_boxes:
        if len(selected) >= minimum_proposals or len(selected) >= max_per_camera:
            break
        if any(_catalog_box_iou(box, existing) >= 0.72 for existing in selected):
            continue
        selected.append(box)
    return selected


def _catalog_class_agnostic_candidates(frames: list[dict[str, Any]]) -> list[dict[str, Any]]:
    try:
        from recognition.embedding import image_embedding
    except ImportError:
        return []

    candidates: list[dict[str, Any]] = []
    for entry in frames:
        frame = entry["frame"]
        for index, (x1, y1, x2, y2) in enumerate(_catalog_class_agnostic_boxes(frame)):
            detection = {
                "track_id": f"proposal-{index}",
                "class_name": "object proposal",
                "object_type": "object proposal",
                "confidence": 0.5,
                "quantity": 1,
                "method": "class-agnostic-edge-proposal",
                "bbox": {"x1": x1, "y1": y1, "x2": x2, "y2": y2},
            }
            crop = _catalog_detection_crop(frame, detection["bbox"])
            if crop is None:
                continue
            candidates.append(
                {
                    "camera_name": entry["camera_name"],
                    "detection": detection,
                    "frame": frame,
                    "crop": crop,
                    "embedding": image_embedding(crop),
                }
            )
    return candidates


def _product_learning_public(session: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in session.items()
        if not key.startswith("_")
    }


def _product_learning_preview(session_id: str, index: int, crop: Any) -> str | None:
    try:
        import cv2
    except ImportError:
        return None
    directory = SNAPSHOT_DIR / "product-learning" / _catalog_visual_slug(session_id)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"view_{index:02d}.jpg"
    if not cv2.imwrite(str(path), crop):
        return None
    return f"/snapshots/product-learning/{quote(_catalog_visual_slug(session_id))}/{quote(path.name)}"


def _run_product_learning_session(session_id: str) -> None:
    """Collect stable, diverse views of the object presented after Learn starts."""
    try:
        import cv2
        from knowledge.similarity import cosine_similarity
        from recognition.embedding import image_embedding

        session = _product_learning_sessions[session_id]
        duration = int(session["duration_seconds"])
        baseline: dict[str, Any] = {}
        ranked: list[dict[str, Any]] = []
        started = time.monotonic()
        sample_index = 0
        while time.monotonic() - started < duration:
            health = _catalog_health_snapshot()
            frames = _catalog_live_frames(health, max_frames=100)
            selected_camera = _catalog_camera_label(session["camera_name"])
            frames = [
                entry
                for entry in frames
                if _catalog_camera_label(entry.get("camera_name")) == selected_camera
            ]
            session["camera_count"] = max(int(session.get("camera_count") or 0), len(frames))
            session["frames_seen"] = int(session.get("frames_seen") or 0) + len(frames)
            session["remaining_seconds"] = max(
                0, int(duration - (time.monotonic() - started))
            )
            for entry in frames:
                frame = entry["frame"]
                camera_name = _catalog_camera_label(entry["camera_name"])
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                gray = cv2.resize(gray, (320, 180))
                previous = baseline.get(camera_name)
                if previous is None:
                    baseline[camera_name] = gray
                    continue
                motion_map = cv2.absdiff(previous, gray)
                baseline[camera_name] = gray
                height, width = frame.shape[:2]
                scored: list[tuple[float, tuple[int, int, int, int], Any]] = []
                for box in _catalog_class_agnostic_boxes(frame):
                    x1, y1, x2, y2 = box
                    crop = _catalog_detection_crop(
                        frame, {"x1": x1, "y1": y1, "x2": x2, "y2": y2}
                    )
                    if crop is None:
                        continue
                    mx1 = max(0, min(319, int(x1 * 320 / width)))
                    my1 = max(0, min(179, int(y1 * 180 / height)))
                    mx2 = max(mx1 + 1, min(320, int(x2 * 320 / width)))
                    my2 = max(my1 + 1, min(180, int(y2 * 180 / height)))
                    motion = float(motion_map[my1:my2, mx1:mx2].mean()) / 255.0
                    area_ratio = ((x2 - x1) * (y2 - y1)) / max(1.0, width * height)
                    sharpness = min(
                        1.0,
                        float(cv2.Laplacian(crop, cv2.CV_64F).var()) / 600.0,
                    )
                    center_x = (x1 + x2) / (2.0 * width)
                    center_y = (y1 + y2) / (2.0 * height)
                    centrality = max(
                        0.0, 1.0 - (((center_x - 0.5) ** 2 + (center_y - 0.5) ** 2) ** 0.5)
                    )
                    score = motion * 5.0 + min(area_ratio, 0.35) + sharpness * 0.15 + centrality * 0.10
                    scored.append((score, box, crop))
                for score, box, crop in sorted(scored, key=lambda row: row[0], reverse=True)[:2]:
                    ranked.append(
                        {
                            "score": score,
                            "camera_name": camera_name,
                            "bbox": box,
                            "crop": crop,
                            "sample_index": sample_index,
                        }
                    )
                    session["proposal_count"] = int(session.get("proposal_count") or 0) + 1
            sample_index += 1
            time.sleep(1.0)

        session["status"] = "processing"
        session["remaining_seconds"] = 0
        embedded: list[dict[str, Any]] = []
        for candidate in sorted(ranked, key=lambda row: row["score"], reverse=True)[:80]:
            candidate["embedding"] = image_embedding(candidate["crop"])
            embedded.append(candidate)
        if not embedded:
            raise RuntimeError("No usable object candidates were found in the live camera frames.")

        seed = embedded[0]
        same_object = [
            candidate
            for candidate in embedded
            if cosine_similarity(candidate["embedding"], seed["embedding"]) >= 0.50
        ]
        selected: list[dict[str, Any]] = []
        seen_samples: set[tuple[str, int]] = set()
        for candidate in same_object:
            sample_key = (candidate["camera_name"], int(candidate["sample_index"]))
            if sample_key in seen_samples:
                continue
            if selected and max(
                cosine_similarity(candidate["embedding"], view["embedding"])
                for view in selected
            ) > 0.992:
                continue
            selected.append(candidate)
            seen_samples.add(sample_key)
            if len(selected) >= 8:
                break
        for candidate in same_object:
            if len(selected) >= 4:
                break
            if all(candidate is not existing for existing in selected):
                selected.append(candidate)
        if len(selected) < 2:
            raise RuntimeError(
                "The object was not visible in enough distinct frames. Move or rotate it and try again."
            )

        catalog = _get_catalog_db()
        existing_matches = []
        for item in catalog.list_items(str(session["scope_id"]), active_only=True):
            references = catalog.list_images(str(item["id"]), include_embeddings=True)
            per_view_scores = [
                max(
                    (
                        cosine_similarity(
                            view["embedding"], reference.get("embedding") or []
                        )
                        for reference in references
                    ),
                    default=0.0,
                )
                for view in selected
            ]
            required_score = float(
                os.getenv("CATALOG_EXISTING_PRODUCT_SIMILARITY_THRESHOLD", "0.82")
            )
            matching_indices = [
                index
                for index, score in enumerate(per_view_scores)
                if score >= required_score
            ]
            strongest = sorted(per_view_scores, reverse=True)[:2]
            score = sum(strongest) / len(strongest) if strongest else 0.0
            existing_matches.append(
                {
                    "item_id": str(item["id"]),
                    "name": str(item["name"]),
                    "confidence": round(float(score), 4),
                    "matching_view_indices": matching_indices,
                    "matching_view_count": len(matching_indices),
                }
            )
        existing_matches.sort(key=lambda match: match["confidence"], reverse=True)
        session["existing_matches"] = existing_matches[:3]
        threshold = float(
            os.getenv("CATALOG_EXISTING_PRODUCT_SIMILARITY_THRESHOLD", "0.82")
        )
        session["existing_match"] = (
            existing_matches[0]
            if existing_matches
            and existing_matches[0]["confidence"] >= threshold
            and existing_matches[0]["matching_view_count"] >= 2
            else None
        )

        previews = []
        for index, view in enumerate(selected, start=1):
            url = _product_learning_preview(session_id, index, view["crop"])
            if url:
                previews.append(
                    {
                        "index": index - 1,
                        "url": url,
                        "camera_name": view["camera_name"],
                        "score": round(float(view["score"]), 4),
                    }
                )
        session["_views"] = selected
        session["views"] = previews
        session["view_count"] = len(selected)
        session["remaining_seconds"] = 0
        session["status"] = "ready"
        session["completed_at"] = _now_iso()
    except Exception as exc:  # noqa: BLE001 - learning failures are returned to the operator
        session = _product_learning_sessions.get(session_id)
        if session is not None:
            session["status"] = "failed"
            session["remaining_seconds"] = 0
            session["error"] = str(exc)
            session["completed_at"] = _now_iso()
        _audit("product_learning_failed", {"session_id": session_id, "error": str(exc)})


def _catalog_fresh_yolo_crop_candidates(
    health: dict[str, Any],
    items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    try:
        from recognition.embedding import image_embedding
    except ImportError:
        return []

    # Catalog recognition must cover every loaded feed. The previous default
    # of eight silently skipped later channels (including channels 9, 10 and
    # 19) even though the persistent detector was counting objects there.
    max_frames = max(1, int(os.getenv("CATALOG_RECOGNITION_MAX_FRAMES", "100")))
    frames = _catalog_live_frames(health, max_frames=max_frames)
    if not frames:
        return []

    scope_id = str(items[0].get("scope_id") or "default")
    prompts = _catalog_detection_prompts(items, scope_id)
    scan = {
        "started_at": _now_iso(),
        "completed_at": None,
        "camera_count": len(frames),
        "cameras": [entry["camera_name"] for entry in frames],
        "prompts": prompts,
        "detection_count": 0,
        "candidate_count": 0,
        "proposal_count": 0,
        "error": None,
    }
    _catalog_yolo_last_scan[scope_id] = scan
    # Reference matching must never wait for a closed-set/prompt detector.
    # Build candidates for every live frame first, then optionally enrich them
    # with detector boxes. On CPU, running YOLO-World sequentially over 26
    # cameras can take longer than the complete live-recognition countdown and
    # previously meant the saved reference images were never consulted.
    candidates = _catalog_class_agnostic_candidates(frames)
    scan["proposal_source"] = "class_agnostic_reference"
    scan["proposal_count"] = len(candidates)
    scan["candidate_count"] = len(candidates)

    run_prompt_detector = str(
        os.getenv("CATALOG_RUN_PROMPT_DETECTOR", "false")
    ).strip().lower() in {"1", "true", "yes", "on"}
    if not run_prompt_detector:
        scan["prompt_detector_skipped"] = True
        scan["completed_at"] = _now_iso()
        return candidates

    detector = _catalog_yolo_for_prompts(prompts)
    if detector is None:
        scan["error"] = "Prompt detector could not be loaded; reference proposals were used."
        scan["completed_at"] = _now_iso()
        return candidates

    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    spatial_cfg = config.get("spatial_analysis", {})
    spatial_analyzer = (
        SpatialAnalyzer.from_config(spatial_cfg)
        if spatial_cfg.get("enabled", False)
        else None
    )
    try:
        for entry in frames:
            frame = entry["frame"]
            detections = detector.detect(frame)
            scan["detection_count"] += len(detections)
            if spatial_analyzer is not None:
                spatial_analyzer.enrich(frame, detections)
            for detection in detections:
                payload = _catalog_detection_payload(detection)
                crop = _catalog_detection_crop(frame, payload["bbox"])
                if crop is None:
                    continue
                candidates.append(
                    {
                        "camera_name": entry["camera_name"],
                        "detection": payload,
                        "frame": frame,
                        "crop": crop,
                        "embedding": image_embedding(crop),
                    }
                )
        if scan["detection_count"]:
            scan["proposal_source"] = "class_agnostic_reference_and_prompt_detector"
    except Exception as exc:
        scan["error"] = str(exc)
        _audit(
            "catalog_yolo_scan_failed",
            {"scope_id": scope_id, "error": str(exc)},
        )
    finally:
        scan["candidate_count"] = len(candidates)
        scan["completed_at"] = _now_iso()
    return candidates


def _catalog_match_current_frame(scope_id: str, include_visuals: bool = False) -> list[dict[str, Any]]:
    """One instantaneous pass: match catalog items (only items enrolled via
    AI Check-in) against whatever the detector's current spatial-object
    snapshot shows. Pure - reads live state but writes nothing, so both a
    single scheduled run and a live-run's repeated sampling can share it."""
    from knowledge.similarity import cosine_similarity

    db = _get_catalog_db()
    items = db.list_items(scope_id, active_only=True)
    health = _catalog_health_snapshot()
    cameras = health.get("cameras") or []
    by_camera = health.get("last_spatial_objects_by_camera") or {}
    if not by_camera and health.get("last_spatial_objects"):
        fallback_name = str((cameras[-1] if cameras else {}).get("name") or "camera")
        by_camera = {fallback_name: health.get("last_spatial_objects") or []}
    crop_candidates = _catalog_crop_candidates(health)
    cached_candidate_count = len(crop_candidates)
    crop_candidates.extend(_catalog_fresh_yolo_crop_candidates(health, items))
    scan = _catalog_yolo_last_scan.get(scope_id)
    if scan is not None:
        scan["cached_candidate_count"] = cached_candidate_count
        scan["total_candidate_count"] = len(crop_candidates)
        scan.setdefault("catalog_scores", {})
    frame_embeddings = _catalog_frame_embeddings(health)
    visual_threshold = float(os.getenv("CATALOG_VISUAL_SIMILARITY_THRESHOLD", "0.94"))
    # The local reference embedding is a compact color histogram + edge
    # density, not a learned re-identification network. Requiring 0.90 made
    # the same package fail after normal changes in angle, scale and lighting.
    crop_threshold = float(os.getenv("CATALOG_CROP_SIMILARITY_THRESHOLD", "0.70"))
    proposal_threshold = float(
        os.getenv("CATALOG_PROPOSAL_SIMILARITY_THRESHOLD", "0.88")
    )
    if scan is not None:
        scan["crop_similarity_threshold"] = crop_threshold
        scan["proposal_similarity_threshold"] = proposal_threshold

    matches: list[dict[str, Any]] = []
    for item in items:
        target = _catalog_normalize_name(item["name"])
        aliases = _catalog_item_prompts(scope_id, str(item["id"]))
        targets = {target, *(_catalog_normalize_name(alias) for alias in aliases)}
        references = db.list_images(str(item["id"]), include_embeddings=True)
        matched_entries: list[tuple[str, dict[str, Any]]] = []
        for camera_name, objects in by_camera.items():
            matched_entries.extend(
                (_catalog_camera_label(camera_name), obj)
                for obj in objects or []
                if _catalog_normalize_name(obj.get("inventory_name")) in targets
            )

        camera_counts: dict[str, int] = {}
        quantity, camera_counts = _catalog_count_objects_by_camera(matched_entries)
        confidence = 1.0 if quantity > 0 else 0.0
        measurement = matched_entries[0][1] if matched_entries else None
        method = str(measurement.get("method") or "catalog-name-and-3d") if measurement else None
        visual_evidence = _catalog_spatial_visuals(health, matched_entries) if include_visuals and matched_entries else {}

        if not matched_entries:
            # Manual-entry recognition is confirmed by the reference photos, not
            # by the detector's class label. For each candidate object we:
            #   1. reject candidates whose product family clearly differs from
            #      the enrolled item (a sack/bag is never counted as a box), so
            #      a mislabelled sack cannot be forced onto a box entry;
            #   2. require the crop to match one of the item's reference images
            #      above a visual-similarity threshold;
            #   3. de-duplicate overlapping candidates (an edge proposal and the
            #      detector box over the same object) so it is counted once.
            # This replaces the old shortcut that counted any box-shaped
            # detection for a single enrolled item by name alone.
            item_category = _catalog_item_category(str(item["name"]), aliases)
            # Regions occupied by a conflicting-family detection (e.g. sacks for
            # a box item). Any candidate overlapping one of these - including
            # label-less edge proposals - is rejected, so a sack cannot be
            # counted as a box merely because its crop looks similar.
            excluded_boxes: list[tuple[int, int, int, int]] = []
            for candidate in crop_candidates:
                detection = candidate.get("detection") or {}
                if _catalog_category_conflicts(item_category, detection):
                    box = _catalog_detection_bbox(detection)
                    if box is not None:
                        excluded_boxes.append(box)
            rejected_family = 0
            rejected_overlap = 0
            rejected_similarity = 0
            crop_matches: list[tuple[float, dict[str, Any]]] = []
            best_crop_score = 0.0
            for candidate in crop_candidates:
                detection = candidate.get("detection") or {}
                if _catalog_category_conflicts(item_category, detection):
                    rejected_family += 1
                    continue
                candidate_box = _catalog_detection_bbox(detection)
                # Only a label-less proposal inherits a conflicting region's
                # exclusion. A candidate that is itself a recognised box is
                # trusted by its own label even when it overlaps an adjacent
                # sack - otherwise boxes stacked next to sacks are lost.
                if (
                    candidate_box is not None
                    and _catalog_detection_category(detection) is None
                    and any(
                        _catalog_boxes_overlap(candidate_box, excluded)
                        for excluded in excluded_boxes
                    )
                ):
                    rejected_overlap += 1
                    continue
                score = max(
                    (
                        cosine_similarity(candidate["embedding"], ref.get("embedding") or [])
                        for ref in references
                    ),
                    default=0.0,
                )
                best_crop_score = max(best_crop_score, score)
                candidate_method = str(detection.get("method") or "")
                accepted_threshold = (
                    proposal_threshold
                    if candidate_method == "class-agnostic-edge-proposal"
                    else crop_threshold
                )
                if score >= accepted_threshold:
                    crop_matches.append((score, candidate))
                else:
                    rejected_similarity += 1
            crop_matches = _catalog_dedupe_candidate_matches(crop_matches)
            if scan is not None:
                scan["catalog_scores"][str(item["name"])] = round(best_crop_score, 4)
                # Rejection breakdown so it is clear where boxes disappear:
                # family mismatch, sack-overlap, or reference similarity.
                scan.setdefault("diagnostics", {})[str(item["name"])] = {
                    "candidates": len(crop_candidates),
                    "accepted": len(crop_matches),
                    "rejected_family": rejected_family,
                    "rejected_overlap": rejected_overlap,
                    "rejected_similarity": rejected_similarity,
                    "best_score": round(best_crop_score, 4),
                    "crop_threshold": crop_threshold,
                    "proposal_threshold": proposal_threshold,
                    "item_category": item_category,
                }

            if crop_matches:
                confidence = max(score for score, _ in crop_matches)
                quantity, camera_counts = _catalog_count_objects_by_camera(
                    [
                        (_catalog_camera_label(candidate.get("camera_name")), candidate["detection"])
                        for _, candidate in crop_matches
                    ]
                )
                measurement = crop_matches[0][1]["detection"]
                method = str(measurement.get("method") or "catalog-crop-reference-and-3d")
                if include_visuals:
                    visual_evidence = _catalog_candidate_visuals([candidate for _, candidate in crop_matches])

        if not matched_entries and quantity <= 0:
            best: tuple[float, str] | None = None
            for camera_name, frame_embedding in frame_embeddings.items():
                score = max(
                    (cosine_similarity(frame_embedding, ref.get("embedding") or []) for ref in references),
                    default=0.0,
                )
                if best is None or score > best[0]:
                    best = (score, camera_name)
            if best and best[0] >= visual_threshold:
                confidence = best[0]
                camera_objects = by_camera.get(best[1]) or []
                quantity = sum(max(1, int(obj.get("quantity") or 1)) for obj in camera_objects) or 1
                camera_counts = {_catalog_camera_label(best[1]): quantity}
                measurement = camera_objects[0] if camera_objects else None
                method = "catalog-reference-and-3d"
                if include_visuals:
                    visual_evidence = _catalog_spatial_visuals(health, [(best[1], measurement or {})])

        match = {
            "item_id": str(item["id"]),
            "item_name": str(item["name"]),
            "quantity": quantity,
            "confidence": confidence,
            "dimensions_m": _catalog_dimensions(measurement),
            "measurement_method": method,
            "camera_counts": _catalog_camera_counts_payload(camera_counts),
        }
        if visual_evidence:
            match["_visual_evidence"] = visual_evidence
        matches.append(match)
    return matches


def _catalog_unidentified_current_frame(
    scope_id: str, include_visuals: bool = False
) -> list[dict[str, Any]]:
    """Count detector objects that do not match anything enrolled in AI Check-in.

    Recognition results historically discarded these objects. A live counting
    session must retain them, grouped by camera, so the reported total can be
    reconciled with the objects visible in each feed.
    """
    db = _get_catalog_db()
    items = db.list_items(scope_id, active_only=True)
    if not items:
        return []

    health = _catalog_health_snapshot()
    cameras = health.get("cameras") or []
    by_camera = health.get("last_spatial_objects_by_camera") or {}
    if not by_camera:
        by_camera = health.get("last_detections_by_camera") or {}
    if not by_camera and health.get("last_spatial_objects"):
        fallback_name = str((cameras[-1] if cameras else {}).get("name") or "camera")
        by_camera = {fallback_name: health.get("last_spatial_objects") or []}

    unknown_entries: list[tuple[str, dict[str, Any]]] = []
    for camera_name, objects in by_camera.items():
        for obj in objects or []:
            if any(
                _catalog_detection_matches_item_prompt(
                    obj,
                    str(item["name"]),
                    _catalog_item_prompts(scope_id, str(item["id"])),
                )
                for item in items
            ):
                continue
            unknown_entries.append((_catalog_camera_label(camera_name), obj))

    quantity, camera_counts = _catalog_count_objects_by_camera(unknown_entries)
    if quantity <= 0:
        return []

    match: dict[str, Any] = {
        # Use an enrolled item FK so existing result databases remain compatible;
        # _state_key keeps this aggregate separate from that enrolled item.
        "item_id": str(items[0]["id"]),
        "item_name": "Unidentified",
        "quantity": quantity,
        "confidence": 0.0,
        "dimensions_m": None,
        "measurement_method": "unidentified-detector-object",
        "camera_counts": _catalog_camera_counts_payload(camera_counts),
        "_state_key": "__unidentified__",
    }
    if include_visuals:
        match["_visual_evidence"] = _catalog_spatial_visuals(health, unknown_entries)
    return [match]


def _run_catalog_recognition(scope_id: str) -> dict[str, Any]:
    """Create one immutable catalog-only count snapshot for a scope."""
    db = _get_catalog_db()
    health = _catalog_health_snapshot()
    cameras = health.get("cameras") or []
    interval = _catalog_interval_hours()
    run_id = db.start_run(scope_id, interval, len(cameras))
    try:
        for match in _catalog_sample_current_frame(scope_id, include_visuals=True):
            db.add_result(run_id=run_id, **_catalog_persist_match_visuals(scope_id, run_id, match))
        db.complete_run(run_id)
    except Exception:
        db.complete_run(run_id, status="failed")
        raise
    return {
        "run": db.latest_run(scope_id),
        "results": db.latest_results(scope_id),
        "schedule": _catalog_schedule(scope_id),
    }


def _live_catalog_status_payload(scope_id: str) -> dict[str, Any]:
    state = _live_catalog_runs.get(scope_id)
    if not state:
        return {"running": False}
    running = state["status"] == "running"
    remaining = 0
    if running:
        ends_at = _catalog_datetime(state["ends_at"])
        if ends_at is not None:
            remaining = max(0, int((ends_at - datetime.now(timezone.utc)).total_seconds()))
    results = sorted(
        state["items"].values(), key=lambda result: (-result["quantity"], result["item_name"])
    )
    return {
        "running": running,
        "status": state["status"],
        "started_at": state["started_at"],
        "ends_at": state["ends_at"],
        "remaining_seconds": remaining,
        "results": results,
        "run_id": state.get("run_id"),
        "yolo_scan": _catalog_yolo_last_scan.get(scope_id),
    }


async def _run_live_catalog_recognition(scope_id: str, ends_at: datetime) -> None:
    global _catalog_run_lock
    state = _live_catalog_runs[scope_id]
    live_visual_run_id = f"live-{_catalog_visual_slug(state['started_at'])}"
    try:
        while datetime.now(timezone.utc) < ends_at:
            try:
                matches = await asyncio.to_thread(_catalog_match_current_frame, scope_id, True)
                matches.extend(
                    await asyncio.to_thread(
                        _catalog_unidentified_current_frame, scope_id, True
                    )
                )
            except Exception as exc:  # keep sampling - one bad sample shouldn't end the run
                state["error"] = str(exc)
            else:
                for match in matches:
                    if match["quantity"] <= 0:
                        continue
                    state_key = str(match.get("_state_key") or match["item_id"])
                    persisted = _catalog_persist_match_visuals(
                        scope_id, live_visual_run_id, match
                    )
                    if match.get("_state_key"):
                        persisted["_state_key"] = match["_state_key"]
                    existing = state["items"].get(state_key)
                    if existing is None:
                        state["items"][state_key] = persisted
                    else:
                        merged = _catalog_merge_match_samples([[existing], [persisted]])[0]
                        media_by_camera = {}
                        for result in (existing, persisted):
                            for entry in result.get("camera_counts") or []:
                                if entry.get("frame_url") or entry.get("crop_url"):
                                    media_by_camera[_catalog_camera_label(entry.get("camera_name"))] = entry
                        for entry in merged.get("camera_counts") or []:
                            media = media_by_camera.get(
                                _catalog_camera_label(entry.get("camera_name"))
                            )
                            if media:
                                for key in ("frame_url", "crop_url", "bbox", "class_name"):
                                    if media.get(key) is not None:
                                        entry[key] = media[key]
                        if persisted.get("_state_key"):
                            merged["_state_key"] = persisted["_state_key"]
                        state["items"][state_key] = merged
            remaining = (ends_at - datetime.now(timezone.utc)).total_seconds()
            if remaining <= 0:
                break
            await asyncio.sleep(min(CATALOG_LIVE_RUN_SAMPLE_INTERVAL_SECONDS, remaining))
    finally:
        if _catalog_run_lock is None:
            _catalog_run_lock = asyncio.Lock()
        async with _catalog_run_lock:
            db = _get_catalog_db()
            health = _catalog_health_snapshot()
            run_id = db.start_run(
                scope_id, _catalog_interval_hours(), len(health.get("cameras") or [])
            )
            for match in state["items"].values():
                db.add_result(run_id=run_id, **_catalog_persist_match_visuals(scope_id, run_id, match))
            db.complete_run(run_id)
        state["status"] = "completed"
        state["run_id"] = run_id
        _audit("catalog_recognition_completed", {"scope_id": scope_id, "run_id": run_id, "mode": "live"})


def _run_due_catalog_scopes() -> None:
    now = datetime.now(timezone.utc)
    for scope_id in _get_catalog_db().list_scopes():
        schedule = _catalog_schedule(scope_id)
        next_run = _catalog_datetime(schedule["next_run_at"])
        if next_run is None or next_run <= now:
            _run_catalog_recognition(scope_id)


def _parse_event_log(limit: int = 40) -> list[dict[str, Any]]:
    if not LOG_PATH.exists():
        return []

    content = LOG_PATH.read_text(encoding="utf-8", errors="replace")
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    entries: list[dict[str, Any]] = []
    i = 0
    while i + 3 < len(lines):
        timestamp = lines[i]
        class_line = lines[i + 1]
        camera = lines[i + 2]
        confidence = lines[i + 3]
        match = re.match(r"^(.*) detected$", class_line, re.IGNORECASE)
        class_name = match.group(1) if match else class_line
        entries.append(
            {
                "timestamp": timestamp,
                "class_name": class_name,
                "camera": camera,
                "confidence": confidence,
            }
        )
        i += 5
    return entries[-limit:]


def _read_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _tail_file(path: Path, limit: int = 80) -> list[str]:
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    return [_redact_sensitive_text(line) for line in lines[-max(1, min(limit, 500)) :]]


@app.get("/api/recognitions")
def recognitions(limit: int = 40) -> dict[str, Any]:
    entries = _parse_event_log(limit)
    counts: dict[str, int] = {}
    for entry in entries:
        counts[entry["class_name"]] = counts.get(entry["class_name"], 0) + 1
    distinct = [
        {"class_name": class_name, "count": count}
        for class_name, count in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
    ]
    warehouse_db = _get_warehouse_db()
    movements = warehouse_db.recent_movements(limit)
    stock = warehouse_db.get_all_stock()
    movement_counts = warehouse_db.movement_counts()
    stock_by_name = {item["name"]: int(item.get("current_stock") or 0) for item in stock}
    all_movements = warehouse_db.recent_movements(500)
    movement_totals: dict[tuple[str, str], int] = {}
    for movement in all_movements:
        key = (movement["product_name"], movement["direction"])
        movement_totals[key] = movement_totals.get(key, 0) + int(movement.get("quantity") or 1)
    vision_items = [
        {
            "product_name": product_name,
            "state": "check-in" if direction == "IN" else "check-out",
            "quantity": stock_by_name.get(product_name, quantity) if direction == "IN" else quantity,
            "current_stock": stock_by_name.get(product_name, 0),
        }
        for (product_name, direction), quantity in sorted(
            movement_totals.items(),
            key=lambda item: (-item[1], item[0][0], item[0][1]),
        )
    ]
    status = _status()
    return {
        "running": status["running"],
        "entries": entries,
        "counts": distinct,
        "vision_items": vision_items,
        "movements": movements,
        "movement_counts": movement_counts,
        "stock": stock,
    }


@app.get("/api/warehouse/stock")
def warehouse_stock() -> dict[str, Any]:
    db = _get_warehouse_db()
    return {"stock": db.get_all_stock(), "movement_counts": db.movement_counts()}


@app.get("/api/warehouse/movements")
def warehouse_movements(limit: int = 50) -> dict[str, Any]:
    db = _get_warehouse_db()
    return {"movements": db.recent_movements(limit=max(1, min(limit, 500)))}


@app.get("/api/warehouse-engine/overview")
def warehouse_engine_overview(limit: int = 100) -> dict[str, Any]:
    db = EngineDatabase(str(WAREHOUSE_ENGINE_DB_PATH))
    objects = db.objects(limit=max(1, min(limit, 500)))
    events = db.events(limit=max(1, min(limit, 500)))
    active = [row for row in objects if row.get("current_status") == "active"]
    zones: dict[str, int] = {}
    for row in active:
        zone = str(row.get("current_zone") or "Unassigned")
        zones[zone] = zones.get(zone, 0) + 1
    return {
        "detected_objects": len(objects),
        "tracked_objects": len(active),
        "inventory_objects": len(
            [row for row in active if row.get("product_name")]
        ),
        "active_events": len(events),
        "zone_statistics": zones,
        "objects": objects,
        "events": events,
    }


@app.get("/api/warehouse-engine/events")
def warehouse_engine_events(limit: int = 200) -> dict[str, Any]:
    return {
        "events": EngineDatabase(str(WAREHOUSE_ENGINE_DB_PATH)).events(
            limit=max(1, min(limit, 1000))
        )
    }


@app.post("/api/warehouse-engine/tasks/parse")
def warehouse_engine_parse_task(request: WarehouseTaskRequest) -> dict[str, Any]:
    return {"prompt": request.prompt, "task": parse_task_prompt(request.prompt)}


def _poll_process() -> None:
    global _last_exit_code, _process, _started_at, _stdout_handle, _stderr_handle
    if _process is None:
        return

    exit_code = _process.poll()
    if exit_code is not None:
        _last_exit_code = exit_code
        _process = None
        _started_at = None
        for handle in (_stdout_handle, _stderr_handle):
            if handle is not None:
                handle.close()
        _stdout_handle = None
        _stderr_handle = None
        _clear_detector_pid()


def _write_detector_pid(pid: int) -> None:
    DETECTION_PID_PATH.parent.mkdir(parents=True, exist_ok=True)
    DETECTION_PID_PATH.write_text(str(pid), encoding="utf-8")


def _clear_detector_pid() -> None:
    try:
        DETECTION_PID_PATH.unlink()
    except FileNotFoundError:
        pass


def _read_detector_pid() -> int | None:
    try:
        value = DETECTION_PID_PATH.read_text(encoding="utf-8").strip()
    except FileNotFoundError:
        return None

    try:
        return int(value)
    except ValueError:
        _clear_detector_pid()
        return None


def _is_detector_command(command_line: str | None) -> bool:
    if not command_line:
        return False
    normalized = command_line.replace("\\", "/").lower()
    root_marker = str(ROOT).replace("\\", "/").lower()
    main_marker = str(ROOT / "main.py").replace("\\", "/").lower()
    return root_marker in normalized and main_marker in normalized and "--config" in normalized


def _process_command_line(pid: int) -> str | None:
    if pid <= 0:
        return None

    if os.name == "nt":
        command = (
            "$p = Get-CimInstance Win32_Process -Filter 'ProcessId = "
            f"{pid}' -ErrorAction SilentlyContinue; "
            "if ($p) { $p.CommandLine }"
        )
        try:
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command", command],
                capture_output=True,
                text=True,
                timeout=3,
            )
        except Exception:
            return None
        if result.returncode != 0:
            return None
        return result.stdout.strip() or None

    proc_cmdline = Path("/proc") / str(pid) / "cmdline"
    try:
        return proc_cmdline.read_text(encoding="utf-8", errors="replace").replace("\x00", " ")
    except OSError:
        return None


def _pid_is_detector(pid: int) -> bool:
    return _is_detector_command(_process_command_line(pid))


def _discover_detector_pid() -> int | None:
    if os.name == "nt":
        command = (
            "Get-CimInstance Win32_Process | "
            "Where-Object { $_.Name -like 'python*' -and $_.CommandLine -like '*main.py*' } | "
            'ForEach-Object { [string]$_.ProcessId + "`t" + $_.CommandLine }'
        )
        try:
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command", command],
                capture_output=True,
                text=True,
                timeout=3,
            )
        except Exception:
            return None
        if result.returncode != 0:
            return None
        lines = result.stdout.splitlines()
    else:
        try:
            result = subprocess.run(
                ["pgrep", "-af", "main.py"],
                capture_output=True,
                text=True,
                timeout=3,
            )
        except Exception:
            return None
        lines = result.stdout.splitlines()

    for line in lines:
        pid_text, _separator, command_line = line.partition("\t")
        if not command_line and " " in pid_text:
            pid_text, _separator, command_line = pid_text.partition(" ")
        try:
            pid = int(pid_text.strip())
        except ValueError:
            continue
        if _is_detector_command(command_line):
            return pid
    return None


def _detector_pid() -> int | None:
    _poll_process()
    if _process is not None:
        return _process.pid

    pid = _read_detector_pid()
    if pid is None:
        discovered_pid = _discover_detector_pid()
        if discovered_pid is not None:
            _write_detector_pid(discovered_pid)
        return discovered_pid
    if _pid_is_detector(pid):
        return pid

    _clear_detector_pid()
    return None


def _terminate_pid(pid: int) -> int | None:
    if os.name == "nt":
        subprocess.run(["taskkill", "/PID", str(pid), "/T"], capture_output=True, text=True, timeout=10)
        if _pid_is_detector(pid):
            subprocess.run(
                ["taskkill", "/PID", str(pid), "/T", "/F"],
                capture_output=True,
                text=True,
                timeout=10,
            )
        return None

    os.kill(pid, signal.SIGTERM)
    deadline = time.time() + 10
    while time.time() < deadline:
        if not _pid_is_detector(pid):
            return 0
        time.sleep(0.2)

    os.kill(pid, signal.SIGKILL)
    return None


def _validate_active_cameras_for_start() -> None:
    db = _get_camera_db()
    active_cameras = db.list_active_cameras(include_secret=True)
    if not active_cameras:
        raise HTTPException(
            status_code=400,
            detail="Assign at least one active camera slot before starting detection.",
        )

    for camera in active_cameras:
        db.set_status(camera["id"], "stream_managed")

    # V2: analytics never opens RTSP for validation. The Stream Manager owns
    # the upstream connection and is allowed to be starting/reconnecting while
    # YOLO comes online as a secondary frame consumer.
    _set_config_active_cameras(active_cameras)


def _status() -> dict[str, Any]:
    pid = _detector_pid()
    return {
        "running": pid is not None,
        "pid": pid,
        "started_at": _started_at,
        "uptime_seconds": round(time.time() - _started_at, 1)
        if _started_at
        else 0,
        "last_exit_code": _last_exit_code,
        "health": _read_json(DETECTION_HEALTH_PATH),
        "streams": _get_stream_manager().status().get("streams", []),
        "stdout_tail": _tail_file(DETECTION_STDOUT_PATH, 40),
        "stderr_tail": _tail_file(DETECTION_STDERR_PATH, 40),
    }


def _should_autostart_detection() -> bool:
    if not _env_bool("AUTO_START_DETECTION", True):
        return False
    try:
        active = _get_camera_db().list_active_cameras(include_secret=False)
    except Exception:
        return False
    return bool(active)


def _clear_live_frames() -> None:
    if not SNAPSHOT_DIR.exists():
        return
    for pattern in ("latest.jpg", "latest_slot_*.jpg", "latest_*.jpg"):
        for path in SNAPSHOT_DIR.glob(pattern):
            try:
                path.unlink()
            except OSError:
                pass


def _ensure_detection_running(reason: str = "watchdog") -> None:
    global _watchdog_last_start_attempt
    if _manual_stop_requested or not _should_autostart_detection():
        return
    if _detector_pid() is not None:
        return

    now = time.time()
    cooldown_seconds = int(os.getenv("DETECTION_WATCHDOG_COOLDOWN_SECONDS", "45"))
    if now - _watchdog_last_start_attempt < cooldown_seconds:
        return

    _watchdog_last_start_attempt = now
    try:
        start_detection(StartRequest())
        _audit(
            "detection_autostart",
            {"reason": reason, "started": True},
            actor="watchdog",
        )
    except HTTPException as exc:
        _audit(
            "detection_autostart_failed",
            {"reason": reason, "status_code": exc.status_code, "detail": str(exc.detail)},
            actor="watchdog",
        )
    except Exception as exc:
        _audit(
            "detection_autostart_failed",
            {"reason": reason, "error": _redact_sensitive_text(str(exc))},
            actor="watchdog",
        )


async def _detection_watchdog() -> None:
    await asyncio.sleep(int(os.getenv("DETECTION_AUTOSTART_DELAY_SECONDS", "8")))
    while _env_bool("DETECTION_WATCHDOG_ENABLED", True):
        await asyncio.to_thread(_ensure_detection_running, "watchdog")
        await asyncio.sleep(int(os.getenv("DETECTION_WATCHDOG_INTERVAL_SECONDS", "30")))


async def _catalog_recognition_scheduler() -> None:
    await asyncio.sleep(int(os.getenv("CATALOG_RECOGNITION_STARTUP_DELAY_SECONDS", "60")))
    while _env_bool("CATALOG_RECOGNITION_SCHEDULER_ENABLED", True):
        try:
            await asyncio.to_thread(_run_due_catalog_scopes)
        except Exception as exc:  # noqa: BLE001
            _audit(
                "catalog_recognition_scheduler_failed",
                {"error": _redact_sensitive_text(str(exc))},
                actor="scheduler",
            )
        await asyncio.sleep(int(os.getenv("CATALOG_RECOGNITION_POLL_SECONDS", "300")))


@app.on_event("startup")
async def start_detection_watchdog() -> None:
    global _catalog_recognition_task, _watchdog_task
    # Self-heal the training dataset on boot: restore from the newest in-volume
    # backup if the working tree came up empty, then ensure the baseline dirs.
    try:
        _ensure_training_dataset()
    except Exception:  # noqa: BLE001 - never block startup on this
        pass
    if _watchdog_task is None or _watchdog_task.done():
        _watchdog_task = asyncio.create_task(_detection_watchdog())
    if _catalog_recognition_task is None or _catalog_recognition_task.done():
        _catalog_recognition_task = asyncio.create_task(_catalog_recognition_scheduler())


_HTML_NO_CACHE = {"Cache-Control": "no-cache"}


@app.get("/")
def dashboard() -> FileResponse:
    return FileResponse(DASHBOARD_V2_DIR / "index.html", headers=_HTML_NO_CACHE)


@app.get("/dashboard-v2")
def dashboard_v2() -> FileResponse:
    return FileResponse(DASHBOARD_V2_DIR / "index.html", headers=_HTML_NO_CACHE)


@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> FileResponse:
    return FileResponse(DASHBOARD_V2_DIR / "favicon.svg", media_type="image/svg+xml")


@app.post("/api/v2/auth/login")
def v2_auth_login(payload: V2LoginRequest, request: Request) -> dict[str, Any]:
    ac = _get_access_control_db()
    user = ac.authenticate_user(payload.email, payload.password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    passkeys = ac.list_passkeys(int(user["id"]))
    if passkeys:
        options = generate_authentication_options(
            rp_id=_v2_rp_id(request),
            allow_credentials=[
                PublicKeyCredentialDescriptor(id=base64url_to_bytes(passkey["credential_id"]))
                for passkey in passkeys
            ],
            user_verification=UserVerificationRequirement.REQUIRED,
        )
        challenge_id = ac.create_challenge(int(user["id"]), bytes_to_base64url(options.challenge), "login")
        return {
            "requires_passkey": True,
            "challenge_id": challenge_id,
            "publicKey": _v2_public_key_options(options),
            "user": {"id": user["id"], "name": user["name"], "email": user["email"]},
        }
    token = ac.create_session(int(user["id"]))
    _audit("v2.auth.login", {"user_id": user["id"], "method": "password"}, actor=user["email"])
    return _v2_auth_response(user, token)


@app.post("/api/v2/auth/setup-password")
def v2_auth_setup_password(payload: V2LoginRequest, request: Request) -> dict[str, Any]:
    ac = _get_access_control_db()
    user = ac.get_user_by_email(payload.email.strip().lower())
    if not user:
        raise HTTPException(status_code=404, detail="Account not found.")
    if user.get("has_password"):
        raise HTTPException(status_code=409, detail="Password is already set for this account.")
    if user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Account is disabled.")
    try:
        user = ac.set_user_password(int(user["id"]), payload.password)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    token = ac.create_session(int(user["id"]))
    _audit("v2.auth.initial_password_set", {"user_id": user["id"]}, actor=user["email"])
    return _v2_auth_response(user, token)


@app.post("/api/v2/auth/login/passkey/options")
def v2_auth_login_passkey_options(payload: V2PasskeyLoginStart, request: Request) -> dict[str, Any]:
    ac = _get_access_control_db()
    user = ac.get_user_by_email(payload.email.strip().lower())
    if not user or user.get("status") != "active":
        raise HTTPException(status_code=401, detail="Account not found or disabled.")
    passkeys = ac.list_passkeys(int(user["id"]))
    if not passkeys:
        raise HTTPException(status_code=404, detail="No fingerprint, Face ID, or passkey is registered for this account yet.")
    options = generate_authentication_options(
        rp_id=_v2_rp_id(request),
        allow_credentials=[
            PublicKeyCredentialDescriptor(id=base64url_to_bytes(passkey["credential_id"]))
            for passkey in passkeys
        ],
        user_verification=UserVerificationRequirement.REQUIRED,
    )
    challenge_id = ac.create_challenge(int(user["id"]), bytes_to_base64url(options.challenge), "login")
    return {
        "requires_passkey": True,
        "challenge_id": challenge_id,
        "publicKey": _v2_public_key_options(options),
        "user": {"id": user["id"], "name": user["name"], "email": user["email"]},
    }


@app.post("/api/v2/auth/login/passkey")
def v2_auth_login_passkey(payload: V2PasskeyLoginFinish, request: Request) -> dict[str, Any]:
    ac = _get_access_control_db()
    user = ac.get_user_by_email(payload.email.strip().lower())
    if not user:
        raise HTTPException(status_code=401, detail="Invalid login.")
    credential_id = str(payload.credential.get("id") or payload.credential.get("rawId") or "")
    passkey = ac.get_passkey(credential_id)
    if not passkey or int(passkey["user_id"]) != int(user["id"]):
        raise HTTPException(status_code=401, detail="Unknown passkey.")
    challenge = ac.consume_challenge(int(user["id"]), payload.challenge_id, "login")
    if not challenge:
        raise HTTPException(status_code=401, detail="Passkey challenge expired.")
    try:
        verified = verify_authentication_response(
            credential=payload.credential,
            expected_challenge=base64url_to_bytes(challenge),
            expected_rp_id=_v2_rp_id(request),
            expected_origin=_v2_expected_origins(request),
            credential_public_key=base64url_to_bytes(passkey["public_key"]),
            credential_current_sign_count=int(passkey["sign_count"] or 0),
            require_user_verification=True,
        )
    except Exception as exc:
        raise HTTPException(status_code=401, detail=f"Passkey verification failed: {exc}") from exc
    ac.update_passkey_sign_count(credential_id, int(verified.new_sign_count))
    token = ac.create_session(int(user["id"]))
    _audit("v2.auth.login", {"user_id": user["id"], "method": "password+passkey"}, actor=user["email"])
    return _v2_auth_response(user, token)


@app.get("/api/v2/auth/me")
def v2_auth_me(request: Request) -> dict[str, Any]:
    user = _v2_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Login required.")
    dashboard = _get_access_control_db().resolve_dashboard(user_id=int(user["id"]))
    return {"user": dashboard["user"], "modules": dashboard["modules"]}


@app.post("/api/v2/auth/passkeys/register/options")
def v2_passkey_register_options(payload: V2PasskeyRegisterStart, request: Request) -> dict[str, Any]:
    user = _v2_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Login required.")
    ac = _get_access_control_db()
    existing = ac.list_passkeys(int(user["id"]))
    options = generate_registration_options(
        rp_id=_v2_rp_id(request),
        rp_name=os.getenv("WEBAUTHN_RP_NAME", "AI Vision"),
        user_id=str(user["id"]).encode("utf-8"),
        user_name=user["email"],
        user_display_name=user["name"],
        authenticator_selection=AuthenticatorSelectionCriteria(
            authenticator_attachment=AuthenticatorAttachment.PLATFORM,
            resident_key=ResidentKeyRequirement.REQUIRED,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
        exclude_credentials=[
            PublicKeyCredentialDescriptor(id=base64url_to_bytes(passkey["credential_id"]))
            for passkey in existing
        ],
    )
    challenge_id = ac.create_challenge(int(user["id"]), bytes_to_base64url(options.challenge), "register")
    return {"challenge_id": challenge_id, "publicKey": _v2_public_key_options(options)}


@app.post("/api/v2/auth/passkeys/register/verify")
def v2_passkey_register_verify(payload: V2PasskeyRegisterFinish, request: Request) -> dict[str, Any]:
    user = _v2_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Login required.")
    ac = _get_access_control_db()
    challenge = ac.consume_challenge(int(user["id"]), payload.challenge_id, "register")
    if not challenge:
        raise HTTPException(status_code=401, detail="Passkey challenge expired.")
    try:
        verified = verify_registration_response(
            credential=payload.credential,
            expected_challenge=base64url_to_bytes(challenge),
            expected_rp_id=_v2_rp_id(request),
            expected_origin=_v2_expected_origins(request),
            require_user_verification=True,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Passkey registration failed: {exc}") from exc
    passkey = ac.add_passkey(
        int(user["id"]),
        bytes_to_base64url(verified.credential_id),
        bytes_to_base64url(verified.credential_public_key),
        int(verified.sign_count),
        payload.name or "Fingerprint / passkey",
    )
    _audit("v2.auth.passkey_registered", {"user_id": user["id"], "credential_id": passkey["credential_id"]}, actor=user["email"])
    return {"ok": True, "passkey": {"name": passkey["name"]}}


@app.get("/api/v2/me/dashboard")
def v2_me_dashboard(request: Request) -> dict[str, Any]:
    return _v2_dashboard(request)


@app.get("/api/v2/me/module/{module_code}")
def v2_module_data(module_code: str, request: Request) -> dict[str, Any]:
    dashboard = _v2_require_module(request, module_code)
    if module_code == "live_monitoring":
        cameras = _get_camera_db().list_active_cameras(include_secret=False)
        allowed_ids = set(dashboard.get("scope", {}).get("camera_ids") or [])
        if allowed_ids:
            cameras = [camera for camera in cameras if str(camera.get("id")) in allowed_ids]
        return {"module": module_code, "cameras": cameras, "status": _status()}
    if module_code == "counting":
        return {
            "module": module_code,
            "stock": _get_warehouse_db().get_all_stock(),
            "movement_counts": _get_warehouse_db().movement_counts(),
        }
    if module_code in {"reports", "activity_history"}:
        return {"module": module_code, "movements": _get_warehouse_db().recent_movements(limit=50)}
    if module_code == "products":
        return {"module": module_code, "stock": _get_warehouse_db().get_all_stock()}
    if module_code == "system_health":
        return {"module": module_code, "status": _status(), "opencv": opencv_diagnostics()}
    if module_code == "audit_logs":
        _v2_require_module(request, module_code, permission="audit.view")
        return security_audit(limit=100)
    return {"module": module_code, "message": "Module shell is assigned and ready for implementation."}


@app.get("/api/v2/admin/overview")
def v2_admin_overview(request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "dashboard.view")
    ac = _get_access_control_db()
    cameras = _get_camera_db().list_cameras(include_secret=False)
    active_cameras = [camera for camera in cameras if camera["is_active"]]
    health = (_status().get("health") or {})
    return {
        "totals": {
            "companies": len([org for org in ac.list_organizations() if org["type"] == "company"]),
            "factories": len([org for org in ac.list_organizations() if org["type"] == "factory"]),
            "warehouses": len([org for org in ac.list_organizations() if org["type"] == "warehouse"]),
            "users": len(ac.list_users()),
            "online_cameras": len(active_cameras),
            "offline_cameras": max(0, len(cameras) - len(active_cameras)),
            "active_ai_processes": 1 if _status()["running"] else 0,
            "products_counted_today": health.get("last_detection_count", 0),
            "active_alerts": 0,
        },
        "server_status": _status(),
        "recent_activity": _get_security_audit_db().recent(limit=12),
    }


@app.get("/api/v2/admin/bootstrap")
def v2_admin_bootstrap(request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.view")
    ac = _get_access_control_db()
    return {
        "users": ac.list_users(),
        "roles": ac.list_roles(),
        "permissions": ac.list_permissions(),
        "modules": ac.list_modules(),
        "organizations": ac.list_organizations(),
    }


@app.post("/api/v2/admin/users")
def v2_admin_create_user(payload: V2UserCreate, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.create")
    user = _get_access_control_db().create_user(payload.name, payload.email.strip().lower())
    _audit("v2.user.created", {"user": user}, actor=_v2_user_email(request))
    return user


@app.post("/api/v2/admin/users/{user_id}/password")
def v2_admin_set_user_password(user_id: int, payload: V2PasswordSet, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.edit")
    try:
        user = _get_access_control_db().set_user_password(user_id, payload.password)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    _audit("v2.user.password_set", {"user_id": user_id}, actor=_v2_user_email(request))
    return user


@app.post("/api/v2/admin/users/{user_id}/auth-preference")
def v2_admin_set_user_auth_preference(user_id: int, payload: V2AuthPreferenceSet, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.edit")
    try:
        user = _get_access_control_db().set_user_auth_preference(user_id, payload.preferred_auth_method)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    _audit("v2.user.auth_preference_set", {"user_id": user_id, "preferred_auth_method": payload.preferred_auth_method}, actor=_v2_user_email(request))
    return user


@app.post("/api/v2/admin/roles")
def v2_admin_create_role(payload: V2RoleCreate, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "role.create")
    role = _get_access_control_db().create_role(payload.name, payload.code)
    _audit("v2.role.created", {"role": role}, actor=_v2_user_email(request))
    return role


@app.post("/api/v2/admin/users/{user_id}/roles")
def v2_admin_assign_role(user_id: int, payload: V2RoleAssignment, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.edit")
    _get_access_control_db().assign_role(user_id, payload.role_code)
    _audit("v2.user.role_assigned", {"user_id": user_id, "role_code": payload.role_code}, actor=_v2_user_email(request))
    return _get_access_control_db().resolve_dashboard(user_id=user_id)


@app.delete("/api/v2/admin/users/{user_id}/roles/{role_code}")
def v2_admin_remove_role(user_id: int, role_code: str, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.edit")
    _get_access_control_db().remove_role(user_id, role_code)
    _audit("v2.user.role_removed", {"user_id": user_id, "role_code": role_code}, actor=_v2_user_email(request))
    return _get_access_control_db().resolve_dashboard(user_id=user_id)


@app.post("/api/v2/admin/users/{user_id}/modules")
def v2_admin_assign_module(user_id: int, payload: V2ModuleAssignment, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "module.assign")
    _get_access_control_db().set_user_module(user_id, payload.module_code, payload.effect, payload.display_order)
    _audit("v2.user.module_changed", {"user_id": user_id, **payload.model_dump()}, actor=_v2_user_email(request))
    return _get_access_control_db().resolve_dashboard(user_id=user_id)


@app.post("/api/v2/admin/users/{user_id}/permissions")
def v2_admin_assign_permission(user_id: int, payload: V2PermissionAssignment, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.edit")
    _get_access_control_db().set_user_permission(user_id, payload.permission_code, payload.effect)
    _audit("v2.user.permission_changed", {"user_id": user_id, **payload.model_dump()}, actor=_v2_user_email(request))
    return _get_access_control_db().resolve_dashboard(user_id=user_id)


@app.post("/api/v2/admin/users/{user_id}/scopes")
def v2_admin_assign_scope(user_id: int, payload: V2ScopeAssignment, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "scope.assign")
    _get_access_control_db().set_user_scope(user_id, payload.scope_type, payload.scope_ids, payload.effect)
    _audit("v2.user.scope_changed", {"user_id": user_id, **payload.model_dump()}, actor=_v2_user_email(request))
    return _get_access_control_db().resolve_dashboard(user_id=user_id)


@app.post("/api/v2/admin/users/{user_id}/disable")
def v2_admin_disable_user(user_id: int, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.disable")
    user = _get_access_control_db().set_user_status(user_id, "disabled")
    _audit("v2.user.disabled", {"user_id": user_id}, actor=_v2_user_email(request))
    return user or {}


@app.post("/api/v2/admin/users/{user_id}/reactivate")
def v2_admin_reactivate_user(user_id: int, request: Request) -> dict[str, Any]:
    _v2_require_permission(request, "user.edit")
    user = _get_access_control_db().set_user_status(user_id, "active")
    _audit("v2.user.reactivated", {"user_id": user_id}, actor=_v2_user_email(request))
    return user or {}


@app.get("/api/v2/rbac/me")
def dashboard_v2_rbac_me(request: Request) -> dict[str, Any]:
    context = _rbac_context(request)
    permissions = set(context["permissions"])
    return {
        **context,
        "surfaces": {
            "head": _authorized_modules("head", permissions),
            "user": _authorized_modules("user", permissions),
        },
        "available_roles": [
            {"id": role, "label": role.replace("_", " ").title()}
            for role in ROLE_PERMISSIONS
        ],
    }


@app.get("/api/v2/navigation")
def dashboard_v2_navigation(request: Request, surface: str = "head") -> dict[str, Any]:
    surface = surface.strip().lower()
    if surface not in DASHBOARD_V2_MODULES:
        raise HTTPException(status_code=400, detail="Unknown dashboard surface.")
    context = _require_permission(request, "view_dashboard")
    return {
        "surface": surface,
        "modules": _authorized_modules(surface, set(context["permissions"])),
        "role": context["role"],
        "scope": context["scope"],
    }


@app.get("/api/v2/head/overview")
def dashboard_v2_head_overview(request: Request) -> dict[str, Any]:
    context = _require_permission(request, "view_dashboard")
    status_data = _status()
    health = status_data.get("health") or {}
    cameras = _get_camera_db().list_cameras(include_secret=False)
    active_cameras = [camera for camera in cameras if camera["is_active"]]
    stock = _get_warehouse_db().get_all_stock()
    movement_counts = _get_warehouse_db().movement_counts()
    audit = _get_security_audit_db().verify()
    return {
        "context": context,
        "summary": {
            "organizations": 1,
            "active_cameras": len(active_cameras),
            "saved_cameras": len(cameras),
            "detector_running": status_data["running"],
            "frames_read": health.get("frames_read", 0),
            "last_frame_at": health.get("last_frame_at"),
            "last_detection_count": health.get("last_detection_count", 0),
            "stock_items": len(stock),
            "audit_verified": audit.get("verified", False),
        },
        "health": health,
        "movement_counts": movement_counts,
        "future_integrations": [
            "ERP",
            "HRM",
            "CRM",
            "Inventory Management",
            "Quality Control",
            "Predictive Analytics",
            "Multi-site Management",
            "API Integrations",
        ],
    }


@app.get("/api/v2/user/overview")
def dashboard_v2_user_overview(request: Request) -> dict[str, Any]:
    context = _require_permission(request, "view_dashboard")
    status_data = _status()
    health = status_data.get("health") or {}
    stock = _get_warehouse_db().get_all_stock()
    movements = _get_warehouse_db().recent_movements(limit=12)
    return {
        "context": context,
        "summary": {
            "detector_running": status_data["running"],
            "active_cameras": health.get("camera_count", 0),
            "frames_read": health.get("frames_read", 0),
            "last_detection_count": health.get("last_detection_count", 0),
            "last_tracked_count": health.get("last_tracked_count", 0),
            "stock_items": len(stock),
            "open_verification_tasks": 0,
            "active_alerts": 0,
        },
        "stock": stock[:12],
        "recent_movements": movements,
    }


@app.get("/api/status")
def status() -> dict[str, Any]:
    data = _status()
    data["security"] = {
        "api_key_required": _security_enabled(),
        "allowed_origins": _env_list("ALLOWED_ORIGINS", DEFAULT_ALLOWED_ORIGINS),
    }
    return data


@app.get("/api/security/audit")
def security_audit(limit: int = 100) -> dict[str, Any]:
    db = _get_security_audit_db()
    return {
        "chain": db.verify(),
        "events": db.recent(limit=limit),
    }


@app.get("/api/diagnostics/opencv")
def opencv_diagnostics() -> dict[str, Any]:
    try:
        import cv2

        return {
            "ok": True,
            "version": getattr(cv2, "__version__", "unknown"),
        }
    except Exception as exc:
        return {
            "ok": False,
            "error": _redact_sensitive_text(str(exc)),
            "error_type": type(exc).__name__,
        }


@app.get("/api/config")
def get_config() -> dict[str, Any]:
    return _redact_config(_read_yaml(CONFIG_PATH))


@app.patch("/api/config")
def update_config(patch: ConfigPatch) -> dict[str, Any]:
    data = _read_yaml(CONFIG_PATH)
    detection = data.setdefault("detection", {})
    display = data.setdefault("display", {})
    spatial = data.setdefault("spatial_analysis", {})
    snapshots = data.setdefault("snapshots", {})
    logging_cfg = data.setdefault("logging", {})
    tracking = data.setdefault("tracking", {})
    warehouse_counting = data.setdefault("warehouse_counting", {})
    recognition = data.setdefault("recognition", {})

    values = patch.model_dump(exclude_unset=True)
    if "model_path" in values:
        detection["model_path"] = values["model_path"]
    if "fallback_model_path" in values:
        detection["fallback_model_path"] = values["fallback_model_path"]
    if "confidence_threshold" in values:
        detection["confidence_threshold"] = values["confidence_threshold"]
    if "iou_threshold" in values:
        detection["iou_threshold"] = values["iou_threshold"]
    if "max_detections" in values:
        detection["max_detections"] = values["max_detections"]
    if "image_size" in values:
        detection["image_size"] = values["image_size"]
    if "device" in values:
        detection["device"] = values["device"] or "auto"
    if "target_fps" in values:
        detection["target_fps"] = values["target_fps"]
    if "stale_after_ms" in values:
        detection["stale_after_ms"] = values["stale_after_ms"]
    if "max_concurrent_cameras" in values:
        detection["max_concurrent_cameras"] = values["max_concurrent_cameras"]
    if "classes" in values:
        detection["classes"] = values["classes"] or None
    if "class_prompts" in values:
        detection["class_prompts"] = values["class_prompts"] or None
    if "class_agnostic_nms" in values:
        detection["class_agnostic_nms"] = values["class_agnostic_nms"]
    if "show_fps" in values:
        display["show_fps"] = values["show_fps"]
    if "live_feed_enabled" in values:
        display["live_feed_enabled"] = values["live_feed_enabled"]
    if "live_frame_width" in values:
        display["live_frame_width"] = values["live_frame_width"]
    if "live_frame_jpeg_quality" in values:
        display["live_frame_jpeg_quality"] = values["live_frame_jpeg_quality"]
    if "spatial_enabled" in values:
        spatial["enabled"] = values["spatial_enabled"]
    if "horizontal_fov_degrees" in values:
        spatial["horizontal_fov_degrees"] = values["horizontal_fov_degrees"]
    if "camera_height_m" in values:
        spatial["camera_height_m"] = values["camera_height_m"]
    if "horizon_y_ratio" in values:
        spatial["horizon_y_ratio"] = values["horizon_y_ratio"]
    if "min_distance_m" in values:
        spatial["min_distance_m"] = values["min_distance_m"]
    if "max_distance_m" in values:
        spatial["max_distance_m"] = values["max_distance_m"]
    if "estimate_depth_layers" in values:
        spatial["estimate_depth_layers"] = values["estimate_depth_layers"]
    if "max_units_per_detection" in values:
        spatial["max_units_per_detection"] = values["max_units_per_detection"]
    if "tracking_enabled" in values:
        tracking["enabled"] = values["tracking_enabled"]
    if "tracking_grace_period_seconds" in values:
        tracking["grace_period_seconds"] = values["tracking_grace_period_seconds"]
    if "warehouse_counting_enabled" in values:
        warehouse_counting["enabled"] = values["warehouse_counting_enabled"]
    if "warehouse_confidence_threshold" in values:
        warehouse_counting["confidence_threshold"] = values["warehouse_confidence_threshold"]
    if "count_low_confidence_as_unknown" in values:
        warehouse_counting["count_low_confidence_as_unknown"] = values["count_low_confidence_as_unknown"]
    if "snapshots_enabled" in values:
        snapshots["enabled"] = values["snapshots_enabled"]
    if "snapshot_trigger_classes" in values:
        snapshots["trigger_classes"] = values["snapshot_trigger_classes"] or []
    if "snapshot_cooldown_seconds" in values:
        snapshots["cooldown_seconds"] = values["snapshot_cooldown_seconds"]
    if "logging_enabled" in values:
        logging_cfg["enabled"] = values["logging_enabled"]
    if "recognition_enabled" in values:
        recognition["enabled"] = values["recognition_enabled"]
    if "recognition_provider" in values:
        recognition["provider"] = values["recognition_provider"]
    if "recognition_model" in values:
        recognition["model"] = values["recognition_model"]
    if "recognition_confidence_threshold" in values:
        recognition["confidence_threshold"] = values["recognition_confidence_threshold"]
    if "recognition_similarity_threshold" in values:
        recognition["similarity_threshold"] = values["recognition_similarity_threshold"]
    if "recognition_cache_enabled" in values:
        recognition["cache_enabled"] = values["recognition_cache_enabled"]
    if "recognition_cache_expiration" in values:
        recognition["cache_expiration"] = values["recognition_cache_expiration"]
    if "recognition_timeout" in values:
        recognition["timeout"] = values["recognition_timeout"]
    if "recognition_retries" in values:
        recognition["retries"] = values["recognition_retries"]
    if "recognition_max_workers" in values:
        recognition["max_workers"] = values["recognition_max_workers"]
    if "recognition_catalog_only" in values:
        recognition["catalog_only"] = values["recognition_catalog_only"]

    _write_yaml(CONFIG_PATH, data)
    _audit("config_updated", {"fields": sorted(values.keys())})
    return _redact_config(data)


def _redact_config(data: dict[str, Any]) -> dict[str, Any]:
    redacted = json.loads(json.dumps(data))
    for camera in redacted.get("cameras", []) or []:
        source = camera.get("source")
        if source is not None:
            camera["source"] = _redact_sensitive_text(str(source))
    return redacted


@app.get("/api/cameras")
def list_cameras() -> dict[str, Any]:
    db = _get_camera_db()
    cameras = db.list_cameras(include_secret=False)
    active_cameras = [camera for camera in cameras if camera["is_active"]]
    active = active_cameras[0] if active_cameras else None
    return {"cameras": cameras, "active_camera": active, "active_cameras": active_cameras}


def _camera_operations_payload() -> list[dict[str, Any]]:
    cameras = _get_camera_db().list_cameras(include_secret=False)
    settings = VisionConfigDB(os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))).get_camera_settings_map(
        [camera["id"] for camera in cameras]
    )
    stream_rows = _get_stream_manager().status().get("streams", [])
    streams = {str(row.get("channel_id")): row for row in stream_rows}
    payload = []
    for camera in cameras:
        stream = streams.get(str(camera["id"])) or {}
        operator_settings = settings.get(str(camera["id"])) or {}
        raw_status = str(stream.get("status") or camera.get("status") or "").lower()
        if raw_status in {"online", "connected"}:
            operator_status = "live"
        elif camera.get("is_active") and raw_status not in {"offline", "failed", "timeout", "stopped"}:
            operator_status = "waiting"
        else:
            operator_status = "offline"
        fps = stream.get("fps") or stream.get("current_fps") or 0
        last_frame = stream.get("last_frame_at")
        decode_errors = stream.get("decode_errors", 0)
        reconnect_count = stream.get("reconnect_count", 0)
        row = dict(camera)
        row.update({
            "camera_id": camera["id"],
            "camera_name": camera["name"],
            "rtsp_url": camera.get("masked_stream_url", ""),
            "status": operator_status,
            "fps": fps,
            "last_frame": last_frame,
            "decode_errors": decode_errors,
            "reconnect_count": reconnect_count,
            "block_id": operator_settings.get("block_id"),
            "block_name": operator_settings.get("block_name"),
            "health": {
                "status": operator_status,
                "fps": fps,
                "last_frame_at": last_frame,
                "frame_age_ms": stream.get("frame_age_ms"),
                "reconnect_count": reconnect_count,
                "decode_errors": decode_errors,
                "stream_latency_ms": stream.get("stream_latency_ms") or stream.get("latency_ms"),
                "last_error": stream.get("last_error"),
            },
        })
        payload.append(row)
    return payload


@app.get("/api/v1/blocks")
def list_camera_assignment_blocks() -> dict[str, Any]:
    blocks = VisionConfigDB(
        os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))
    ).list_blocks()
    return {"data": blocks, "meta": {"count": len(blocks)}}


@app.get("/api/v1/blocks/{block_id}/cameras")
def list_block_cameras(block_id: int) -> dict[str, Any]:
    config_db = VisionConfigDB(
        os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))
    )
    block = next(
        (row for row in config_db.list_blocks() if int(row["id"]) == block_id),
        None,
    )
    if block is None:
        raise HTTPException(status_code=404, detail="Block not found.")
    cameras = [
        row
        for row in _camera_operations_payload()
        if row.get("block_id") is not None and int(row["block_id"]) == block_id
    ]
    return {"data": cameras, "meta": {"count": len(cameras), "block": block}}


@app.get("/api/v1/cameras")
def list_operator_cameras() -> dict[str, Any]:
    rows = _camera_operations_payload()
    blocks = VisionConfigDB(os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))).list_blocks()
    return {"data": rows, "meta": {"count": len(rows), "blocks": blocks}}


@app.put("/api/v1/cameras/{camera_id}")
def update_operator_camera(camera_id: int, body: CameraOperationsUpdate) -> dict[str, Any]:
    db = _get_camera_db()
    current = db.get_camera(camera_id, include_secret=True)
    if current is None:
        raise HTTPException(status_code=404, detail="Camera not found.")
    block_name = body.block_name.strip() if body.block_name is not None else None
    if body.block_id is None and not block_name:
        raise HTTPException(status_code=422, detail="Enter a block name before saving this camera.")
    if body.stream_url is not None:
        _endpoint, validation_error = _camera_stream_endpoint(body.stream_url)
        if validation_error:
            raise HTTPException(status_code=400, detail=validation_error)
    try:
        db.update_camera(camera_id, name=body.name, stream_url=body.stream_url)
        config_db = VisionConfigDB(os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db")))
        block_id = body.block_id
        if block_name:
            block = next((row for row in config_db.list_blocks() if row["name"].casefold() == block_name.casefold()), None)
            if block is None:
                block = config_db.create_block(block_name)
            block_id = int(block["id"])
        config_db.assign_camera_block(camera_id, block_id)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if current.get("is_active") and (body.name is not None or body.stream_url is not None):
        refreshed = db.get_camera(camera_id, include_secret=True)
        _get_stream_manager().stop(str(camera_id))
        _start_stream_for_camera(refreshed)
        _sync_config_active_cameras(db)
    return {"data": next(row for row in _camera_operations_payload() if row["id"] == camera_id), "meta": {}}


@app.post("/api/v1/cameras/{camera_id}/reconnect")
def reconnect_operator_camera(camera_id: int) -> dict[str, Any]:
    camera = _get_camera_db().get_camera(camera_id, include_secret=True)
    if camera is None:
        raise HTTPException(status_code=404, detail="Camera not found.")
    if not camera.get("is_active"):
        raise HTTPException(status_code=409, detail="Assign the camera to a live slot before reconnecting.")
    manager = _get_stream_manager()
    manager.stop(str(camera_id))
    stream = _start_stream_for_camera(camera)
    return {"data": next(row for row in _camera_operations_payload() if row["id"] == camera_id), "meta": {"stream": stream}}


@app.post("/api/v1/cameras/{camera_id}/test-frame")
def test_operator_camera_frame(camera_id: int) -> dict[str, Any]:
    camera = _get_camera_db().get_camera(camera_id, include_secret=True)
    if camera is None:
        raise HTTPException(status_code=404, detail="Camera not found.")
    result = _test_camera_stream(camera["stream_url"])
    _get_camera_db().set_status(camera_id, result["status"])
    return {"data": {"camera_id": camera_id, **result}, "meta": {}}


@app.post("/api/cameras/test")
def test_camera_stream(request: CameraTestRequest) -> dict[str, Any]:
    return _test_camera_stream(request.stream_url)


@app.post("/api/cameras")
def save_camera(camera: CameraCreate) -> dict[str, Any]:
    db = _get_camera_db()
    _endpoint, validation_error = _camera_stream_endpoint(camera.stream_url)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    test_result = (
        _test_camera_stream(camera.stream_url)
        if camera.test_connection
        else {"status": "unknown", "message": "Saved without testing."}
    )
    saved = db.add_camera(
        name=camera.name.strip(),
        stream_url=camera.stream_url.strip(),
        status=test_result["status"],
    )

    active = None
    stream = None
    if camera.make_active and test_result["status"] == "connected":
        slot_number = camera.slot_number or _next_available_slot(db.list_cameras(include_secret=False))
        active = db.assign_slot(saved["id"], slot_number)
        _sync_config_active_cameras(db)
        stream = _start_stream_for_camera(active)
        if _status()["running"]:
            stop_detection()
            start_detection(StartRequest())

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    return {
        "camera": db.get_camera(saved["id"], include_secret=False),
        "active_camera": db.get_camera(active["id"], include_secret=False) if active else None,
        "active_cameras": active_cameras,
        "test": test_result,
        "stream": stream,
        "cameras": cameras,
    }


def _register_controller_channels(
    controller: CameraControllerCreate, db: CameraDB
) -> dict[str, Any]:
    """Test and save/activate every channel on a controller.

    Shared by the Add NVR endpoint and the environment-based boot-time seed
    so both paths apply the exact same validation instead of drifting apart
    - the seed path used to have its own separate, untested activation loop
    that could silently occupy real slots with channels that had never
    actually been checked.
    """
    endpoint = _controller_endpoint(controller)
    if not endpoint["host"]:
        raise HTTPException(status_code=400, detail="Controller IP/host is required.")

    private_host_message = _private_controller_host_message(endpoint["host"])
    if controller.require_public and private_host_message:
        raise HTTPException(status_code=400, detail=private_host_message)

    controller_error = None

    saved_cameras = []
    test_results = []
    controller_reachable = controller_error is None

    # A camera can be registered (saved, tested, shown on the dashboard)
    # without being active (occupying a real slot the detector actually
    # opens a connection to). MAX_CAMERA_SLOTS bounds how many cameras can
    # be active - i.e. actually connected - at once, to protect the
    # droplet's CPU/memory/bandwidth; it does not bound how many cameras
    # can be registered. A controller with more channels than there are
    # free slots right now still saves and tests every channel - it just
    # leaves the channels beyond the free-slot budget registered but
    # inactive instead of rejecting the whole request.
    channel_numbers = (
        [int(channel) for channel in controller.channels]
        if controller.channels
        else [controller.channel_start + index for index in range(controller.channel_count)]
    )
    controller_stream_urls = [
        _controller_stream_url(controller, channel) for channel in channel_numbers
    ]
    controller_stream_url_set = set(controller_stream_urls)
    all_cameras = db.list_cameras(include_secret=True)
    used_slots = {
        int(camera["slot_number"])
        for camera in all_cameras
        if camera.get("slot_number") is not None
        and camera.get("stream_url") not in controller_stream_url_set
    }
    next_slot = max(controller.start_slot, 1)
    while next_slot in used_slots and next_slot <= MAX_CAMERA_SLOTS:
        next_slot += 1

    for index, stream_url in enumerate(controller_stream_urls):
        channel = channel_numbers[index]

        if controller.test_streams and controller_reachable:
            test_result = _test_camera_stream(stream_url)
        elif controller_reachable:
            test_result = {
                "status": "connected",
                "message": "Controller channels registered; Stream Manager owns connection validation.",
            }
        else:
            test_result = {
                "status": "failed",
                "message": controller_error or "Controller endpoint is not reachable.",
            }

        saved = db.upsert_camera_by_stream_url(
            name=_controller_camera_name(controller, channel, next_slot),
            stream_url=stream_url,
            status=test_result["status"],
        )
        _delete_duplicate_stream_url_cameras(db, stream_url, int(saved["id"]))

        active = None
        assigned_slot = None
        stream_status = None
        message = test_result["message"]
        if controller.make_active and test_result["status"] == "connected":
            active, assigned_slot, next_slot = _activate_stream_managed_camera(
                db,
                int(saved["id"]),
                next_slot,
                used_slots,
                reuse_existing_slot=False,
            )
            if active is not None:
                stream_status = _start_stream_for_camera(active)
            else:
                message = (
                    f"Reachable, but no free camera slot is available right now "
                    f"({MAX_CAMERA_SLOTS} active slot limit reached). Deactivate "
                    "another camera to free one up for this one."
                )

        saved_cameras.append(db.get_camera(saved["id"], include_secret=False))
        test_results.append(
            {
                "camera_id": saved["id"],
                "slot_number": assigned_slot,
                "channel": channel,
                "status": test_result["status"],
                "message": message,
                "active": active is not None,
                "stream": stream_status,
            }
        )

    return {
        "endpoint": endpoint,
        "private_host_message": private_host_message,
        "controller_error": controller_error,
        "controller_reachable": controller_reachable,
        "saved_cameras": saved_cameras,
        "test_results": test_results,
    }


@app.post("/api/discovery/scan")
async def discovery_scan(request: DiscoveryScanRequest) -> dict[str, Any]:
    """Device-first discovery: given only an IP/hostname, report what the
    device is and which streaming services it exposes, with no RTSP URL,
    stream path, or vendor supplied by the operator.

    The scan is CPU-light but network-bound (several socket probes), so it runs
    off the event loop. The Discovery Engine itself enforces the target-host
    safety policy (see discovery.portscan.resolve_and_guard)."""
    result = await asyncio.to_thread(discover_device, request.host.strip())
    _audit(
        "discovery_scan",
        {
            "host": request.host.strip(),
            "reachable": result.reachable,
            "vendor": result.fingerprint.vendor,
            "service_count": len(result.services),
        },
    )
    return result.to_dict()


@app.post("/api/v2/devices/discover")
async def v2_discover_device(request: V2DeviceDiscoverRequest) -> dict[str, Any]:
    """V2 entry point: discovery starts from only an IP address or hostname."""
    result = await asyncio.to_thread(discover_device, request.host.strip())
    payload = result.to_dict()
    device = _get_device_db().upsert_device_from_discovery(
        name=(request.name or request.host).strip(),
        host=request.host.strip(),
        result=payload,
    )
    _audit(
        "v2_device_discover",
        {
            "device_id": device.get("id"),
            "host": request.host.strip(),
            "reachable": payload.get("reachable"),
            "service_count": len(payload.get("services") or []),
        },
    )
    return {"device": device, "discovery": payload}


@app.get("/api/v2/devices")
def v2_list_devices() -> dict[str, Any]:
    return {"devices": _get_device_db().list_devices()}


@app.post("/api/v2/devices/{device_id}/authenticate")
def v2_authenticate_device(device_id: int, request: V2DeviceAuthenticateRequest) -> dict[str, Any]:
    """Authenticate only after service selection, then enumerate channels."""
    db = _get_device_db()
    device = db.get_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found.")

    port = request.port or STREAM_DEFAULT_PORTS.get(request.protocol.lower(), 554)
    username = request.username or ""
    password = request.password or ""
    if not username:
        # Reuse credentials already stored server-side for this NVR host. This
        # allows a rediscovered recorder to expand its channel count without
        # sending secrets back to the browser or asking the operator again.
        for camera in _get_camera_db().list_cameras(include_secret=True):
            try:
                parsed = urlsplit(str(camera.get("stream_url") or ""))
            except ValueError:
                continue
            if parsed.hostname == str(device["host"]) and parsed.username:
                username = parsed.username
                password = parsed.password or ""
                break
    credentials = StreamCredentials(username, password)
    enumeration = enumerate_streams(
        host=str(device["host"]),
        port=port,
        protocol=request.protocol,
        credentials=credentials,
        vendor=_v2_stream_vendor_hint(device, request),
        channel_count=request.channel_count,
    )
    if enumeration.requires_auth and not request.username:
        raise HTTPException(
            status_code=401,
            detail="This service requires authentication. Provide credentials after selecting the service.",
        )
    if not enumeration.channels:
        raise HTTPException(
            status_code=502,
            detail=_redact_sensitive_text(enumeration.error or "No channels were discovered."),
        )

    camera_db = _get_camera_db()
    stream_url_set = {stream.stream_url for stream in enumeration.channels}
    all_cameras = camera_db.list_cameras(include_secret=True)
    used_slots = {
        int(camera["slot_number"])
        for camera in all_cameras
        if camera.get("slot_number") is not None
        and camera.get("stream_url") not in stream_url_set
    }
    next_slot = 1
    while next_slot in used_slots and next_slot <= MAX_CAMERA_SLOTS:
        next_slot += 1
    channels = []
    stream_statuses = []
    for stream in enumeration.channels:
        test_result = (
            _test_camera_stream(stream.stream_url)
            if request.test_streams
            else {"status": "connected", "message": "Stream registered without a pre-flight test."}
        )
        camera = camera_db.upsert_camera_by_stream_url(
            name=f"{device['name']} {stream.name}",
            stream_url=stream.stream_url,
            status=test_result["status"],
        )
        _delete_duplicate_stream_url_cameras(camera_db, stream.stream_url, int(camera["id"]))

        assigned_slot = None
        active = None
        if request.make_active and test_result["status"] == "connected":
            active, assigned_slot, next_slot = _activate_stream_managed_camera(
                camera_db,
                int(camera["id"]),
                next_slot,
                used_slots,
                reuse_existing_slot=False,
            )

        channel = db.upsert_channel(
            device_id=device_id,
            external_channel_id=str(stream.channel),
            name=f"{device['name']} {stream.name}",
            profile=stream.description,
            stream_reference=stream.stream_url,
            camera_id=camera["id"],
            slot_number=assigned_slot,
        )
        if active is not None:
            try:
                status = _start_stream_for_camera(active)
            except Exception as exc:
                # Registration is a database operation; a camera that is
                # temporarily offline or already reconnecting must not abort
                # the remaining NVR channels or produce a raw CORS-masked 500.
                status = {
                    "channel_id": str(active["id"]),
                    "slot_number": active.get("slot_number"),
                    "status": "reconnecting",
                    "last_error": _redact_sensitive_text(str(exc)),
                }
            db.update_stream_session(channel["id"], status)
            stream_statuses.append(status)
        channels.append(channel)

    config_sync_warning = None
    try:
        _sync_config_active_cameras(camera_db)
    except Exception as exc:
        # The database is the source of truth. A stale/read-only compatibility
        # YAML file must not turn a successful device registration into a raw
        # 500 response (which browsers misleadingly report as a CORS failure).
        config_sync_warning = _redact_sensitive_text(str(exc))
        _audit(
            "v2_device_config_sync_failed",
            {"device_id": device_id, "error": config_sync_warning},
        )
    _audit("v2_device_authenticate", {"device_id": device_id, "channels": len(channels)})
    return {
        "provider": enumeration.provider,
        "device": db.get_device(device_id),
        "channels": channels,
        "streams": stream_statuses,
        "warning": config_sync_warning,
    }


@app.get("/api/v2/devices/{device_id}/channels")
def v2_device_channels(device_id: int) -> dict[str, Any]:
    device = _get_device_db().get_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found.")
    return {"device": device, "channels": device.get("channels", [])}


def _register_discovered_channels(
    name_prefix: str,
    channels: list[Any],
    make_active: bool,
    test_streams: bool,
    db: CameraDB,
) -> list[dict[str, Any]]:
    """Save/activate a provider's enumerated channels, reusing the same
    register-without-activate + slot-budget behaviour as the controller path:
    every channel is saved, and channels beyond the free active-slot budget
    are left registered-but-inactive rather than rejecting the whole request.
    """
    stream_url_set = {stream.stream_url for stream in channels}
    all_cameras = db.list_cameras(include_secret=True)
    used_slots = {
        int(camera["slot_number"])
        for camera in all_cameras
        if camera.get("slot_number") is not None
        and camera.get("stream_url") not in stream_url_set
    }
    next_slot = 1
    while next_slot in used_slots and next_slot <= MAX_CAMERA_SLOTS:
        next_slot += 1
    results: list[dict[str, Any]] = []

    for stream in channels:
        if test_streams:
            test_result = _test_camera_stream(stream.stream_url)
        else:
            test_result = {"status": "connected", "message": "Stream registered without a pre-flight test."}

        saved = db.upsert_camera_by_stream_url(
            name=f"{name_prefix} {stream.name}",
            stream_url=stream.stream_url,
            status=test_result["status"],
        )
        _delete_duplicate_stream_url_cameras(db, stream.stream_url, int(saved["id"]))

        active = None
        assigned_slot = None
        stream_status = None
        message = test_result["message"]
        if make_active and test_result["status"] == "connected":
            active, assigned_slot, next_slot = _activate_stream_managed_camera(
                db,
                int(saved["id"]),
                next_slot,
                used_slots,
                reuse_existing_slot=False,
            )
            if active is not None:
                stream_status = _start_stream_for_camera(active)
            else:
                message = (
                    f"Reachable, but no free camera slot is available right now "
                    f"({MAX_CAMERA_SLOTS} active slot limit reached). Deactivate "
                    "another camera to free one up for this one."
                )

        results.append(
            {
                "camera_id": saved["id"],
                "channel": stream.channel,
                "slot_number": assigned_slot,
                "status": test_result["status"],
                "message": message,
                "active": active is not None,
                "stream": stream_status,
            }
        )

    return results


@app.post("/api/discovery/connect")
def discovery_connect(request: DiscoveryConnectRequest) -> dict[str, Any]:
    """Complete the device-first flow: enumerate the chosen service's streams
    via the provider stack (ONVIF-first, vendor/generic fallback) and register
    them through the existing camera store + slot allocator + ffmpeg stream
    layer. The operator supplies only the device, the selected service, and -
    when required - credentials; never a stream path."""
    host = request.host.strip()
    try:
        resolve_and_guard(host)
    except DiscoveryHostError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    port = request.port or STREAM_DEFAULT_PORTS.get(request.protocol.lower(), 554)
    credentials = StreamCredentials(request.username or "", request.password or "")
    enumeration = enumerate_streams(
        host=host,
        port=port,
        protocol=request.protocol,
        credentials=credentials,
        vendor=request.vendor,
        channel_count=request.channel_count,
    )

    if enumeration.requires_auth and not request.username:
        raise HTTPException(
            status_code=401,
            detail="This service requires authentication. Provide a username and password.",
        )
    if not enumeration.channels:
        raise HTTPException(
            status_code=502,
            detail=_redact_sensitive_text(
                enumeration.error or "No connectable streams were found on this device."
            ),
        )

    db = _get_camera_db()
    results = _register_discovered_channels(
        name_prefix=request.name.strip(),
        channels=enumeration.channels,
        make_active=request.make_active,
        test_streams=request.test_streams,
        db=db,
    )

    if request.make_active and any(result["active"] for result in results):
        _sync_config_active_cameras(db)
        if _status()["running"]:
            stop_detection()
        try:
            start_detection(StartRequest())
        except HTTPException:
            pass

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    _audit(
        "discovery_connect",
        {"host": host, "provider": enumeration.provider, "channels": len(results)},
    )
    return {
        "provider": enumeration.provider,
        "results": results,
        "cameras": cameras,
        "active_cameras": active_cameras,
    }


@app.post("/api/camera-controller")
def save_camera_controller(controller: CameraControllerCreate) -> dict[str, Any]:
    db = _get_camera_db()
    registration = _register_controller_channels(controller, db)
    endpoint = registration["endpoint"]
    private_host_message = registration["private_host_message"]
    controller_error = registration["controller_error"]
    controller_reachable = registration["controller_reachable"]
    saved_cameras = registration["saved_cameras"]
    test_results = registration["test_results"]

    if controller.make_active and any(result["active"] for result in test_results):
        _sync_config_active_cameras(db)
        if _status()["running"]:
            stop_detection()
        # Always attempt a (re)start so newly added channels start
        # transmitting immediately instead of waiting on the next NVR
        # change to happen to restart detection. A freshly reconnected RTSP
        # stream can fail the stricter start-time stream check (e.g. no
        # keyframe yet) even though the endpoint itself is reachable; that
        # shouldn't fail this request since the cameras were already saved
        # and activated above — the watchdog retries shortly after.
        try:
            start_detection(StartRequest())
        except HTTPException:
            pass

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    return {
        "controller": {
            "name": controller.name.strip(),
            "host": endpoint["host"],
            "port": endpoint["port"],
            "protocol": endpoint["scheme"],
            "reachable": controller_reachable,
            "public_reachable_required": controller.require_public,
            "public_reachability_warning": private_host_message,
            "message": controller_error
            or f"Controller endpoint {endpoint['host']}:{endpoint['port']} is reachable.",
        },
        "created": saved_cameras,
        "results": test_results,
        "cameras": cameras,
        "active_cameras": active_cameras,
        "active_camera": active_cameras[0] if active_cameras else None,
    }


@app.post("/api/v2/channels/{channel_id}/stream/start")
def v2_start_stream(channel_id: int, request: V2StreamStartRequest | None = None) -> dict[str, Any]:
    request = request or V2StreamStartRequest()
    device_db = _get_device_db()
    channel = device_db.get_channel(channel_id, include_secret=True)
    if not channel:
        raise HTTPException(status_code=404, detail="Channel not found.")

    slot_number = request.slot_number or channel.get("slot_number")
    if slot_number is None and channel.get("camera_id"):
        camera = _get_camera_db().assign_slot(
            int(channel["camera_id"]),
            _next_available_slot(_get_camera_db().list_cameras(include_secret=False)),
        )
        slot_number = camera.get("slot_number") if camera else None
        _sync_config_active_cameras(_get_camera_db())

    status = _get_stream_manager().start(
        StreamSessionConfig(
            channel_id=str(channel_id),
            name=str(channel["name"]),
            source=str(channel["stream_reference"]),
            slot_number=slot_number,
            snapshot_dir=SNAPSHOT_DIR,
        )
    )
    device_db.update_stream_session(channel_id, status)
    return {"channel": device_db.get_channel(channel_id, include_secret=False), "stream": status}


@app.post("/api/v2/channels/{channel_id}/stream/stop")
def v2_stop_stream(channel_id: int) -> dict[str, Any]:
    stopped = _get_stream_manager().stop(str(channel_id))
    status = {"channel_id": str(channel_id), "status": "offline"}
    _get_device_db().update_stream_session(channel_id, status)
    return {"stopped": stopped, "stream": status}


@app.get("/api/v2/streams/health")
def v2_stream_health() -> dict[str, Any]:
    return _get_stream_manager().status()


@app.get("/api/v2/channels/{channel_id}/stream/routes")
def v2_channel_stream_routes(channel_id: int) -> dict[str, Any]:
    """Return internal AI plus browser WebRTC/HLS routes for one camera."""
    manager = _get_stream_manager()
    route = manager.route_info(str(channel_id))
    status = manager.status(str(channel_id))
    media_path = manager.media_client.path_status(route["dashboard_path"])
    return {
        "channel_id": str(channel_id),
        "transport": status.get("transport", "direct"),
        "routes": route,
        "health": status,
        "clients": len(media_path.get("readers") or []),
        "media_path": media_path,
    }


@app.post("/api/v2/channels/{channel_id}/stream/restart")
def v2_restart_stream(channel_id: int) -> dict[str, Any]:
    status = _get_stream_manager().restart(str(channel_id))
    if status.get("status") == "offline":
        raise HTTPException(status_code=404, detail="Stream is not active.")
    return {"restarted": True, "stream": status}


@app.get("/api/v2/channels/{channel_id}/stream/snapshot")
def v2_channel_stream_snapshot(channel_id: int) -> Response:
    data = _get_stream_manager().latest_frame_bytes(channel_id=str(channel_id))
    if data is None:
        raise HTTPException(status_code=503, detail="No fresh camera frame is available.")
    return Response(
        content=data,
        media_type="image/jpeg",
        headers={"Cache-Control": "no-store", "X-AI-Frame-Source": "stream-manager"},
    )


@app.get("/api/v2/channels/{channel_id}/recording/status")
def v2_channel_recording_status(channel_id: int) -> dict[str, Any]:
    manager = _get_stream_manager()
    route = manager.route_info(str(channel_id))
    source_path = manager.media_client.path_status(route["source_path"])
    return {
        "channel_id": str(channel_id),
        "enabled": bool(manager.media_client.enabled),
        "recording": bool(source_path.get("record")),
        "path": route["source_path"],
        "storage_root": "/recordings",
        "media_path": source_path,
    }


@app.get("/api/v2/channels/{channel_id}/stream/health")
def v2_channel_stream_health(channel_id: int) -> dict[str, Any]:
    return _get_stream_manager().status(str(channel_id))


@app.get("/api/v2/channels/{channel_id}/live")
async def v2_channel_live_frame(channel_id: int):
    channel = _get_device_db().get_channel(channel_id, include_secret=False)
    if not channel:
        raise HTTPException(status_code=404, detail="Channel not found.")
    return await live_frame(slot=channel.get("slot_number"))


@app.post("/api/v2/channels/{channel_id}/analytics/start")
def v2_start_analytics(channel_id: int) -> dict[str, Any]:
    channel = _get_device_db().set_analytics_enabled(channel_id, True)
    if not channel:
        raise HTTPException(status_code=404, detail="Channel not found.")
    if _detector_pid() is None:
        try:
            start_detection(StartRequest())
        except HTTPException as exc:
            if exc.status_code != 409:
                raise
    return {"channel": channel, "analytics": _status()}


@app.post("/api/v2/channels/{channel_id}/analytics/stop")
def v2_stop_analytics(channel_id: int) -> dict[str, Any]:
    channel = _get_device_db().set_analytics_enabled(channel_id, False)
    if not channel:
        raise HTTPException(status_code=404, detail="Channel not found.")
    return {"channel": channel, "analytics": _status()}


@app.get("/api/v2/analytics/health")
def v2_analytics_health() -> dict[str, Any]:
    status = _status()
    return {
        "status": "online" if status.get("running") else "offline",
        "pid": status.get("pid"),
        "health": status.get("health"),
    }


@app.get("/api/v2/detections/latest")
def v2_latest_detections() -> dict[str, Any]:
    health = _read_json(DETECTION_HEALTH_PATH) or {}
    return {
        "state": health.get("state", "offline"),
        "updated_at": health.get("updated_at"),
        "cameras": health.get("cameras") or [],
        "detections": health.get("last_detections_by_camera") or {},
        "spatial": health.get("last_spatial_objects_by_camera")
        or health.get("last_spatial_objects")
        or {},
    }


@app.post("/api/cameras/{camera_id}/test")
def test_saved_camera(camera_id: int) -> dict[str, Any]:
    db = _get_camera_db()
    camera = db.get_camera(camera_id, include_secret=True)
    if camera is None:
        raise HTTPException(status_code=404, detail="Camera not found.")

    result = _test_camera_stream(camera["stream_url"])
    updated = db.set_status(camera_id, result["status"])
    return {"camera": updated, "test": result}


@app.delete("/api/cameras/{camera_id}")
def delete_saved_camera(camera_id: int) -> dict[str, Any]:
    db = _get_camera_db()
    deleted = db.delete_camera(camera_id)
    if deleted:
        try:
            _sync_config_active_cameras(db)
        except Exception as exc:
            _audit(
                "camera_config_sync_failed",
                {"camera_id": camera_id, "error": _redact_sensitive_text(str(exc))},
            )

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    return {
        "deleted": deleted,
        "already_absent": not deleted,
        "cameras": cameras,
        "active_cameras": active_cameras,
        "active_camera": active_cameras[0] if active_cameras else None,
    }


@app.post("/api/cameras/cleanup")
def cleanup_cameras_by_name(request: CameraCleanupRequest) -> dict[str, Any]:
    """Bulk-delete every camera whose name starts with the given prefix.

    For clearing out stale rows left behind by the environment-based boot
    seed (see _seed_cameras_from_environment) - those get created once with
    whatever stream URL was correct/available at the time, stored as a
    literal string, and never retroactively updated by a later code or env
    var change. A camera named e.g. "Warehouse NVR Camera 7" with a wrong
    URL baked in from months ago has to be deleted and re-added, not
    patched.
    """
    prefix = request.name_prefix.strip()
    if not prefix:
        raise HTTPException(status_code=400, detail="name_prefix is required.")

    db = _get_camera_db()
    matches = [
        camera
        for camera in db.list_cameras(include_secret=False)
        if str(camera.get("name", "")).startswith(prefix)
    ]

    deleted = []
    any_active = False
    for camera in matches:
        if camera.get("is_active"):
            any_active = True
        if db.delete_camera(camera["id"]):
            deleted.append({"id": camera["id"], "name": camera["name"]})

    if any_active:
        _sync_config_active_cameras(db)
        if _status()["running"]:
            stop_detection()
        try:
            start_detection(StartRequest())
        except HTTPException:
            pass

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    return {
        "deleted": deleted,
        "deleted_count": len(deleted),
        "cameras": cameras,
        "active_cameras": active_cameras,
    }


@app.post("/api/cameras/{camera_id}/activate")
def set_active_camera(
    camera_id: int, request: CameraSlotRequest | None = None
) -> dict[str, Any]:
    db = _get_camera_db()
    request = request or CameraSlotRequest()
    active = db.assign_slot(camera_id, request.slot_number)
    if active is None:
        raise HTTPException(status_code=404, detail="Camera not found.")

    _sync_config_active_cameras(db)
    stream = _start_stream_for_camera(active)
    restarted = False
    if _status()["running"]:
        stop_detection()
        start_detection(StartRequest())
        restarted = True

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    return {
        "active_camera": db.get_camera(camera_id, include_secret=False),
        "active_cameras": active_cameras,
        "cameras": cameras,
        "stream": stream,
        "restarted": restarted,
    }


@app.delete("/api/camera-slots/{slot_number}")
def clear_camera_slot(slot_number: int) -> dict[str, Any]:
    if slot_number < 1 or slot_number > MAX_CAMERA_SLOTS:
        raise HTTPException(
            status_code=400,
            detail=f"Slot number must be between 1 and {MAX_CAMERA_SLOTS}.",
        )

    db = _get_camera_db()
    active_before = [
        camera for camera in db.list_active_cameras(include_secret=False)
        if camera.get("slot_number") == slot_number
    ]
    db.clear_slot(slot_number)
    for camera in active_before:
        _get_stream_manager().stop(str(camera["id"]))
    _sync_config_active_cameras(db)
    restarted = False
    if _status()["running"]:
        stop_detection()
        start_detection(StartRequest())
        restarted = True

    cameras = db.list_cameras(include_secret=False)
    active_cameras = [row for row in cameras if row["is_active"]]
    return {
        "active_camera": active_cameras[0] if active_cameras else None,
        "active_cameras": active_cameras,
        "cameras": cameras,
        "restarted": restarted,
    }


@app.post("/api/start")
def start_detection(request: StartRequest | None = None) -> dict[str, Any]:
    global _process, _started_at, _last_exit_code, _stdout_handle, _stderr_handle, _manual_stop_requested
    request = request or StartRequest()
    if _detector_pid() is not None:
        # Starting an already-running persistent service is a successful no-op.
        # The dashboard may repeat this request after reconnects, navigation, or
        # React remounts; treating that as a conflict creates a noisy 409 even
        # though the requested state has already been reached.
        status = _status()
        status["already_running"] = True
        return status
    # Clear the manual-stop latch as soon as a start is attempted, not after
    # validation succeeds. Otherwise a start that's triggered right after a
    # stop (e.g. restarting to pick up a newly added camera) and then fails
    # validation (a stream that hasn't sent its first keyframe yet, briefly
    # unreachable, etc.) leaves _manual_stop_requested stuck True, which
    # permanently blocks the watchdog from ever retrying.
    _manual_stop_requested = False
    # Treat the camera database as the source of truth. This prevents a stale
    # config/config.yaml (for example the demo camera checked into the repo) from
    # making the detector process only slot 1 while the dashboard has many active
    # NVR/controller channels saved in SQLite.
    _sync_config_active_cameras(_get_camera_db())
    stream_status = _ensure_streams_from_active_cameras()
    _validate_active_cameras_for_start()

    DETECTION_STDOUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    _stdout_handle = DETECTION_STDOUT_PATH.open("w", encoding="utf-8", buffering=1)
    _stderr_handle = DETECTION_STDERR_PATH.open("w", encoding="utf-8", buffering=1)
    _stdout_handle.write(f"\n--- detection start {_now_iso()} config={request.config_path} ---\n")
    DETECTION_HEALTH_PATH.write_text(
        json.dumps(
            {
                "state": "starting",
                "error": None,
                "frames_read": 0,
                "last_frame_at": None,
                "last_detection_count": 0,
                "last_tracked_count": 0,
                "stream_manager": stream_status,
                "updated_at": _now_iso(),
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    command = [
        sys.executable,
        str(ROOT / "main.py"),
        "--config",
        request.config_path,
    ]
    if request.no_display:
        command.append("--no-display")

    _process = subprocess.Popen(
        command,
        cwd=ROOT,
        stdout=_stdout_handle,
        stderr=_stderr_handle,
        env={**os.environ, "PYTHONUNBUFFERED": "1", "AI_VISION_STREAM_FIRST": "1"},
        start_new_session=os.name != "nt",
    )
    _started_at = time.time()
    _last_exit_code = None
    _write_detector_pid(_process.pid)
    return _status()


@app.post("/api/stop")
def stop_detection() -> dict[str, Any]:
    global _process, _started_at, _last_exit_code, _stdout_handle, _stderr_handle, _manual_stop_requested
    _manual_stop_requested = True
    process = _process
    pid = _detector_pid()
    if pid is None:
        return _status()

    if process is None:
        _last_exit_code = _terminate_pid(pid)
    elif os.name == "nt":
        _terminate_pid(process.pid)
        try:
            process.wait(timeout=10)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait(timeout=5)
        _last_exit_code = process.returncode
    else:
        os.killpg(process.pid, signal.SIGTERM)

        try:
            process.wait(timeout=10)
        except subprocess.TimeoutExpired:
            os.killpg(process.pid, signal.SIGKILL)
            process.wait(timeout=5)
        _last_exit_code = process.returncode

    _process = None
    _started_at = None
    _clear_detector_pid()
    DETECTION_HEALTH_PATH.write_text(
        json.dumps(
            {
                "state": "stopped",
                "error": None,
                "frames_read": 0,
                "last_frame_at": None,
                "last_detection_count": 0,
                "last_tracked_count": 0,
                "updated_at": _now_iso(),
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    for handle in (_stdout_handle, _stderr_handle):
        if handle is not None:
            handle.close()
    _stdout_handle = None
    _stderr_handle = None
    return _status()


@app.post("/api/restart")
def restart_detection(request: StartRequest | None = None) -> dict[str, Any]:
    global _manual_stop_requested
    _manual_stop_requested = False
    stop_detection()
    _manual_stop_requested = False
    return start_detection(request)


@app.get("/api/logs")
def recent_logs(limit: int = 80) -> dict[str, Any]:
    if not LOG_PATH.exists():
        return {"lines": []}

    lines = LOG_PATH.read_text(encoding="utf-8", errors="replace").splitlines()
    return {"lines": lines[-max(1, min(limit, 500)) :]}


@app.get("/api/detection/logs")
def detection_logs(limit: int = 120) -> dict[str, Any]:
    return {
        "health": _read_json(DETECTION_HEALTH_PATH),
        "stdout": _tail_file(DETECTION_STDOUT_PATH, limit),
        "stderr": _tail_file(DETECTION_STDERR_PATH, limit),
    }


@app.get("/api/snapshots")
def snapshots(limit: int = 24) -> dict[str, Any]:
    if not SNAPSHOT_DIR.exists():
        return {"snapshots": []}

    files = sorted(
        SNAPSHOT_DIR.glob("*.jpg"), key=lambda path: path.stat().st_mtime, reverse=True
    )
    return {
        "snapshots": [
            {
                "name": path.name,
                "url": f"/snapshots/{path.name}",
                "modified_at": path.stat().st_mtime,
            }
            for path in files[: max(1, min(limit, 100))]
        ]
    }


@app.get("/api/occupancy")
def occupancy(camera: str | None = None) -> dict[str, Any]:
    """Currently checked-in tracked objects (from ByteTrack + SQLite),
    plus per-class counts. Distinct from /api/inventory, which is the
    manually-operated warehouse item ledger."""
    db = _get_tracking_db()
    current = db.current_occupancy(camera_name=camera)
    counts = db.occupancy_counts(camera_name=camera)
    return {
        "current": current,
        "counts": [
            {"class_name": name, "count": count}
            for name, count in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
        ],
    }


@app.get("/api/occupancy/events")
def occupancy_events(limit: int = 50, camera: str | None = None) -> dict[str, Any]:
    """Recent check-in / check-out events, most recent first."""
    db = _get_tracking_db()
    events = db.recent_events(limit=max(1, min(limit, 500)), camera_name=camera)
    return {"events": events}


@app.get("/api/inventory")
def inventory() -> dict[str, Any]:
    data = _ensure_inventory()
    return {"items": data["items"], "history": data["history"]}


@app.post("/api/inventory/item")
def add_inventory_item(item: ItemCreate) -> dict[str, Any]:
    data = _ensure_inventory()
    if _find_item(data, item.item_id):
        raise HTTPException(status_code=409, detail="Item ID already exists.")

    record = {
        "item_id": item.item_id,
        "name": item.name,
        "item_type": item.item_type or "unknown",
        "quantity": 0,
        "created_at": _now_iso(),
        "last_updated_at": _now_iso(),
    }
    data["items"].append(record)
    _save_inventory(data)
    return record


@app.post("/api/inventory/checkin")
def inventory_checkin(action: InventoryAction) -> dict[str, Any]:
    data = _ensure_inventory()
    item = _find_item(data, action.item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found.")

    item["quantity"] += action.quantity
    item["last_updated_at"] = _now_iso()
    _record_inventory_event(data, "check-in", action.item_id, action.quantity, action.note)
    _save_inventory(data)
    return item


@app.post("/api/inventory/checkout")
def inventory_checkout(action: InventoryAction) -> dict[str, Any]:
    data = _ensure_inventory()
    item = _find_item(data, action.item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found.")
    if action.quantity > item["quantity"]:
        raise HTTPException(status_code=400, detail="Insufficient quantity for checkout.")

    item["quantity"] -= action.quantity
    item["last_updated_at"] = _now_iso()
    _record_inventory_event(data, "check-out", action.item_id, action.quantity, action.note)
    _save_inventory(data)
    return item


@app.post("/api/inventory/upload-image")
async def upload_inventory_image(item_id: str = Form(...), file: UploadFile = File(...)) -> dict[str, Any]:
    INVENTORY_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    item = _find_item(_ensure_inventory(), item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found.")

    filename = f"{item_id}_{int(time.time())}_{file.filename}"
    path = INVENTORY_IMAGE_DIR / filename
    contents = await file.read()
    path.write_bytes(contents)
    return {"url": f"/snapshots/inventory/{filename}", "name": filename}


@app.get("/api/catalog/items")
def catalog_items(scope_id: str) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    db = _get_catalog_db()
    items = db.list_items(scope)
    for item in items:
        item["detection_prompts"] = _catalog_item_prompts(scope, str(item["id"]))
    return {
        "items": items,
        "schedule": _catalog_schedule(scope),
        "latest_run": db.latest_run(scope),
    }


@app.post("/api/catalog/learning/start")
async def start_product_learning(
    scope_id: str, request: ProductLearningStart
) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    active = next(
        (
            session
            for session in _product_learning_sessions.values()
            if session.get("scope_id") == scope
            and session.get("status") in {"capturing", "processing"}
        ),
        None,
    )
    if active:
        raise HTTPException(status_code=409, detail="A product learning session is already active.")
    session_id = secrets.token_urlsafe(18)
    session = {
        "session_id": session_id,
        "scope_id": scope,
        "status": "capturing",
        "duration_seconds": request.duration_seconds,
        "camera_name": _catalog_camera_label(request.camera_name),
        "remaining_seconds": request.duration_seconds,
        "started_at": _now_iso(),
        "completed_at": None,
        "camera_count": 0,
        "frames_seen": 0,
        "proposal_count": 0,
        "view_count": 0,
        "views": [],
        "error": None,
    }
    _product_learning_sessions[session_id] = session
    task = asyncio.create_task(asyncio.to_thread(_run_product_learning_session, session_id))
    _product_learning_tasks[session_id] = task
    _audit("product_learning_started", {"scope_id": scope, "session_id": session_id})
    return _product_learning_public(session)


@app.get("/api/catalog/learning/{session_id}")
def product_learning_status(session_id: str, scope_id: str) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    session = _product_learning_sessions.get(session_id)
    if not session or session.get("scope_id") != scope:
        raise HTTPException(status_code=404, detail="Product learning session not found.")
    return _product_learning_public(session)


@app.post("/api/catalog/learning/save")
def save_learned_product(scope_id: str, request: ProductLearningSave) -> dict[str, Any]:
    import cv2

    scope = _catalog_scope(scope_id)
    session = _product_learning_sessions.get(request.session_id)
    if not session or session.get("scope_id") != scope:
        raise HTTPException(status_code=404, detail="Product learning session not found.")
    if session.get("status") != "ready" or len(session.get("_views") or []) < 1:
        raise HTTPException(status_code=409, detail="Product learning has not completed.")
    product_name = " ".join(request.product_name.split()).strip()
    db = _get_catalog_db()
    all_views = session["_views"]
    indices = sorted(set(int(index) for index in request.view_indices))
    if any(index < 0 or index >= len(all_views) for index in indices):
        raise HTTPException(status_code=400, detail="One or more selected views are invalid.")
    views = [all_views[index] for index in indices]
    if len(views) < 1:
        raise HTTPException(status_code=400, detail="Select at least one product view.")
    if request.existing_item_id:
        item = db.get_item(request.existing_item_id)
        if not item or str(item.get("scope_id")) != scope:
            raise HTTPException(status_code=404, detail="Existing catalog product not found.")
        product_name = str(item["name"])
    else:
        if any(
            _catalog_normalize_name(item["name"]) == _catalog_normalize_name(product_name)
            for item in db.list_items(scope)
        ):
            raise HTTPException(
                status_code=409,
                detail="This product already exists. Select the suggested existing product instead.",
            )
        item = db.create_item(scope, product_name)

    item_dir = CATALOG_IMAGE_DIR / scope / str(item["id"])
    item_dir.mkdir(parents=True, exist_ok=True)
    stamp = int(time.time())
    for index, view in enumerate(views, start=1):
        filename = f"learned_{stamp}_{index:02d}.jpg"
        path = item_dir / filename
        if not cv2.imwrite(str(path), view["crop"]):
            continue
        frame = view["crop"]
        db.add_image(
            item_id=str(item["id"]),
            filename=filename,
            url=f"/snapshots/catalog/{quote(scope)}/{quote(str(item['id']))}/{quote(filename)}",
            embedding=view["embedding"],
            width_px=int(frame.shape[1]),
            height_px=int(frame.shape[0]),
        )

    prompts = _catalog_save_item_prompts(
        scope,
        str(item["id"]),
        [product_name, *_catalog_item_prompts(scope, str(item["id"]))],
    )
    global _catalog_yolo_detector, _catalog_yolo_detector_key
    _catalog_yolo_detector = None
    _catalog_yolo_detector_key = None
    fingerprints = _read_json(PRODUCT_FINGERPRINTS_PATH) or {"products": {}}
    previous_fingerprint = fingerprints.setdefault("products", {}).get(str(item["id"])) or {}
    fingerprints["products"][str(item["id"])] = {
        "product_id": str(item["id"]),
        "product_name": product_name,
        "scope_id": scope,
        "created_at": previous_fingerprint.get("created_at") or _now_iso(),
        "last_updated_at": _now_iso(),
        "matching_confidence": float(
            os.getenv("CATALOG_PROPOSAL_SIMILARITY_THRESHOLD", "0.62")
        ),
        "learning_statistics": {
            "duration_seconds": session["duration_seconds"],
            "camera_count": session["camera_count"],
            "frames_seen": session["frames_seen"],
            "proposal_count": session["proposal_count"],
            "reference_count": len(db.list_images(str(item["id"]))),
            "latest_learning_reference_count": len(views),
            "updated_existing_product": bool(request.existing_item_id),
        },
        "detection_prompts": prompts,
    }
    PRODUCT_FINGERPRINTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    PRODUCT_FINGERPRINTS_PATH.write_text(
        json.dumps(fingerprints, indent=2), encoding="utf-8"
    )
    session["status"] = "saved"
    session["product_id"] = str(item["id"])
    session["product_name"] = product_name
    _audit(
        "product_learning_saved",
        {
            "scope_id": scope,
            "session_id": request.session_id,
            "item_id": item["id"],
            "name": product_name,
            "reference_count": len(views),
        },
    )
    return {
        "item": db.get_item(str(item["id"])),
        "fingerprint": fingerprints["products"][str(item["id"])],
        "active": True,
    }


@app.put("/api/catalog/items/{item_id}/prompts")
def update_catalog_item_prompts(
    item_id: str, scope_id: str, update: CatalogPromptUpdate
) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    item = _get_catalog_db().get_item(item_id)
    if not item or item["scope_id"] != scope:
        raise HTTPException(status_code=404, detail="Catalog item not found.")
    prompts = _catalog_save_item_prompts(scope, item_id, update.prompts)
    global _catalog_yolo_detector, _catalog_yolo_detector_key
    _catalog_yolo_detector = None
    _catalog_yolo_detector_key = None
    _audit(
        "catalog_prompts_updated",
        {"scope_id": scope, "item_id": item_id, "prompts": prompts},
    )
    return {"item_id": item_id, "prompts": prompts}


@app.post("/api/catalog/results/correct")
async def correct_catalog_result(
    scope_id: str, correction: CatalogCorrection
) -> dict[str, Any]:
    """Teach the catalog from a human correction on the Result Analytics page.

    The operator confirms what an object really is; the corrected crop is saved
    as a reference image for that item (creating the item if needed) and the
    optional description is stored as a recognition prompt. Future recognition
    then matches this object to the correct name instead of the mislabel.
    """
    import cv2
    import numpy as np
    from recognition.embedding import image_embedding

    scope = _catalog_scope(scope_id)
    correct_name = " ".join(correction.correct_name.split()).strip()
    if not correct_name:
        raise HTTPException(status_code=400, detail="A correct name is required.")

    crop_path = _catalog_snapshot_path(correction.crop_url)
    if crop_path is None or not crop_path.exists():
        raise HTTPException(status_code=404, detail="Crop image not found for this result.")
    contents = crop_path.read_bytes()
    frame = cv2.imdecode(np.frombuffer(contents, dtype=np.uint8), cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(status_code=400, detail="Crop image could not be decoded.")

    db = _get_catalog_db()
    item = next(
        (
            existing
            for existing in db.list_items(scope)
            if _catalog_normalize_name(existing["name"]) == _catalog_normalize_name(correct_name)
        ),
        None,
    )
    created = item is None
    if item is None:
        item = db.create_item(scope, correct_name)
    item_id = str(item["id"])

    item_dir = CATALOG_IMAGE_DIR / scope / item_id
    item_dir.mkdir(parents=True, exist_ok=True)
    existing_images = db.list_images(item_id)
    filename = f"correction_{len(existing_images) + 1:02d}.jpg"
    (item_dir / filename).write_bytes(contents)
    url = f"/snapshots/catalog/{quote(scope)}/{quote(item_id)}/{quote(filename)}"
    db.add_image(
        item_id=item_id,
        filename=filename,
        url=url,
        embedding=image_embedding(frame),
        width_px=int(frame.shape[1]),
        height_px=int(frame.shape[0]),
    )

    prompts = _catalog_item_prompts(scope, item_id)
    if correction.prompt:
        cleaned_prompt = " ".join(correction.prompt.split()).strip()
        existing_norm = {_catalog_normalize_name(value) for value in prompts}
        if cleaned_prompt and _catalog_normalize_name(cleaned_prompt) not in existing_norm:
            prompts = _catalog_save_item_prompts(scope, item_id, [*prompts, cleaned_prompt])

    global _catalog_yolo_detector, _catalog_yolo_detector_key
    _catalog_yolo_detector = None
    _catalog_yolo_detector_key = None

    _audit(
        "catalog_result_corrected",
        {
            "scope_id": scope,
            "item_id": item_id,
            "correct_name": correct_name,
            "predicted_name": correction.predicted_name,
            "item_created": created,
        },
    )
    return {
        "item": db.get_item(item_id),
        "item_created": created,
        "reference_count": len(db.list_images(item_id)),
        "prompts": prompts,
    }


TRAINING_DATASET_ROOT = ROOT / "datasets" / "baget_box"
TRAINING_DATASET_YAML = TRAINING_DATASET_ROOT / "data.yaml"
TRAINING_PROMPTS_PATH = TRAINING_DATASET_ROOT / "prompts.json"
TRAINING_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}


_training_detector_obj = None
_training_detector_key = None
_baget_detector_obj = None
_baget_detector_key = None


def _training_detector(class_prompts: list[str] | None = None):
    """Light, cached detector for the training tools - uses config detection
    settings (conf/imgsz) instead of the heavy 1280px/0.01-conf catalog one.

    Uses the light config settings and runs off the event loop, so it stays
    responsive. Set TRAINING_USE_MODEL=0 to force the model-free fallback (edge
    proposals for uploads, running-detector output for live) if the box is too
    small to hold a second model copy.
    """
    if os.getenv("TRAINING_USE_MODEL", "1").strip().lower() not in {"1", "true", "yes", "on"}:
        return None
    global _training_detector_obj, _training_detector_key
    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    det = config.get("detection", {}) or {}
    prompts = list(det.get("class_prompts") or []) if class_prompts is None else list(class_prompts)
    key = (
        det.get("model_path"), tuple(prompts),
        float(det.get("confidence_threshold", 0.4)),
        str(det.get("device", "auto")), int(det.get("image_size", 640)),
    )
    if _training_detector_obj is not None and _training_detector_key == key:
        return _training_detector_obj
    try:
        _training_detector_obj = Detector(
            model_path=str(det.get("model_path") or "yoloe-26s-seg.pt"),
            confidence_threshold=float(det.get("confidence_threshold", 0.4)),
            device=str(det.get("device", "auto")), class_prompts=prompts or None,
            image_size=int(det.get("image_size", 640)),
            class_agnostic_nms=bool(det.get("class_agnostic_nms", False)),
            fallback_model_path=det.get("fallback_model_path"),
        )
        _training_detector_key = key
        return _training_detector_obj
    except Exception as exc:  # noqa: BLE001
        _audit("training_detector_failed", {"error": str(exc)})
        _training_detector_obj = None
        _training_detector_key = None
        return None


def _baget_inventory_detector():
    """Load only the explicit factory Baget detector; never use generic fallback."""
    global _baget_detector_obj, _baget_detector_key
    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    inventory = config.get("inventory", {}) or {}
    active = _read_json(ACTIVE_MODELS_PATH) or {}
    model_path = str(
        (active.get("baget_box") or {}).get("weights_path")
        or os.getenv("BAGET_MODEL_PATH")
        or inventory.get("baget_model_path")
        or "models/baget_box_best.pt"
    )
    confidence = float(inventory.get("minimum_confidence", 0.5))
    key = (model_path, confidence, str(inventory.get("device", "auto")))
    if _baget_detector_obj is not None and _baget_detector_key == key:
        return _baget_detector_obj
    candidates = [Path(model_path), ROOT / model_path, ROOT / "models" / Path(model_path).name]
    resolved = next((path for path in candidates if path.exists()), None)
    if resolved is None:
        raise RuntimeError("BAGET DETECTOR UNAVAILABLE")
    detector = Detector(
        model_path=str(resolved), confidence_threshold=confidence,
        device=str(inventory.get("device", "auto")),
        image_size=int(inventory.get("image_size", 960)),
        iou_threshold=float(inventory.get("detector_iou", 0.55)),
        allow_fallback=False, detector_mode="baget_box_custom",
    )
    if not detector.health().get("ready"):
        raise RuntimeError("BAGET DETECTOR UNAVAILABLE")
    _baget_detector_obj, _baget_detector_key = detector, key
    return detector


def _deployed_product_detector(product_name: str):
    """Resolve an explicitly deployed custom detector for any normalized product."""
    key_name = re.sub(r"[^a-z0-9]+", "_", product_name.strip().lower()).strip("_")
    deployment = ((_read_json(ACTIVE_MODELS_PATH) or {}).get(key_name) or {})
    model_path = str(deployment.get("weights_path") or "")
    if not model_path:
        return None
    cache_key = (key_name, model_path)
    if cache_key in _active_product_detectors:
        return _active_product_detectors[cache_key]
    path = Path(model_path)
    if not path.exists():
        raise RuntimeError(f"{key_name.upper()} DETECTOR UNAVAILABLE")
    detector = Detector(
        model_path=str(path), confidence_threshold=0.5, device="auto",
        image_size=960, iou_threshold=0.55, allow_fallback=False,
        detector_mode=f"{key_name}_custom",
    )
    if not detector.health().get("ready"):
        raise RuntimeError(f"{key_name.upper()} DETECTOR UNAVAILABLE")
    _active_product_detectors[cache_key] = detector
    return detector


# --- Dataset durability: in-volume backups + auto-restore ---------------------
# The dataset lives on the persistent `ai-vision-datasets` Docker volume, so it
# already survives redeploys (git reset / compose up). These backups add a
# second line of defence against the working tree being cleared: snapshots kept
# inside the same volume, plus auto-restore on boot if the tree comes up empty.
# The user can also download/upload the whole dataset for an off-server copy.
TRAINING_BACKUP_KEEP = 15
_training_last_backup = 0.0


def _training_backup_dir() -> Path:
    """Backups live next to the dataset root (…/datasets/_backups) so they sit
    on the same persistent volume and always track the dataset's location."""
    return TRAINING_DATASET_ROOT.parent / "_backups"


def _training_dataset_has_images() -> bool:
    for split in ("train", "val"):
        img_dir = TRAINING_DATASET_ROOT / "images" / split
        if img_dir.exists() and any(
            p.suffix.lower() in TRAINING_IMAGE_EXTS for p in img_dir.iterdir()
        ):
            return True
    return False


def _training_backup_dataset(reason: str = "", *, min_interval: float = 120.0, force: bool = False):
    """Snapshot the dataset into a tar.gz inside the persistent volume so an
    accidental wipe of the working tree can be recovered. Throttled so rapid
    edits don't cause a backup storm; keeps the newest TRAINING_BACKUP_KEEP.
    Never raises - a failed backup must not break a save."""
    global _training_last_backup
    try:
        now = time.time()
        if not force and (now - _training_last_backup) < min_interval:
            return None
        if not _training_dataset_has_images():
            return None
        import tarfile

        backup_dir = _training_backup_dir()
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        archive = backup_dir / f"baget_box-{stamp}.tar.gz"
        with tarfile.open(archive, "w:gz") as tar:
            tar.add(TRAINING_DATASET_ROOT, arcname="baget_box")
        _training_last_backup = now
        for old in sorted(backup_dir.glob("baget_box-*.tar.gz"))[:-TRAINING_BACKUP_KEEP]:
            old.unlink()
        _audit("training_dataset_backup", {"archive": archive.name, "reason": reason})
        return archive
    except Exception as exc:  # noqa: BLE001
        _audit("training_dataset_backup_failed", {"error": str(exc)})
        return None


def _training_restore_from_backup_if_empty() -> bool:
    """If the dataset tree has no images but a backup exists, restore the newest
    one. Runs on boot so a cleared/reset working tree self-heals. Never raises."""
    try:
        if _training_dataset_has_images():
            return False
        backup_dir = _training_backup_dir()
        backups = sorted(backup_dir.glob("baget_box-*.tar.gz")) if backup_dir.exists() else []
        if not backups:
            return False
        import tarfile

        newest = backups[-1]
        with tarfile.open(newest, "r:gz") as tar:
            try:
                tar.extractall(TRAINING_DATASET_ROOT.parent, filter="data")
            except TypeError:
                # Older Python without the extraction filter argument.
                tar.extractall(TRAINING_DATASET_ROOT.parent)
        _audit("training_dataset_restored", {"archive": newest.name})
        return True
    except Exception as exc:  # noqa: BLE001
        _audit("training_dataset_restore_failed", {"error": str(exc)})
        return False


def _ensure_training_dataset() -> None:
    """Make sure the dataset dirs + data.yaml exist. On a fresh persistent
    volume (first boot after the dataset was added to docker-compose) the tree
    is empty, so seed the baseline classes and the split folders. If a backup
    exists it is restored first. Idempotent and never raises - training tools
    must degrade gracefully."""
    try:
        import yaml as _yaml

        # Recover a wiped working tree from the newest in-volume backup before
        # falling back to seeding an empty baseline.
        _training_restore_from_backup_if_empty()

        for sub in ("images/train", "images/val", "labels/train", "labels/val"):
            (TRAINING_DATASET_ROOT / sub).mkdir(parents=True, exist_ok=True)
        if not TRAINING_DATASET_YAML.exists():
            TRAINING_DATASET_YAML.write_text(
                _yaml.safe_dump(
                    {
                        "path": "datasets/baget_box",
                        "train": "images/train",
                        "val": "images/val",
                        "names": {0: "baget_box_stack_individual", 1: "sack"},
                    },
                    sort_keys=False,
                ),
                encoding="utf-8",
            )
        else:
            # Preserve every existing class id while migrating the original
            # Baget Box class name. Existing class-0 labels remain valid.
            data = _yaml.safe_load(TRAINING_DATASET_YAML.read_text(encoding="utf-8")) or {}
            names = data.get("names") or {}
            class_zero = names.get(0, names.get("0"))
            normalized = " ".join(str(class_zero or "").replace("_", " ").split()).lower()
            if normalized == "baget box":
                if 0 in names:
                    names[0] = "baget_box_stack_individual"
                else:
                    names["0"] = "baget_box_stack_individual"
                data["names"] = names
                TRAINING_DATASET_YAML.write_text(
                    _yaml.safe_dump(data, sort_keys=False), encoding="utf-8"
                )
    except Exception as exc:  # noqa: BLE001
        _audit("training_dataset_seed_failed", {"error": str(exc)})


def _training_dataset_stats() -> dict[str, Any]:
    import yaml as _yaml

    _ensure_training_dataset()
    names: dict[Any, Any] = {}
    if TRAINING_DATASET_YAML.exists():
        data = _yaml.safe_load(TRAINING_DATASET_YAML.read_text(encoding="utf-8")) or {}
        names = data.get("names") or {}
    splits: dict[str, Any] = {}
    class_instances: dict[str, int] = {}
    for split in ("train", "val"):
        img_dir = TRAINING_DATASET_ROOT / "images" / split
        lbl_dir = TRAINING_DATASET_ROOT / "labels" / split
        images = (
            [p for p in img_dir.iterdir() if p.suffix.lower() in TRAINING_IMAGE_EXTS]
            if img_dir.exists()
            else []
        )
        labeled = negatives = 0
        for image in images:
            label = lbl_dir / f"{image.stem}.txt"
            if not label.exists():
                continue
            content = label.read_text(encoding="utf-8").strip()
            if not content:
                negatives += 1
                continue
            labeled += 1
            for line in content.splitlines():
                parts = line.split()
                if parts:
                    class_instances[parts[0]] = class_instances.get(parts[0], 0) + 1
        splits[split] = {
            "images": len(images),
            "labeled": labeled,
            "negatives": negatives,
            "unlabeled": len(images) - labeled - negatives,
        }
    store = _read_json(TRAINING_PROMPTS_PATH) or {}
    prompts = store.get("prompts") if isinstance(store, dict) else []
    return {
        "classes": names,
        "splits": splits,
        "class_instances": class_instances,
        "prompts": prompts if isinstance(prompts, list) else [],
    }


@app.get("/api/training/dataset")
def training_dataset() -> dict[str, Any]:
    return _training_dataset_stats()


@app.get("/api/training/dataset/export")
def training_dataset_export() -> Response:
    """Download the whole dataset as a tar.gz so an off-server copy can be kept
    (the firmest protection against losing it). Also refreshes the newest
    in-volume backup."""
    import io as _io
    import tarfile

    _ensure_training_dataset()
    _training_backup_dataset("export", force=True)
    buffer = _io.BytesIO()
    with tarfile.open(fileobj=buffer, mode="w:gz") as tar:
        if TRAINING_DATASET_ROOT.exists():
            tar.add(TRAINING_DATASET_ROOT, arcname="baget_box")
    buffer.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return Response(
        content=buffer.getvalue(),
        media_type="application/gzip",
        headers={"Content-Disposition": f'attachment; filename="baget_box-dataset-{stamp}.tar.gz"'},
    )


@app.post("/api/training/dataset/import")
async def training_dataset_import(file: UploadFile = File(...)) -> dict[str, Any]:
    """Restore the dataset from a previously exported tar.gz. Backs up whatever
    is there first, then merges the archive's contents in (never deletes)."""
    import io as _io
    import tarfile

    contents = await file.read()
    _training_backup_dataset("pre-import", force=True)
    try:
        with tarfile.open(fileobj=_io.BytesIO(contents), mode="r:gz") as tar:
            members = [m for m in tar.getmembers() if m.name.startswith("baget_box")]
            if not members:
                raise HTTPException(status_code=400, detail="Archive does not contain a baget_box dataset.")
            try:
                tar.extractall(TRAINING_DATASET_ROOT.parent, members=members, filter="data")
            except TypeError:
                tar.extractall(TRAINING_DATASET_ROOT.parent, members=members)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read archive: {exc}") from exc
    _ensure_training_dataset()
    _training_backup_dataset("post-import", force=True)
    _audit("training_dataset_import", {"bytes": len(contents)})
    return {"restored": True, "dataset": _training_dataset_stats()}


@app.post("/api/training/inject")
async def training_inject(
    split: str = Form("train"),
    prompts: str = Form(""),
    files: list[UploadFile] = File(default=[]),
) -> dict[str, Any]:
    """Save-and-Inject: drop images + prompts straight into the YOLO dataset."""
    import cv2
    import numpy as np

    split = "val" if str(split).strip().lower().startswith("v") else "train"
    img_dir = TRAINING_DATASET_ROOT / "images" / split
    img_dir.mkdir(parents=True, exist_ok=True)
    (TRAINING_DATASET_ROOT / "labels" / split).mkdir(parents=True, exist_ok=True)

    saved = 0
    stamp = int(time.time())
    for index, upload in enumerate(files or [], start=1):
        if not (upload.content_type or "").startswith("image/"):
            continue
        contents = await upload.read()
        if not contents or len(contents) > 12 * 1024 * 1024:
            continue
        frame = cv2.imdecode(np.frombuffer(contents, dtype=np.uint8), cv2.IMREAD_COLOR)
        if frame is None:
            continue
        suffix = Path(upload.filename or "").suffix.lower()
        if suffix not in TRAINING_IMAGE_EXTS:
            suffix = ".jpg"
        (img_dir / f"inject_{stamp}_{index:03d}{suffix}").write_bytes(contents)
        saved += 1

    store = _read_json(TRAINING_PROMPTS_PATH) or {}
    existing = store.get("prompts") if isinstance(store, dict) else []
    if not isinstance(existing, list):
        existing = []
    added: list[str] = []
    seen = {str(value).lower() for value in existing}
    for part in re.split(r"[,\n]+", prompts or ""):
        cleaned = " ".join(part.split()).strip()
        if cleaned and cleaned.lower() not in seen:
            existing.append(cleaned)
            added.append(cleaned)
            seen.add(cleaned.lower())
    TRAINING_PROMPTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    TRAINING_PROMPTS_PATH.write_text(
        json.dumps({"prompts": existing}, indent=2), encoding="utf-8"
    )

    _audit("training_inject", {"split": split, "images_saved": saved, "prompts_added": added})
    _training_backup_dataset("inject")
    return {
        "images_saved": saved,
        "prompts_added": added,
        "dataset": _training_dataset_stats(),
    }


def _training_resolve_class(name: str) -> int:
    """Return the class id for a name, adding it to data.yaml if it is new."""
    import yaml as _yaml

    data = {}
    if TRAINING_DATASET_YAML.exists():
        data = _yaml.safe_load(TRAINING_DATASET_YAML.read_text(encoding="utf-8")) or {}
    names = data.get("names") or {}
    normalized = _catalog_normalize_name(name)
    for class_id, class_name in names.items():
        if _catalog_normalize_name(str(class_name)) == normalized:
            return int(class_id)
    new_id = (max((int(k) for k in names.keys()), default=-1) + 1) if names else 0
    names[new_id] = name
    data["names"] = names
    data.setdefault("path", "datasets/baget_box")
    data.setdefault("train", "images/train")
    data.setdefault("val", "images/val")
    TRAINING_DATASET_YAML.parent.mkdir(parents=True, exist_ok=True)
    TRAINING_DATASET_YAML.write_text(_yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return int(new_id)


@app.post("/api/training/detect")
async def training_detect(files: list[UploadFile] = File(default=[])) -> dict[str, Any]:
    """Run the detector over each image so the operator can name/keep detections."""
    payloads: list[tuple[str, bytes]] = []
    for upload in files or []:
        if not (upload.content_type or "").startswith("image/"):
            continue
        payloads.append((upload.filename or "image", await upload.read()))
    # Detection is heavy + synchronous; run it off the event loop so the
    # dashboard stays responsive.
    return await asyncio.to_thread(_training_detect_sync, payloads)


def _training_detect_sync(payloads: list[tuple[str, bytes]]) -> dict[str, Any]:
    import base64
    import cv2
    import numpy as np

    detector = _training_detector()
    images: list[dict[str, Any]] = []
    for filename, contents in payloads:
        frame = cv2.imdecode(np.frombuffer(contents, dtype=np.uint8), cv2.IMREAD_COLOR)
        if frame is None:
            continue
        height, width = frame.shape[:2]
        boxes: list[tuple[int, int, int, int, str, float]] = []
        if detector is not None:
            try:
                for detection in detector.detect(frame):
                    x1, y1, x2, y2 = detection.box
                    boxes.append((int(x1), int(y1), int(x2), int(y2), detection.class_name, float(detection.confidence)))
            except Exception:  # noqa: BLE001 - fall back to proposals below
                boxes = []
        if not boxes:
            try:
                for (x1, y1, x2, y2) in _catalog_class_agnostic_boxes(frame):
                    boxes.append((int(x1), int(y1), int(x2), int(y2), "object", 0.0))
            except Exception:  # noqa: BLE001 - a bad frame must not fail the request
                boxes = []

        detections: list[dict[str, Any]] = []
        for (x1, y1, x2, y2, name, confidence) in boxes[:30]:
            crop = frame[max(0, y1):max(0, y2), max(0, x1):max(0, x2)]
            if crop.size == 0:
                continue
            ok, buffer = cv2.imencode(".jpg", crop)
            crop_data = (
                f"data:image/jpeg;base64,{base64.b64encode(buffer.tobytes()).decode()}"
                if ok
                else ""
            )
            detections.append(
                {
                    "bbox": {"x1": x1, "y1": y1, "x2": x2, "y2": y2},
                    "name": name,
                    "confidence": round(confidence, 3),
                    "crop": crop_data,
                }
            )
        images.append(
            {"filename": filename, "width": width, "height": height, "detections": detections}
        )
    return {"images": images}


@app.post("/api/training/annotate")
async def training_annotate(
    split: str = Form("train"),
    boxes: str = Form("[]"),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """Save one labelled image: necessary boxes become YOLO labels, others are
    left unlabelled (background). No necessary boxes = a negative image."""
    import cv2
    import numpy as np

    split = "val" if str(split).strip().lower().startswith("v") else "train"
    contents = await file.read()
    frame = cv2.imdecode(np.frombuffer(contents, dtype=np.uint8), cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(status_code=400, detail="Image could not be decoded.")
    height, width = frame.shape[:2]
    try:
        items = json.loads(boxes) or []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid boxes payload.")

    lines: list[str] = []
    for item in items:
        if not item.get("necessary"):
            continue
        name = " ".join(str(item.get("name", "")).split()).strip()
        bbox = item.get("bbox") or {}
        try:
            x1 = float(bbox["x1"]); y1 = float(bbox["y1"]); x2 = float(bbox["x2"]); y2 = float(bbox["y2"])
        except (KeyError, TypeError, ValueError):
            continue
        if not name:
            continue
        x1, x2 = sorted((max(0.0, min(width, x1)), max(0.0, min(width, x2))))
        y1, y2 = sorted((max(0.0, min(height, y1)), max(0.0, min(height, y2))))
        if x2 <= x1 or y2 <= y1:
            continue
        class_id = _training_resolve_class(name)
        cx = ((x1 + x2) / 2.0) / width
        cy = ((y1 + y2) / 2.0) / height
        bw = (x2 - x1) / width
        bh = (y2 - y1) / height
        lines.append(f"{class_id} {cx:.6f} {cy:.6f} {bw:.6f} {bh:.6f}")

    img_dir = TRAINING_DATASET_ROOT / "images" / split
    lbl_dir = TRAINING_DATASET_ROOT / "labels" / split
    img_dir.mkdir(parents=True, exist_ok=True)
    lbl_dir.mkdir(parents=True, exist_ok=True)
    stem = f"label_{int(time.time() * 1000)}"
    ok, buffer = cv2.imencode(".jpg", frame)
    (img_dir / f"{stem}.jpg").write_bytes(buffer.tobytes() if ok else contents)
    (lbl_dir / f"{stem}.txt").write_text(("\n".join(lines) + "\n") if lines else "", encoding="utf-8")

    _audit("training_annotate", {"split": split, "boxes": len(lines), "negative": not lines})
    return {"labeled": len(lines), "negative": not lines, "dataset": _training_dataset_stats()}


TRAINING_STAGING_DIR = SNAPSHOT_DIR / "training-staging"
TRAINING_APPLIED_PATH = TRAINING_DATASET_ROOT / "applied.json"


class TrainingApply(BaseModel):
    group_id: str = Field(min_length=1, max_length=120)
    name: str = Field(default="", max_length=80)
    keep: bool = True
    split: str = "train"
    count: int | None = Field(default=None, ge=0, le=100000)


class TrainingSearch(BaseModel):
    query: str = Field(default="", max_length=120)


class BlockScanStart(BaseModel):
    block_id: int = Field(ge=1)
    camera_ids: list[int] = Field(min_length=1)
    product_id: int = Field(ge=1)


class BenchmarkSubmission(BaseModel):
    inventory_result_id: int = Field(ge=1)
    ground_truth_count: int = Field(ge=0)
    notes: str | None = Field(default=None, max_length=1000)


class DatasetCreate(BaseModel):
    product_id: str = Field(min_length=1, max_length=120)
    product_name: str = Field(min_length=1, max_length=160)


class DatasetCapture(BaseModel):
    camera_id: str = Field(min_length=1, max_length=120)
    block_id: str | None = Field(default=None, max_length=120)


class DatasetAutoCapture(DatasetCapture):
    interval_seconds: float = Field(default=3, ge=0.5, le=3600)
    frames: int = Field(default=30, ge=1, le=1000)


class DatasetAnnotationSave(BaseModel):
    annotations: list[dict[str, Any]] = Field(default_factory=list, max_length=2000)


class DatasetTrainingStart(BaseModel):
    model_name: str = Field(min_length=1, max_length=100)
    base_model: str = Field(default="yolov8s.pt", min_length=1, max_length=240)
    epochs: int = Field(default=100, ge=1, le=1000)
    image_size: int = Field(default=960, ge=320, le=2048)


class DatasetModelBenchmark(BaseModel):
    benchmark_ids: list[int] = Field(min_length=1, max_length=10000)


class ReviewQueueCreate(BaseModel):
    product_id: str = Field(min_length=1, max_length=120)
    dataset_id: str | None = Field(default=None, max_length=160)
    camera_id: str | None = Field(default=None, max_length=120)
    image_path: str = Field(min_length=1, max_length=1000)
    reason: str = Field(min_length=1, max_length=240)
    confidence: float | None = Field(default=None, ge=0, le=1)
    payload: dict[str, Any] = Field(default_factory=dict)


def _training_gemini_suggestion(crop) -> tuple[str | None, float]:
    """Return a best-effort naming-service label and its real confidence."""
    try:
        from recognition.gemini_client import GeminiClient
    except Exception:  # noqa: BLE001
        return None, 0.0
    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    recognition = config.get("recognition", {}) or {}
    try:
        scan_timeout = max(
            1,
            int(os.getenv("AI_VISION_SCAN_NAMING_TIMEOUT_SECONDS", "5")),
        )
        result = GeminiClient(
            model=recognition.get("model", "gemini-3.1-flash-lite"),
            timeout=min(int(recognition.get("timeout", 30)), scan_timeout),
            retries=0,
        ).recognize(crop)
        name = str(getattr(result, "name", "") or "").strip()
        if not name or name.lower() == "unknown product":
            return None, 0.0
        return name, max(0.0, min(1.0, float(getattr(result, "confidence", 0.0) or 0.0)))
    except Exception:  # noqa: BLE001 - naming fallback must not abort a scan
        return None, 0.0


def _training_camera_map(health: dict[str, Any]) -> dict[int, str]:
    """Active Stream Manager cameras, keyed by slot.

    Snapshot files are deliberately not discovery inputs: they may outlive a
    stream session and previously inflated a seven-camera deployment to 26.
    `_catalog_live_frame_image` may still use a snapshot as a frame fallback,
    but only after this function has established that the camera is active.
    """
    cameras: dict[int, str] = {}
    try:
        streams = _get_stream_manager().status().get("streams", [])
    except Exception:  # noqa: BLE001 - a stopped manager means no active cameras
        streams = []
    for camera in streams:
        status = str(camera.get("status") or "").strip().lower()
        slot = camera.get("slot_number")
        if slot is None:
            continue
        if status in {"offline", "stopped"}:
            continue
        cameras[int(slot)] = str(camera.get("name") or f"slot-{slot}")
    return cameras


# --- Dataset-centered recognition -------------------------------------------
# The Recognition button compares live crops against the trained dataset (not
# Gemini): per-class mean embeddings built from the dataset's own labeled crops.
_training_refs_obj: dict[str, list[float]] | None = None
_training_refs_key: tuple[int, float] | None = None


def _training_dataset_signature() -> tuple[int, float]:
    count = 0
    latest = 0.0
    for split in ("train", "val"):
        img_dir = TRAINING_DATASET_ROOT / "images" / split
        if not img_dir.exists():
            continue
        for path in img_dir.glob("*.*"):
            if path.suffix.lower() in TRAINING_IMAGE_EXTS:
                count += 1
                latest = max(latest, path.stat().st_mtime)
    return count, latest


def _training_dataset_reference_embeddings() -> dict[str, list[float]]:
    """Mean color/shape embedding per dataset class, cached until the dataset
    changes. Empty labels (negatives) are skipped."""
    global _training_refs_obj, _training_refs_key
    try:
        import cv2
        import numpy as np

        from recognition.embedding import image_embedding
    except Exception:  # noqa: BLE001
        return {}

    signature = _training_dataset_signature()
    if _training_refs_obj is not None and _training_refs_key == signature:
        return _training_refs_obj

    names: dict[Any, Any] = {}
    if TRAINING_DATASET_YAML.exists():
        import yaml as _yaml

        data = _yaml.safe_load(TRAINING_DATASET_YAML.read_text(encoding="utf-8")) or {}
        names = data.get("names") or {}

    accum: dict[str, list[list[float]]] = {}
    for split in ("train", "val"):
        img_dir = TRAINING_DATASET_ROOT / "images" / split
        lbl_dir = TRAINING_DATASET_ROOT / "labels" / split
        if not img_dir.exists():
            continue
        for img_path in img_dir.glob("*.*"):
            if img_path.suffix.lower() not in TRAINING_IMAGE_EXTS:
                continue
            label_path = lbl_dir / (img_path.stem + ".txt")
            lines = (
                [ln for ln in label_path.read_text(encoding="utf-8").splitlines() if ln.strip()]
                if label_path.exists()
                else []
            )
            if not lines:
                continue  # negative example, no class to learn
            image = cv2.imread(str(img_path))
            if image is None:
                continue
            h, w = image.shape[:2]
            for line in lines[:8]:
                parts = line.split()
                if len(parts) < 5:
                    continue
                try:
                    cid = int(float(parts[0]))
                    cx, cy, bw, bh = (float(parts[1]), float(parts[2]), float(parts[3]), float(parts[4]))
                except ValueError:
                    continue
                name = str(names.get(cid, names.get(str(cid), cid)))
                if bw >= 0.99 and bh >= 0.99:
                    crop = image
                else:
                    x1 = max(0, int((cx - bw / 2) * w))
                    y1 = max(0, int((cy - bh / 2) * h))
                    x2 = min(w, int((cx + bw / 2) * w))
                    y2 = min(h, int((cy + bh / 2) * h))
                    crop = image[y1:y2, x1:x2] if x2 > x1 and y2 > y1 else image
                try:
                    accum.setdefault(name, []).append(image_embedding(crop))
                except Exception:  # noqa: BLE001
                    continue

    refs: dict[str, list[float]] = {}
    for name, vectors in accum.items():
        arr = np.array(vectors, dtype="float32")
        mean = arr.mean(axis=0)
        norm = float(np.linalg.norm(mean))
        refs[name] = [float(v) for v in (mean / norm if norm > 0 else mean)]

    _training_refs_obj = refs
    _training_refs_key = signature
    return refs


def _training_match_dataset(crop, refs: dict[str, list[float]]) -> tuple[str | None, float]:
    """Best-matching dataset class name and cosine similarity for a crop."""
    if not refs or crop is None or getattr(crop, "size", 0) == 0:
        return None, 0.0
    try:
        import numpy as np

        from recognition.embedding import image_embedding

        vector = np.array(image_embedding(crop), dtype="float32")
    except Exception:  # noqa: BLE001
        return None, 0.0
    best_name: str | None = None
    best_score = -1.0
    for name, ref in refs.items():
        ref_arr = np.array(ref, dtype="float32")
        if ref_arr.shape != vector.shape:
            continue
        score = float(np.dot(vector, ref_arr))
        if score > best_score:
            best_score = score
            best_name = name
    return best_name, max(0.0, best_score)


# --- Per-location clustering + box counting ---------------------------------
def _training_box_iou_gap(a, b) -> float:
    """0 = overlapping/touching, grows with the normalized gap between boxes."""
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    dx = max(0, max(ax1, bx1) - min(ax2, bx2))
    dy = max(0, max(ay1, by1) - min(ay2, by2))
    scale = max(1.0, (ax2 - ax1 + bx2 - bx1 + ay2 - ay1 + by2 - by1) / 4.0)
    return (dx + dy) / scale


def _training_cluster_boxes(boxes: list[tuple[int, int, int, int]], gap: float = 0.6) -> list[list[int]]:
    """Union-find grouping of boxes that touch or sit close together (one stack
    at one place). Returns lists of member indices."""
    n = len(boxes)
    parent = list(range(n))

    def find(i: int) -> int:
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    for i in range(n):
        for j in range(i + 1, n):
            if _training_box_iou_gap(boxes[i], boxes[j]) <= gap:
                parent[find(i)] = find(j)
    clusters: dict[int, list[int]] = {}
    for i in range(n):
        clusters.setdefault(find(i), []).append(i)
    return list(clusters.values())


def _training_tiled_count(crop, detector, base_count: int = 0) -> int:
    """Stronger recount for a dense stack: slice the crop into overlapping tiles,
    detect in each at higher resolution, merge with NMS. Returns the larger of
    the tiled count and the original so a recount never lowers a good count."""
    try:
        import numpy as np
    except Exception:  # noqa: BLE001
        return base_count
    if detector is None or crop is None or getattr(crop, "size", 0) == 0:
        return base_count
    h, w = crop.shape[:2]
    tiles_x = 2 if w >= 260 else 1
    tiles_y = 2 if h >= 260 else 1
    overlap = 0.2
    merged: list[tuple[float, float, float, float]] = []
    step_x = w / tiles_x
    step_y = h / tiles_y
    for ty in range(tiles_y):
        for tx in range(tiles_x):
            x0 = int(max(0, step_x * tx - step_x * overlap))
            y0 = int(max(0, step_y * ty - step_y * overlap))
            x1 = int(min(w, step_x * (tx + 1) + step_x * overlap))
            y1 = int(min(h, step_y * (ty + 1) + step_y * overlap))
            tile = crop[y0:y1, x0:x1]
            if tile.size == 0:
                continue
            try:
                dets = detector.detect(tile)
            except Exception:  # noqa: BLE001
                continue
            for det in dets or []:
                box = getattr(det, "box", None)
                if not box or len(box) < 4:
                    continue
                merged.append((box[0] + x0, box[1] + y0, box[2] + x0, box[3] + y0))
    # Greedy NMS to drop tile-overlap duplicates.
    merged.sort(key=lambda b: (b[2] - b[0]) * (b[3] - b[1]), reverse=True)
    kept: list[tuple[float, float, float, float]] = []
    for box in merged:
        keep = True
        for k in kept:
            ix1, iy1 = max(box[0], k[0]), max(box[1], k[1])
            ix2, iy2 = min(box[2], k[2]), min(box[3], k[3])
            inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
            area_b = max(1, (box[2] - box[0]) * (box[3] - box[1]))
            if inter / area_b > 0.45:
                keep = False
                break
        if keep:
            kept.append(box)
    return max(base_count, len(kept))


@app.post("/api/training/analytics/apply")
async def training_analytics_apply(payload: TrainingApply) -> dict[str, Any]:
    """Directly edit the dataset from an analytics row: keep = save the item's
    detected crops as positive examples under the name; ignore = save them as
    negatives. Re-applying the same item replaces its previous entries."""
    group_id = _catalog_visual_slug(payload.group_id)
    name = " ".join(payload.name.split()).strip()
    keep = bool(payload.keep)
    split = "val" if str(payload.split).strip().lower().startswith("v") else "train"
    if not group_id:
        raise HTTPException(status_code=400, detail="group_id is required.")

    stage = TRAINING_STAGING_DIR / group_id
    crops = sorted(stage.glob("*.jpg")) if stage.exists() else []
    if not crops:
        raise HTTPException(status_code=404, detail="No staged crops; run the test again.")
    if keep and not name:
        raise HTTPException(status_code=400, detail="A name is required to keep an item.")

    applied = _read_json(TRAINING_APPLIED_PATH) or {}
    if not isinstance(applied, dict):
        applied = {}
    for old_rel in applied.get(group_id, []):
        image_path = TRAINING_DATASET_ROOT / old_rel
        if image_path.exists():
            image_path.unlink()
        # old_rel is images/<split>/<stem>.jpg -> labels/<split>/<stem>.txt
        label_path = TRAINING_DATASET_ROOT / "labels" / Path(old_rel).parent.name / (Path(old_rel).stem + ".txt")
        if label_path.exists():
            label_path.unlink()

    class_id = _training_resolve_class(name) if keep else None
    img_dir = TRAINING_DATASET_ROOT / "images" / split
    lbl_dir = TRAINING_DATASET_ROOT / "labels" / split
    img_dir.mkdir(parents=True, exist_ok=True)
    lbl_dir.mkdir(parents=True, exist_ok=True)
    written: list[str] = []
    for index, crop_path in enumerate(crops):
        stem = f"analytics_{group_id}_{index:02d}"
        (img_dir / f"{stem}.jpg").write_bytes(crop_path.read_bytes())
        if keep and class_id is not None:
            (lbl_dir / f"{stem}.txt").write_text(f"{class_id} 0.5 0.5 1.0 1.0\n", encoding="utf-8")
        else:
            (lbl_dir / f"{stem}.txt").write_text("", encoding="utf-8")  # negative
        written.append(f"images/{split}/{stem}.jpg")

    applied[group_id] = written
    TRAINING_APPLIED_PATH.parent.mkdir(parents=True, exist_ok=True)
    TRAINING_APPLIED_PATH.write_text(json.dumps(applied, indent=2), encoding="utf-8")

    # Keep the persisted scan result aligned with the operator's final decision
    # so reopening Analytics and exporting Excel show the corrected values.
    with _training_search_lock:
        for row in _training_search_state.get("rows", []):
            if row.get("group_id") == group_id:
                row["name"] = name
                row["keep"] = keep
                break
        current_search = dict(_training_search_state)
    if current_search:
        try:
            TRAINING_SEARCH_STATE_PATH.write_text(json.dumps(current_search), encoding="utf-8")
        except Exception:  # noqa: BLE001 - dataset save already succeeded
            pass

    # Persist the human-confirmed box count for this item as inventory metadata,
    # so the exact number survives even though YOLO labels can't carry a count.
    if payload.count is not None:
        counts_path = TRAINING_DATASET_ROOT / "counts.json"
        counts = _read_json(counts_path) or {}
        if not isinstance(counts, dict):
            counts = {}
        if keep:
            counts[group_id] = {"name": name, "count": int(payload.count)}
        else:
            counts.pop(group_id, None)
        counts_path.write_text(json.dumps(counts, indent=2), encoding="utf-8")

    _audit(
        "training_analytics_apply",
        {"group_id": group_id, "name": name, "keep": keep, "count": len(written), "boxes": payload.count},
    )
    if written:
        _training_backup_dataset("apply")
    return {"applied": len(written), "keep": keep, "dataset": _training_dataset_stats()}


# --- Recognition search as a persistent background job -----------------------
# The sweep runs in a daemon thread and writes its progress + results to disk
# after every camera, so it keeps running even when the browser tab is closed
# and the last results are restored the next time the page is opened.
TRAINING_SEARCH_STATE_PATH = TRAINING_STAGING_DIR / "search_job.json"
_training_search_lock = threading.Lock()
_training_search_state: dict[str, Any] = {}
_training_search_generation = 0
_training_scan_trackers: dict[str, Any] = {}
_training_scan_sequences: dict[str, int] = {}
# True only while a worker thread is actually alive. A "running" state with no
# live worker (e.g. after a container restart mid-sweep, whose state was
# persisted to disk) is stale and must not make the page poll forever.
_training_search_active = False


def _training_search_idle_state() -> dict[str, Any]:
    return {
        "status": "idle",
        "query": "",
        "rows": [],
        "diagnostics": {
            "total_active_cameras": 0,
            "attempted_cameras": 0,
            "frames_read": 0,
            "cameras_failed": 0,
            "cameras_completed": 0,
            "detections": 0,
            "raw_detection_count": 0,
            "accepted_detection_count": 0,
            "rejected_detection_count": 0,
            "final_inventory_count": 0,
            "model": True,
        },
        "progress": {"done": 0, "total": 0},
        "stage": "idle",
        "message": "Ready to scan live cameras.",
        "current_camera": None,
        "started_at": None,
        "finished_at": None,
    }


def _training_search_status() -> dict[str, Any]:
    """Current job state: the in-memory copy if present, else the last one
    persisted to disk (survives a backend restart), else idle. A "running"
    state with no live worker is corrected to "done" so the page never polls
    forever after a restart killed the worker."""
    with _training_search_lock:
        state = dict(_training_search_state) if _training_search_state else None
    if state is None:
        try:
            state = json.loads(TRAINING_SEARCH_STATE_PATH.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return _training_search_idle_state()
    if state.get("status") == "running" and not _training_search_active:
        state["status"] = "done"
        state["stage"] = "interrupted"
        state["message"] = "The previous scan was interrupted before completion. Start a new scan."
        state["current_camera"] = None
        state["finished_at"] = state.get("finished_at") or datetime.now(timezone.utc).isoformat()
        diag = dict(state.get("diagnostics") or {})
        diag["stopped_early"] = True
        diag["interrupted"] = True
        state["diagnostics"] = diag
    return state


def _training_search_write_state(state: dict[str, Any], generation: int) -> bool:
    """Persist state only while this worker is still the current generation, so
    a restarted search (new query) cannot be clobbered by a superseded one."""
    with _training_search_lock:
        if generation != _training_search_generation:
            return False
        _training_search_state.clear()
        _training_search_state.update(state)
    try:
        TRAINING_STAGING_DIR.mkdir(parents=True, exist_ok=True)
        TRAINING_SEARCH_STATE_PATH.write_text(json.dumps(state), encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass
    return True


def _training_scan_tracker(camera_name: str):
    tracker = _training_scan_trackers.get(camera_name)
    if tracker is None:
        from tracking.bytetrack_adapter import ByteTrackAdapter

        tracker = ByteTrackAdapter(camera_name, str(ROOT / "config" / "warehouse_bytetrack.yaml"))
        _training_scan_trackers[camera_name] = tracker
    return tracker


def _training_detection_context(query: str) -> tuple[Any, dict[str, Any]]:
    """Build universal detection; the query is used only after recognition."""
    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    detection = config.get("detection", {}) or {}
    configured = [str(value).strip() for value in detection.get("class_prompts") or [] if str(value).strip()]
    broad = [
        str(value).strip()
        for value in detection.get("broad_discovery_prompts")
        or [
            "person", "box", "package", "bag", "sack", "pallet", "cart", "forklift",
            "bottle", "container", "machine", "vehicle", "tool", "furniture", "animal",
        ]
        if str(value).strip()
    ]
    query_prompt = " ".join(query.split()).strip()
    deployed = _deployed_product_detector(query_prompt)
    if deployed is not None:
        health = deployed.health()
        class_name = re.sub(r"[^a-z0-9]+", "_", query_prompt.lower()).strip("_")
        return deployed, {
            "query_prompt": query_prompt, "configured_prompts": configured,
            "broad_discovery": False, "stock_closed_class": True, "active_prompts": [],
            "requested_model": health["requested_model"], "loaded_model": health["active_model"],
            "detector_mode": f"{class_name}_custom", "detection_mode": f"{class_name}_custom",
            "fallback_used": False,
        }
    if _catalog_normalize_name(query_prompt) == "baget box":
        detector = _baget_inventory_detector()
        health = detector.health()
        return detector, {
            "query_prompt": query_prompt,
            "configured_prompts": configured,
            "broad_discovery": False,
            "stock_closed_class": True,
            "active_prompts": [],
            "requested_model": health["requested_model"],
            "loaded_model": health["active_model"],
            "detector_mode": "baget_box_custom",
            "detection_mode": "baget_box_custom",
            "fallback_used": False,
        }
    requested_prompts = broad
    try:
        detector = _training_detector(requested_prompts)
    except TypeError:
        # Compatibility for tests and deployments that temporarily replace the
        # detector factory with the historical no-argument callable.
        detector = _training_detector()
    model = getattr(detector, "model", None) if detector is not None else None
    open_vocabulary = callable(getattr(model, "set_classes", None))
    health = detector.health() if detector is not None and hasattr(detector, "health") else {}
    return detector, {
        "query_prompt": query_prompt or None,
        "configured_prompts": configured,
        "broad_discovery": bool(open_vocabulary),
        "stock_closed_class": bool(detector is not None and not open_vocabulary),
        "active_prompts": requested_prompts if open_vocabulary else [],
        "requested_model": health.get("requested_model"),
        "loaded_model": health.get("active_model"),
        "detector_mode": health.get("detector_mode", "universal"),
        "fallback_used": bool(health.get("fallback_used")),
        "detection_mode": (
            "universal_detection"
            if open_vocabulary
            else "stock_closed_class"
        ),
    }


def _scan_product_database():
    from knowledge.product_database import ProductDatabase

    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    recognition = config.get("recognition", {}) or {}
    path = os.getenv("PRODUCT_DB_PATH", recognition.get("db_path", "database/products.db"))
    return ProductDatabase(path)


def _scan_products() -> list[dict[str, Any]]:
    """Load scan products from the recognition store or legacy warehouse store."""
    try:
        products = _scan_product_database().list_products()
        if products:
            return products
    except Exception:
        pass
    warehouse = WarehouseDB(
        os.getenv("WAREHOUSE_DB_PATH", str(ROOT / "database" / "warehouse.db"))
    )
    with warehouse.db.connect() as connection:
        rows = connection.execute(
            "SELECT id, name, category FROM products ORDER BY name"
        ).fetchall()
    return [dict(row) for row in rows]


@app.get("/api/v1/products")
def list_scan_products() -> dict[str, Any]:
    products = _scan_products()
    return {"data": products, "meta": {"count": len(products)}}


def _dataset_error(exc: DatasetError, status: int = 400) -> HTTPException:
    return HTTPException(status_code=status, detail=str(exc))


@app.get("/api/v1/datasets/models")
def list_dataset_models() -> dict[str, Any]:
    rows = _dataset_training().models()
    return {"data": rows, "meta": {"count": len(rows)}}


@app.get("/api/v1/datasets/models/{model_id}")
def get_dataset_model(model_id: str) -> dict[str, Any]:
    try:
        return {"data": _dataset_training().model(model_id)}
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc


@app.post("/api/v1/datasets/models/{model_id}/benchmark")
def benchmark_dataset_model(model_id: str, body: DatasetModelBenchmark) -> dict[str, Any]:
    manager = _dataset_training()
    try:
        model = manager.model(model_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    if model["status"] != "COMPLETED":
        raise HTTPException(status_code=409, detail="Only completed models can be benchmarked.")
    dataset = manager.builder.get_dataset(model["dataset_id"])
    placeholders = ",".join("?" for _ in body.benchmark_ids)
    rows = manager.builder.database.dataset_fetchall(
        f"SELECT * FROM vision_benchmarks WHERE id IN ({placeholders})", tuple(body.benchmark_ids)
    )
    if len(rows) != len(set(body.benchmark_ids)):
        raise HTTPException(status_code=404, detail="One or more real-camera benchmark records were not found.")
    if any(_catalog_normalize_name(row.get("target_product")) != _catalog_normalize_name(dataset["product_name"]) for row in rows):
        raise HTTPException(status_code=400, detail="Every benchmark record must belong to the model Product.")
    truth = sum(int(row["ground_truth_count"]) for row in rows)
    predicted = sum(int(row["predicted_count"]) for row in rows)
    accuracy = 1.0 if truth == predicted == 0 else (max(0.0, 1.0 - abs(predicted - truth) / truth) if truth else 0.0)
    metrics = dict(model.get("metrics") or {})
    metrics["real_camera_benchmark"] = {
        "benchmark_ids": body.benchmark_ids, "ground_truth": truth,
        "predicted": predicted, "accuracy": accuracy,
    }
    manager.builder.database.dataset_execute(
        "UPDATE vision_models SET benchmark_accuracy=?,metrics=? WHERE id=?",
        (accuracy, json.dumps(metrics), model_id),
    )
    return {"data": manager.model(model_id)}


def _activate_dataset_model(model: dict) -> None:
    global _baget_detector_obj, _baget_detector_key
    weights = Path(str(model.get("weights_path") or ""))
    if model["status"] != "COMPLETED" or not weights.exists():
        raise DatasetError("Completed model weights are unavailable.")
    if float(model.get("benchmark_accuracy") or 0) < 0.90:
        raise DatasetError("Real-camera benchmark accuracy must be at least 90% before deployment.")
    dataset = _dataset_builder().get_dataset(model["dataset_id"])
    active = _read_json(ACTIVE_MODELS_PATH) or {}
    product_key = dataset["class_name"]
    previous = (active.get(product_key) or {}).get("model_id")
    active[product_key] = {
        "model_id": model["id"], "model_name": model["name"],
        "weights_path": str(weights), "detector_mode": f"{product_key}_custom",
        "benchmark_accuracy": model["benchmark_accuracy"], "deployed_at": datetime.now(timezone.utc).isoformat(),
    }
    ACTIVE_MODELS_PATH.parent.mkdir(parents=True, exist_ok=True)
    temporary = ACTIVE_MODELS_PATH.with_suffix(".tmp")
    temporary.write_text(json.dumps(active, indent=2), encoding="utf-8")
    temporary.replace(ACTIVE_MODELS_PATH)
    database = _dataset_builder().database
    database.dataset_execute(
        "UPDATE vision_models SET deployment_status='INACTIVE' WHERE dataset_id IN (SELECT id FROM vision_datasets WHERE product_id=?)",
        (dataset["product_id"],),
    )
    deployed_at = datetime.now(timezone.utc).isoformat()
    database.dataset_execute(
        "UPDATE vision_models SET deployment_status='ACTIVE',deployed_at=? WHERE id=?", (deployed_at, model["id"])
    )
    database.dataset_execute(
        "INSERT INTO vision_model_deployments(product_id,model_id,previous_model_id,deployed_at) VALUES(?,?,?,?)",
        (dataset["product_id"], model["id"], previous, deployed_at),
    )
    _active_product_detectors.clear()
    _baget_detector_obj = _baget_detector_key = None


@app.post("/api/v1/datasets/models/{model_id}/deploy")
def deploy_dataset_model(model_id: str) -> dict[str, Any]:
    try:
        model = _dataset_training().model(model_id)
        _activate_dataset_model(model)
        return {"data": _dataset_training().model(model_id)}
    except DatasetError as exc:
        raise _dataset_error(exc, 409) from exc


@app.post("/api/v1/datasets/models/{model_id}/rollback")
def rollback_dataset_model(model_id: str) -> dict[str, Any]:
    manager = _dataset_training()
    try:
        current = manager.model(model_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    dataset = manager.builder.get_dataset(current["dataset_id"])
    row = manager.builder.database.dataset_fetchone(
        "SELECT previous_model_id FROM vision_model_deployments WHERE product_id=? AND model_id=? ORDER BY id DESC LIMIT 1",
        (dataset["product_id"], model_id),
    )
    if not row or not row.get("previous_model_id"):
        raise HTTPException(status_code=409, detail="No previous deployed model is available.")
    try:
        previous = manager.model(row["previous_model_id"])
        _activate_dataset_model(previous)
        return {"data": previous}
    except DatasetError as exc:
        raise _dataset_error(exc, 409) from exc


@app.get("/api/v1/datasets/review-queue")
def list_dataset_review_queue(status: str = "PENDING") -> dict[str, Any]:
    rows = _dataset_builder().database.dataset_fetchall(
        "SELECT * FROM vision_review_queue WHERE status=? ORDER BY created_at DESC", (status.upper(),)
    )
    for row in rows:
        row["payload"] = json.loads(row.get("payload") or "{}")
    return {"data": rows, "meta": {"count": len(rows)}}


@app.post("/api/v1/datasets/review-queue")
def create_dataset_review_case(body: ReviewQueueCreate) -> dict[str, Any]:
    case_id, now = secrets.token_hex(16), datetime.now(timezone.utc).isoformat()
    _dataset_builder().database.dataset_execute(
        "INSERT INTO vision_review_queue(id,product_id,dataset_id,camera_id,image_path,reason,confidence,status,payload,created_at,reviewed_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        (case_id, body.product_id, body.dataset_id, body.camera_id, body.image_path,
         body.reason, body.confidence, "PENDING", json.dumps(body.payload), now, None),
    )
    return {"data": {"id": case_id, "status": "PENDING"}}


@app.post("/api/v1/datasets/review-queue/{case_id}/approve")
def approve_dataset_review_case(case_id: str) -> dict[str, Any]:
    database = _dataset_builder().database
    row = database.dataset_fetchone("SELECT * FROM vision_review_queue WHERE id=?", (case_id,))
    if row is None:
        raise HTTPException(status_code=404, detail="Review case not found.")
    database.dataset_execute(
        "UPDATE vision_review_queue SET status='APPROVED',reviewed_at=? WHERE id=?",
        (datetime.now(timezone.utc).isoformat(), case_id),
    )
    return {"data": {"id": case_id, "status": "APPROVED", "dataset_id": row.get("dataset_id")}}


@app.get("/api/v1/datasets/capture-jobs/{job_id}")
def get_dataset_capture_job(job_id: str) -> dict[str, Any]:
    job = _dataset_capture_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Capture job not found.")
    return {"data": job}


@app.get("/api/v1/datasets")
def list_datasets(include_archived: bool = False) -> dict[str, Any]:
    rows = _dataset_builder().list_datasets(include_archived)
    return {"data": rows, "meta": {"count": len(rows)}}


@app.post("/api/v1/datasets")
def create_dataset(body: DatasetCreate) -> dict[str, Any]:
    product = next((row for row in _scan_products() if str(row.get("id")) == body.product_id), None)
    if product is None or _catalog_normalize_name(product.get("name")) != _catalog_normalize_name(body.product_name):
        raise HTTPException(status_code=400, detail="Dataset product must match the existing Product database.")
    try:
        return {"data": _dataset_builder().create_dataset(body.product_id, str(product["name"]))}
    except DatasetError as exc:
        raise _dataset_error(exc) from exc


@app.get("/api/v1/datasets/{dataset_id}")
def get_dataset(dataset_id: str) -> dict[str, Any]:
    try:
        dataset = _dataset_builder().get_dataset(dataset_id)
        dataset["images"] = _dataset_builder().images(dataset_id)
        return {"data": dataset}
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc


@app.get("/api/v1/datasets/{dataset_id}/health")
def get_dataset_health(dataset_id: str) -> dict[str, Any]:
    try:
        return {"data": _dataset_builder().health(dataset_id)}
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc


@app.post("/api/v1/datasets/{dataset_id}/archive")
def archive_dataset(dataset_id: str) -> dict[str, Any]:
    try:
        _dataset_builder().get_dataset(dataset_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    now = datetime.now(timezone.utc).isoformat()
    _dataset_builder().database.dataset_execute(
        "UPDATE vision_datasets SET archived=1,status='ARCHIVED',updated_at=? WHERE id=?", (now, dataset_id)
    )
    return {"data": _dataset_builder().get_dataset(dataset_id)}


@app.post("/api/v1/datasets/{dataset_id}/duplicate")
def duplicate_dataset(dataset_id: str) -> dict[str, Any]:
    builder = _dataset_builder()
    try:
        source = builder.get_dataset(dataset_id)
        target = builder.create_dataset(source["product_id"], source["product_name"])
        for source_image in builder.images(dataset_id):
            imported = builder.ingest(
                target["id"], Path(source_image["original_path"]).read_bytes(),
                suffix=Path(source_image["original_path"]).suffix, source="dataset_version",
                camera_id=source_image.get("camera_id"), block_id=source_image.get("block_id"),
            )
            if imported.get("skipped") or source_image["annotation_status"] != "VERIFIED":
                continue
            boxes = builder.get_image(dataset_id, source_image["id"])["annotations"]
            builder.save_annotations(target["id"], imported["id"], [{
                "x1": box["x1"], "y1": box["y1"], "x2": box["x2"], "y2": box["y2"],
                "provenance": "manual", "confidence": box.get("confidence"),
            } for box in boxes if box["status"] == "VERIFIED"])
            builder.approve(target["id"], imported["id"])
        return {"data": builder.get_dataset(target["id"])}
    except (DatasetError, OSError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/v1/datasets/{dataset_id}/export")
def export_dataset(dataset_id: str) -> StreamingResponse:
    import zipfile
    builder = _dataset_builder()
    try:
        dataset = builder.get_dataset(dataset_id)
        root = builder.materialize(dataset_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in root.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(root))
    output.seek(0)
    return StreamingResponse(output, media_type="application/zip", headers={
        "Content-Disposition": f'attachment; filename="{dataset["class_name"]}_v{dataset["version"]}.zip"'
    })


@app.post("/api/v1/datasets/{dataset_id}/import-legacy")
def import_legacy_dataset_images(dataset_id: str) -> dict[str, Any]:
    """Import legacy images only; historical labels/classes are intentionally ignored."""
    builder = _dataset_builder()
    try:
        dataset = builder.get_dataset(dataset_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    candidates = [ROOT / "datasets" / dataset["class_name"], ROOT / "datasets" / "baget_box"]
    paths = sorted({path.resolve() for root in candidates if root.exists() for path in root.rglob("*") if path.suffix.lower() in builder.allowed_extensions})
    results = []
    for path in paths:
        # Do not re-import Dataset Builder's own materialized version.
        if f"{os.sep}v{dataset['version']}{os.sep}" in str(path):
            continue
        results.append(builder.ingest(dataset_id, path.read_bytes(), suffix=path.suffix, source="legacy_import"))
    return {"data": results, "meta": {"found": len(paths), "accepted": sum(not row.get("skipped") for row in results), "labels_imported": 0}}


def _capture_dataset_once(dataset_id: str, body: DatasetCapture) -> dict[str, Any]:
    camera = next((row for row in _camera_operations_payload() if str(row.get("id")) == body.camera_id), None)
    if camera is None:
        raise DatasetError("Camera not found.")
    if body.block_id is not None and str(camera.get("block_id")) != body.block_id:
        raise DatasetError("Camera does not belong to the selected Block.")
    frame = _catalog_live_frame_image(slot=camera.get("slot_number"), camera=camera.get("name"))
    if frame is None:
        raise DatasetError("No live frame is currently available from this camera.")
    import cv2
    ok, encoded = cv2.imencode(".jpg", frame, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
    if not ok:
        raise DatasetError("Live frame could not be encoded.")
    return _dataset_builder().ingest(
        dataset_id, encoded.tobytes(), suffix=".jpg", source="camera",
        camera_id=body.camera_id, block_id=body.block_id,
        camera_number=camera.get("slot_number") or camera.get("id"),
        block_name=camera.get("block_name"),
    )


@app.post("/api/v1/datasets/{dataset_id}/capture")
def capture_dataset_frame(dataset_id: str, body: DatasetCapture) -> dict[str, Any]:
    try:
        return {"data": _capture_dataset_once(dataset_id, body)}
    except DatasetError as exc:
        raise _dataset_error(exc) from exc


@app.post("/api/v1/datasets/{dataset_id}/auto-capture")
def start_dataset_auto_capture(dataset_id: str, body: DatasetAutoCapture) -> dict[str, Any]:
    try:
        _dataset_builder().get_dataset(dataset_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    job_id = secrets.token_hex(12)
    _dataset_capture_jobs[job_id] = {"id": job_id, "status": "RUNNING", "requested": body.frames, "accepted": 0, "skipped": 0, "errors": []}

    def worker() -> None:
        job = _dataset_capture_jobs[job_id]
        for index in range(body.frames):
            try:
                result = _capture_dataset_once(dataset_id, body)
                job["skipped" if result.get("skipped") else "accepted"] += 1
            except Exception as exc:  # noqa: BLE001
                job["errors"].append(str(exc))
            if index + 1 < body.frames:
                time.sleep(body.interval_seconds)
        job["status"] = "COMPLETED" if not job["errors"] else "COMPLETED_WITH_ERRORS"

    threading.Thread(target=worker, daemon=True).start()
    return {"data": _dataset_capture_jobs[job_id]}


@app.post("/api/v1/datasets/{dataset_id}/upload")
async def upload_dataset_images(dataset_id: str, files: list[UploadFile] = File(...)) -> dict[str, Any]:
    results = []
    for upload in files:
        suffix = Path(upload.filename or "").suffix.lower()
        try:
            results.append(_dataset_builder().ingest(dataset_id, await upload.read(), suffix=suffix, source="upload"))
        except DatasetError as exc:
            results.append({"filename": upload.filename, "error": str(exc)})
    return {"data": results, "meta": {"accepted": sum(not row.get("skipped") and not row.get("error") for row in results)}}


@app.get("/api/v1/datasets/{dataset_id}/images/{image_id}")
def get_dataset_image(dataset_id: str, image_id: str, metadata: bool = False):
    try:
        row = _dataset_builder().get_image(dataset_id, image_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    return {"data": row} if metadata else FileResponse(row["original_path"])


@app.post("/api/v1/datasets/{dataset_id}/images/{image_id}/suggestions")
def suggest_dataset_annotations(dataset_id: str, image_id: str) -> dict[str, Any]:
    builder = _dataset_builder()
    try:
        dataset = builder.get_dataset(dataset_id)
        image_row = builder.get_image(dataset_id, image_id)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc
    if image_row["annotations"]:
        raise HTTPException(status_code=409, detail="Resolve or delete the existing annotation draft before requesting suggestions.")
    import cv2
    frame = cv2.imread(image_row["original_path"])
    if frame is None:
        raise HTTPException(status_code=500, detail="Dataset image cannot be decoded.")
    detector = _training_detector([dataset["product_name"]])
    if detector is None:
        raise HTTPException(status_code=503, detail="AI annotation suggestions are currently unavailable; manual annotation remains available.")
    height, width = frame.shape[:2]
    suggestions = []
    try:
        detections = detector.detect(frame)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"AI annotation suggestions failed: {exc}") from exc
    for detection in detections:
        box = getattr(detection, "box", None)
        if not box or len(box) != 4:
            continue
        x1, y1, x2, y2 = [float(value) for value in box]
        suggestions.append({
            "x1": max(0.0, min(1.0, x1 / width)), "y1": max(0.0, min(1.0, y1 / height)),
            "x2": max(0.0, min(1.0, x2 / width)), "y2": max(0.0, min(1.0, y2 / height)),
            "provenance": "ai_suggested", "confidence": float(getattr(detection, "confidence", 0.0) or 0.0),
        })
    result = builder.save_annotations(dataset_id, image_id, suggestions)
    return {"data": result, "meta": {"status": "UNVERIFIED", "count": len(suggestions)}}


@app.put("/api/v1/datasets/{dataset_id}/images/{image_id}/annotations")
def save_dataset_annotations(dataset_id: str, image_id: str, body: DatasetAnnotationSave) -> dict[str, Any]:
    try:
        return {"data": _dataset_builder().save_annotations(dataset_id, image_id, body.annotations)}
    except DatasetError as exc:
        raise _dataset_error(exc) from exc


@app.post("/api/v1/datasets/{dataset_id}/images/{image_id}/approve")
def approve_dataset_image(dataset_id: str, image_id: str) -> dict[str, Any]:
    try:
        return {"data": _dataset_builder().approve(dataset_id, image_id)}
    except DatasetError as exc:
        raise _dataset_error(exc, 409) from exc


@app.delete("/api/v1/datasets/{dataset_id}/images/{image_id}", status_code=204)
def reject_dataset_image(dataset_id: str, image_id: str) -> Response:
    try:
        _dataset_builder().reject(dataset_id, image_id)
        return Response(status_code=204)
    except DatasetError as exc:
        raise _dataset_error(exc, 404) from exc


@app.post("/api/v1/datasets/{dataset_id}/train")
def train_dataset_model(dataset_id: str, body: DatasetTrainingStart) -> dict[str, Any]:
    try:
        return {"data": _dataset_training().start(dataset_id, body.model_name, body.base_model, body.epochs, body.image_size)}
    except DatasetError as exc:
        raise _dataset_error(exc, 409) from exc


@app.get("/api/v1/benchmarks")
def list_inventory_benchmarks(limit: int = 200) -> dict[str, Any]:
    rows = VisionDB(
        os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))
    ).list_benchmarks(limit)
    return {"data": rows, "meta": {"count": len(rows)}}


@app.post("/api/v1/benchmarks")
def create_inventory_benchmark(body: BenchmarkSubmission) -> dict[str, Any]:
    database = VisionDB(
        os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))
    )
    try:
        benchmark_id = database.record_benchmark(
            body.inventory_result_id, body.ground_truth_count, body.notes
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return {"data": {"id": benchmark_id}}


def _training_track_ids(detections, tracked_objects) -> dict[int, int]:
    """Best-effort raw-detection to track mapping; never filters detections."""
    matches: dict[int, int] = {}
    unused = list(tracked_objects or [])
    for index, detection in enumerate(detections):
        box = getattr(detection, "box", None)
        if not box or not unused:
            continue
        best = max(unused, key=lambda tracked: _catalog_box_iou(tuple(box), tuple(tracked.box)))
        if _catalog_box_iou(tuple(box), tuple(best.box)) > 0:
            matches[index] = int(best.track_id)
            unused.remove(best)
    return matches


def _training_scan_camera(slot, cam_name, term, detector, refs, diag=None, seq_start=0,
                          block_id=None):
    """Capture and create one row per raw detector result; tracking is optional enrichment."""
    import cv2

    local_diag = {
        "frames_read": 0, "detections": 0, "tracked": 0,
        "raw_detection_count": 0, "accepted_detection_count": 0,
        "rejected_detection_count": 0, "final_inventory_count": 0,
        "failure_reason": None,
    }
    rows: list[dict[str, Any]] = []
    seq = seq_start
    frame = _catalog_live_frame_image(slot=slot, camera=cam_name)
    if frame is None:
        local_diag["failure_reason"] = "no_frame"
        if diag is not None:
            diag.update({key: int(diag.get(key, 0)) + value for key, value in local_diag.items() if isinstance(value, int)})
        return rows, seq, local_diag
    local_diag["frames_read"] = 1
    try:
        detections = detector.detect(frame)
    except Exception as exc:  # noqa: BLE001
        _audit("training_search_detect_failed", {"slot": slot, "error": str(exc)})
        local_diag["failure_reason"] = "detector_unavailable"
        return rows, seq, local_diag
    detections = [det for det in (detections or []) if getattr(det, "box", None)]
    local_diag["detections"] = len(detections)
    local_diag["raw_detection_count"] = len(detections)
    if not detections:
        local_diag["failure_reason"] = "no_detections"
        return rows, seq, local_diag
    sequence = _training_scan_sequences.get(cam_name, 0) + 1
    _training_scan_sequences[cam_name] = sequence
    try:
        tracked_objects = _training_scan_tracker(cam_name).update(
            detections,
            frame.shape,
            sequence,
            time.time(),
        )
    except Exception as exc:  # noqa: BLE001
        _audit("training_search_tracking_failed", {"camera": cam_name, "error": str(exc)})
        tracked_objects = []
    local_diag["tracked"] = len(tracked_objects)
    track_ids = _training_track_ids(detections, tracked_objects)
    config = _read_yaml(CONFIG_PATH) if CONFIG_PATH.exists() else {}
    match_threshold = float((config.get("recognition", {}) or {}).get("similarity_threshold", 0.62))
    for detection_index, detection in enumerate(detections):
        box = detection.box
        crop = _catalog_detection_crop(
            frame,
            {"x1": box[0], "y1": box[1], "x2": box[2], "y2": box[3]},
        )
        if crop is None or crop.size == 0:
            continue
        matched, score = _training_match_dataset(crop, refs)
        source = "dataset"
        confidence = score
        suggested_name = matched if (matched and score >= match_threshold) else None
        if not suggested_name:
            detector_label = str(
                getattr(detection, "inventory_name", None)
                or detection.class_name
                or "object"
            ).strip()
            detector_label_key = _catalog_normalize_name(detector_label)
            generic_detector_labels = {
                "", "object", "object proposal", "unknown", "unknown object",
            }
            if detector_label_key not in generic_detector_labels:
                # YOLO already identified this object. Returning that label is
                # immediate; Gemini is reserved for genuinely unknown boxes.
                suggested_name = detector_label
                confidence = float(detection.confidence or 0.0)
                source = "yolo"
            else:
                suggested_name, naming_confidence = _training_gemini_suggestion(crop)
                confidence = naming_confidence
                source = "naming_service" if suggested_name else "yolo"
                if not suggested_name:
                    suggested_name = detector_label or "object"
                    confidence = float(detection.confidence or 0.0)
        if term == "baget box" and _catalog_normalize_name(detection.class_name or "") == "baget box":
            # The factory detector's explicit class is authoritative. Local
            # recognition may enrich it, but cannot rename it away from the
            # selected inventory target.
            suggested_name = "Baget Box"
            confidence = float(detection.confidence or 0.0)
            source = "baget_box_custom"
        seq += 1
        track_id = track_ids.get(detection_index)
        identity = track_id if track_id is not None else f"raw-{detection_index}"
        group_id = _catalog_visual_slug(f"scan-{slot}-{identity}-{sequence}-{seq}")
        stage = TRAINING_STAGING_DIR / group_id
        if stage.exists():
            for old in stage.glob("*.jpg"):
                old.unlink()
        stage.mkdir(parents=True, exist_ok=True)
        ok, buffer = cv2.imencode(".jpg", crop)
        if ok:
            (stage / "crop_00.jpg").write_bytes(buffer.tobytes())
        rows.append(
            {
                "group_id": group_id,
                "suggested_name": suggested_name,
                "name": suggested_name,
                "camera": cam_name,
                "slot": slot,
                "track_id": track_id,
                "confidence": round(confidence, 3),
                "source": source,
                "detector_class": str(detection.class_name or "object"),
                "detector_confidence": round(float(detection.confidence or 0.0), 3),
                "bbox": [int(value) for value in box],
                "crop_url": f"/snapshots/training-staging/{quote(group_id)}/crop_00.jpg" if ok else "",
                "keep": True,
            }
        )
    inventory_config = (config.get("inventory", {}) or {})
    counter = VisibleInventoryCounter(
        target_product=term or (rows[0]["suggested_name"] if rows else "Baget Box"),
        minimum_confidence=float(inventory_config.get("minimum_confidence", 0.5)),
        minimum_area_px=int(inventory_config.get("minimum_object_area_px", 64)),
        duplicate_iou=float(inventory_config.get("duplicate_iou", 0.65)),
        inventory_roi=tuple(tuple(point) for point in inventory_config.get("roi", []) or []),
        ignore_zones=tuple(
            tuple(tuple(point) for point in zone)
            for zone in inventory_config.get("ignore_zones", []) or []
        ),
    )
    candidates = [
        InventoryCandidate(
            index=index, box=tuple(row["bbox"]),
            detector_class=row["detector_class"],
            detector_confidence=float(row["detector_confidence"]),
            recognized_name=row["suggested_name"],
            recognition_confidence=float(row["confidence"]),
            recognition_source=row["source"],
        )
        for index, row in enumerate(rows)
    ]
    inventory_result = counter.evaluate(candidates) if term else None
    if inventory_result is None:
        local_diag["accepted_detection_count"] = len(rows)
        local_diag["rejected_detection_count"] = max(0, len(detections) - len(rows))
        local_diag["final_inventory_count"] = len(rows)
    else:
        for decision in inventory_result.decisions:
            rows[decision.candidate.index]["keep"] = decision.accepted
            rows[decision.candidate.index]["rule_decision"] = decision.reason
        local_diag["accepted_detection_count"] = inventory_result.final_inventory_count
        local_diag["rejected_detection_count"] = len(inventory_result.rejected) + max(0, len(detections) - len(rows))
        local_diag["final_inventory_count"] = inventory_result.final_inventory_count
    if term and inventory_result is not None and inventory_result.final_inventory_count == 0:
        local_diag["failure_reason"] = "all_filtered_by_query"
    if term and inventory_result is not None:
        frame_uuid = secrets.token_hex(16)
        evidence_dir = TRAINING_STAGING_DIR / "benchmark"
        evidence_dir.mkdir(parents=True, exist_ok=True)
        evidence_path = evidence_dir / f"{frame_uuid}.jpg"
        cv2.imwrite(str(evidence_path), frame)
        health = detector.health() if hasattr(detector, "health") else {}
        local_diag["inventory_result_id"] = VisionDB(
            os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))
        ).record_inventory_result({
            "camera_id": cam_name, "block_id": str(block_id) if block_id is not None else None,
            "frame_uuid": frame_uuid, "target_product": term,
            "requested_model": health.get("requested_model"),
            "loaded_model": health.get("active_model"),
            "detector_mode": health.get("detector_mode", "unknown"),
            "fallback_used": bool(health.get("fallback_used")),
            "raw_detection_count": inventory_result.raw_detection_count,
            "accepted_detection_count": inventory_result.final_inventory_count,
            "rejected_detection_count": len(inventory_result.rejected),
            "final_inventory_count": inventory_result.final_inventory_count,
            "detections": [
                {"bbox": list(decision.candidate.box),
                 "class": decision.candidate.detector_class,
                 "confidence": decision.candidate.detector_confidence,
                 "recognized_name": decision.candidate.recognized_name,
                 "accepted": decision.accepted, "reason": decision.reason}
                for decision in inventory_result.decisions
            ],
            "evidence_path": str(evidence_path),
        })
    if diag is not None:
        for key in ("frames_read", "detections", "tracked", "raw_detection_count",
                    "accepted_detection_count", "rejected_detection_count", "final_inventory_count"):
            if key in diag or key in {"frames_read", "detections", "tracked"}:
                diag[key] = int(diag.get(key, 0)) + int(local_diag[key])
    return rows, seq, local_diag


def _training_search_sync(query: str) -> dict[str, Any]:
    """Synchronous single-pass sweep returning {rows, diagnostics}. The live
    UI uses the background job (start/status) instead, but this keeps a simple
    blocking entry point for internal callers and tests."""
    term = _catalog_normalize_name(query)
    health = _catalog_health_snapshot()
    cameras = _training_camera_map(health)
    detector, mode = _training_detection_context(query)
    refs = _training_dataset_reference_embeddings()
    diag = {
        "total_active_cameras": len(cameras), "attempted_cameras": 0,
        "frames_read": 0, "cameras_failed": 0, "cameras_completed": 0,
        "detections": 0, "tracked": 0, "raw_detection_count": 0,
        "accepted_detection_count": 0, "rejected_detection_count": 0,
        "final_inventory_count": 0, "model": detector is not None,
        "failures": [], **mode,
    }
    if detector is None:
        return {"rows": [], "diagnostics": diag}
    TRAINING_STAGING_DIR.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, Any]] = []
    seq = 0
    for slot, cam_name in sorted(cameras.items()):
        new_rows, seq, camera_diag = _training_scan_camera(slot, cam_name, term, detector, refs, None, seq)
        rows.extend(new_rows)
        _training_merge_camera_diagnostics(diag, slot, cam_name, camera_diag)
    rows.sort(key=lambda r: (str(r["camera"]), int(r["track_id"] or -1)))
    return {"rows": rows, "diagnostics": diag}


def _training_merge_camera_diagnostics(diag, slot, cam_name, camera_diag) -> None:
    diag["attempted_cameras"] = int(diag.get("attempted_cameras", 0)) + 1
    for key in ("frames_read", "detections", "tracked", "raw_detection_count",
                "accepted_detection_count", "rejected_detection_count", "final_inventory_count"):
        diag[key] = int(diag.get(key, 0)) + int(camera_diag.get(key, 0))
    reason = camera_diag.get("failure_reason")
    if reason in {"no_frame", "detector_unavailable", "camera_timeout"}:
        diag["cameras_failed"] = int(diag.get("cameras_failed", 0)) + 1
        diag.setdefault("failures", []).append({"slot": slot, "camera": cam_name, "reason": reason})
    else:
        diag["cameras_completed"] = int(diag.get("cameras_completed", 0)) + 1
    if reason:
        reasons = diag.setdefault("outcomes", {})
        reasons[reason] = int(reasons.get(reason, 0)) + 1


def _training_search_worker(
    query: str,
    generation: int,
    selected_slots: set[int] | None = None,
    block_id: int | None = None,
    camera_ids: list[int] | None = None,
) -> None:
    started = datetime.now(timezone.utc).isoformat()
    try:
        term = _catalog_normalize_name(query)
        health = _catalog_health_snapshot()
        cameras = _training_camera_map(health)
        if selected_slots is not None:
            cameras = {
                slot: name for slot, name in cameras.items() if slot in selected_slots
            }
        detector, mode = _training_detection_context(query)
        refs = _training_dataset_reference_embeddings()
    except Exception as exc:  # noqa: BLE001
        _training_search_write_state(
            {
                **_training_search_idle_state(),
                "status": "error",
                "stage": "error",
                "message": str(exc),
                "query": query,
                "error": str(exc),
                "started_at": started,
                "finished_at": datetime.now(timezone.utc).isoformat(),
            },
            generation,
        )
        return

    cam_items = sorted(cameras.items())
    diag = {
        "total_active_cameras": len(cameras), "attempted_cameras": 0,
        "frames_read": 0, "cameras_failed": 0, "cameras_completed": 0,
        "detections": 0, "tracked": 0, "raw_detection_count": 0,
        "accepted_detection_count": 0, "rejected_detection_count": 0,
        "final_inventory_count": 0, "model": detector is not None,
        "failures": [], "outcomes": {}, **mode,
    }
    base = {
        "query": query,
        "block_id": block_id,
        "camera_ids": camera_ids or [],
        "started_at": started,
        "finished_at": None,
    }
    if not _training_search_write_state(
        {**base, "status": "running", "stage": "capturing", "message": "Connecting to live cameras.", "current_camera": None, "rows": [], "diagnostics": diag, "progress": {"done": 0, "total": len(cam_items)}},
        generation,
    ):
        return

    rows: list[dict[str, Any]] = []
    try:
        workers = max(1, int(os.getenv("AI_VISION_SEARCH_WORKERS", "3")))
    except ValueError:
        workers = 3
    workers = min(workers, max(1, len(cam_items)), 8)
    try:
        camera_timeout = max(0.01, float(os.getenv("AI_VISION_CAMERA_SCAN_TIMEOUT_SECONDS", "90")))
    except ValueError:
        camera_timeout = 90.0
    if detector is not None:
        from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait

        TRAINING_STAGING_DIR.mkdir(parents=True, exist_ok=True)
        pool = ThreadPoolExecutor(max_workers=workers, thread_name_prefix="camera-scan")
        submitted = {
            pool.submit(_training_scan_camera, slot, cam_name, term, detector, refs, None, 0, block_id):
            (slot, cam_name, time.monotonic())
            for slot, cam_name in cam_items
        }
        try:
            while submitted:
                completed, _ = wait(submitted, timeout=0.25, return_when=FIRST_COMPLETED)
                now = time.monotonic()
                timed_out = {
                    future for future, (_, _, submitted_at) in submitted.items()
                    if not future.done() and now - submitted_at >= camera_timeout
                }
                for future in completed | timed_out:
                    slot, cam_name, _ = submitted.pop(future)
                    camera_diag = {
                        "frames_read": 0, "detections": 0, "tracked": 0,
                        "raw_detection_count": 0, "accepted_detection_count": 0,
                        "rejected_detection_count": 0, "final_inventory_count": 0,
                        "failure_reason": None,
                    }
                    if future in timed_out:
                        future.cancel()
                        new_rows = []
                        camera_diag["failure_reason"] = "camera_timeout"
                    else:
                        try:
                            new_rows, _, camera_diag = future.result()
                            rows.extend(new_rows)
                        except Exception as exc:  # noqa: BLE001
                            camera_diag["failure_reason"] = "detector_unavailable"
                            _audit("training_search_camera_failed", {"slot": slot, "error": str(exc)})
                    _training_merge_camera_diagnostics(diag, slot, cam_name, camera_diag)
                    rows.sort(key=lambda row: (str(row["camera"]), int(row["track_id"] or -1)))
                    done = int(diag["attempted_cameras"])
                    if not _training_search_write_state(
                        {**base, "status": "running", "stage": "processing", "message": f"Processed {cam_name}.", "current_camera": cam_name, "rows": rows, "diagnostics": diag, "progress": {"done": done, "total": len(cam_items)}},
                        generation,
                    ):
                        return
        finally:
            # A timed-out inference cannot be force-killed safely in Python;
            # abandon it without blocking job completion. It owns only local
            # result state, so a late return cannot mutate the published scan.
            pool.shutdown(wait=False, cancel_futures=True)

    rows.sort(key=lambda row: (str(row["camera"]), int(row["track_id"] or -1)))
    processed = int(diag.get("attempted_cameras", 0))
    final_count = int(diag.get("final_inventory_count", 0))
    final_message = (
        f"Scan complete: {processed}/{len(cam_items)} active cameras; "
        f"{final_count} target product(s) visible."
    )
    _training_search_write_state(
        {
            **base,
            "status": "done",
            "stage": "completed",
            "message": final_message,
            "current_camera": None,
            "rows": rows,
            "diagnostics": diag,
            "progress": {"done": processed, "total": len(cam_items)},
            "finished_at": datetime.now(timezone.utc).isoformat(),
        },
        generation,
    )


def _training_search_run(
    query: str,
    generation: int,
    selected_slots: set[int] | None = None,
    block_id: int | None = None,
    camera_ids: list[int] | None = None,
) -> None:
    """Thread entry point that tracks worker liveness so a restart-orphaned
    'running' state can be detected as stale."""
    global _training_search_active
    try:
        _training_search_worker(
            query, generation, selected_slots, block_id, camera_ids
        )
    finally:
        if generation == _training_search_generation:
            _training_search_active = False


def _training_search_start(
    query: str,
    selected_slots: set[int] | None = None,
    block_id: int | None = None,
    camera_ids: list[int] | None = None,
) -> dict[str, Any]:
    global _training_search_generation, _training_search_active
    with _training_search_lock:
        running = _training_search_active and _training_search_state.get("status") == "running"
        if running:
            # There is one recognition workflow and one tracker timeline per
            # camera. Never overlap scans or mutate a tracker from two workers.
            return dict(_training_search_state)
        _training_search_generation += 1
        generation = _training_search_generation
        # Mark alive synchronously (before the thread starts) so a status poll
        # in the start-up window doesn't mis-flag the fresh job as stale.
        _training_search_active = True
    _training_search_write_state(
        {
            **_training_search_idle_state(),
            "status": "running",
            "stage": "queued",
            "message": "Scan queued; preparing live camera capture.",
            "query": query,
            "block_id": block_id,
            "camera_ids": camera_ids or [],
            "started_at": datetime.now(timezone.utc).isoformat(),
        },
        generation,
    )
    threading.Thread(
        target=_training_search_run,
        args=(query, generation, selected_slots, block_id, camera_ids),
        daemon=True,
    ).start()
    return _training_search_status()


@app.post("/api/v1/scan/start")
async def start_block_scan(payload: BlockScanStart) -> dict[str, Any]:
    product = next(
        (row for row in _scan_products()
         if int(row["id"]) == payload.product_id),
        None,
    )
    if product is None:
        raise HTTPException(status_code=404, detail="Selected product was not found.")
    assigned = {
        int(row["camera_id"]): row
        for row in _camera_operations_payload()
        if row.get("block_id") is not None
        and int(row["block_id"]) == payload.block_id
    }
    requested = set(payload.camera_ids)
    outside_block = sorted(requested - set(assigned))
    if outside_block:
        raise HTTPException(
            status_code=422,
            detail="Every selected camera must belong to the selected Block.",
        )
    selected_slots = {
        int(assigned[camera_id]["slot_number"])
        for camera_id in requested
        if assigned[camera_id].get("is_active")
        and assigned[camera_id].get("slot_number") is not None
    }
    if not selected_slots:
        raise HTTPException(
            status_code=409,
            detail="This Block has no active cameras available to scan.",
        )
    VisionDB(os.getenv("VISION_DB_PATH", str(ROOT / "database" / "vision.db"))).record_operator_action(
        "block_scan_started", "operator",
        {"block_id": payload.block_id, "camera_ids": sorted(requested),
         "product_id": payload.product_id, "product_name": product["name"]},
    )
    return await asyncio.to_thread(
        _training_search_start,
        str(product["name"]),
        selected_slots,
        payload.block_id,
        sorted(requested),
    )


@app.post("/api/training/search")
async def training_search(payload: TrainingSearch) -> dict[str, Any]:
    """Start the single camera-scan workflow as a background job.

    Every live frame passes through YOLO and the camera's persistent ByteTrack
    adapter. Each tracked object is cropped and compared with the local dataset;
    only insufficient local matches fall back to the naming service. An empty
    query keeps every object. Poll the status endpoint for real progress and
    individual-object rows.
    """
    return await asyncio.to_thread(_training_search_start, payload.query)


@app.get("/api/training/search/status")
async def training_search_status() -> dict[str, Any]:
    """Current recognition-job state (running/done/idle) with progress and the
    rows found so far - used to restore results when the page is reopened."""
    return _training_search_status()


@app.get("/api/training/search/export")
async def training_search_export() -> Response:
    """Export the latest individual-object scan without starting another scan."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    state = _training_search_status()
    rows = list(state.get("rows") or [])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Camera Scan"
    headers = [
        "Object Preview",
        "Suggested Name",
        "Final Name",
        "Camera",
        "Track ID",
        "Confidence",
        "Decision",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for row in rows:
        sheet.append(
            [
                row.get("crop_url") or "",
                row.get("suggested_name") or "",
                row.get("name") or row.get("suggested_name") or "",
                row.get("camera") or "",
                row.get("track_id"),
                row.get("confidence", 0.0),
                "Keep" if row.get("keep", True) else "Ignore",
            ]
        )
    for column, width in zip("ABCDEFG", [52, 28, 28, 28, 12, 14, 12]):
        sheet.column_dimensions[column].width = width
    stream = io.BytesIO()
    workbook.save(stream)
    filename = f"camera-scan-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/catalog/items")
async def create_catalog_item(
    scope_id: str = Form(...),
    name: str = Form(...),
    files: list[UploadFile] = File(...),
) -> dict[str, Any]:
    import cv2
    import numpy as np
    from recognition.embedding import image_embedding

    scope = _catalog_scope(scope_id)
    item_name = " ".join(name.split()).strip()
    if not item_name:
        raise HTTPException(status_code=400, detail="Item name is required.")
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="Upload at least two reference images.")
    if len(files) > 12:
        raise HTTPException(status_code=400, detail="Upload no more than twelve reference images.")

    decoded: list[tuple[str, bytes, Any, list[float]]] = []
    for upload in files:
        if not (upload.content_type or "").startswith("image/"):
            raise HTTPException(status_code=400, detail=f"{upload.filename or 'File'} is not an image.")
        contents = await upload.read()
        if not contents or len(contents) > 8 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Each image must be between 1 byte and 8 MB.")
        frame = cv2.imdecode(np.frombuffer(contents, dtype=np.uint8), cv2.IMREAD_COLOR)
        if frame is None:
            raise HTTPException(status_code=400, detail=f"{upload.filename or 'File'} could not be decoded.")
        decoded.append(
            (
                _catalog_safe_name(upload.filename or "reference.jpg"),
                contents,
                frame,
                image_embedding(frame),
            )
        )

    db = _get_catalog_db()
    if any(_catalog_normalize_name(item["name"]) == _catalog_normalize_name(item_name) for item in db.list_items(scope)):
        raise HTTPException(status_code=409, detail="An item with this name already exists.")
    item = db.create_item(scope, item_name)
    item_dir = CATALOG_IMAGE_DIR / scope / str(item["id"])
    item_dir.mkdir(parents=True, exist_ok=True)
    for index, (original_name, contents, frame, embedding) in enumerate(decoded, start=1):
        suffix = Path(original_name).suffix.lower()
        if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
            suffix = ".jpg"
        filename = f"reference_{index:02d}{suffix}"
        path = item_dir / filename
        path.write_bytes(contents)
        url = f"/snapshots/catalog/{quote(scope)}/{quote(str(item['id']))}/{quote(filename)}"
        db.add_image(
            item_id=str(item["id"]),
            filename=filename,
            url=url,
            embedding=embedding,
            width_px=int(frame.shape[1]),
            height_px=int(frame.shape[0]),
        )
    _audit(
        "catalog_item_created",
        {"scope_id": scope, "item_id": item["id"], "name": item_name, "image_count": len(decoded)},
    )
    recognition = await asyncio.to_thread(_run_catalog_recognition, scope)
    return {
        "item": db.get_item(str(item["id"])),
        "schedule": recognition["schedule"],
        "recognition": recognition,
    }


@app.delete("/api/catalog/items/{item_id}")
def delete_catalog_item(item_id: str, scope_id: str) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    db = _get_catalog_db()
    item = db.get_item(item_id)
    if not item or item["scope_id"] != scope:
        raise HTTPException(status_code=404, detail="Catalog item not found.")
    filenames = db.delete_item(item_id)
    item_dir = (CATALOG_IMAGE_DIR / scope / item_id).resolve()
    catalog_root = CATALOG_IMAGE_DIR.resolve()
    if item_dir.is_relative_to(catalog_root):
        for filename in filenames:
            path = (item_dir / filename).resolve()
            if path.is_relative_to(item_dir) and path.exists():
                path.unlink()
        try:
            item_dir.rmdir()
        except OSError:
            pass
    _audit("catalog_item_deleted", {"scope_id": scope, "item_id": item_id, "name": item["name"]})
    return {"deleted": True, "item_id": item_id}


@app.get("/api/catalog/results")
def catalog_results(scope_id: str) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    db = _get_catalog_db()
    return {
        "run": db.latest_run(scope),
        "results": db.latest_results(scope, detected_only=True),
        "schedule": _catalog_schedule(scope),
    }


@app.get("/api/catalog/results/history")
def catalog_results_history(scope_id: str, limit: int = 200) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    db = _get_catalog_db()
    return {
        "scope_id": scope,
        "results": db.result_history(scope, limit=limit),
        "schedule": _catalog_schedule(scope),
    }


@app.post("/api/catalog/recognition/run")
async def run_catalog_recognition(scope_id: str) -> dict[str, Any]:
    global _catalog_run_lock
    scope = _catalog_scope(scope_id)
    if _catalog_run_lock is None:
        _catalog_run_lock = asyncio.Lock()
    if _catalog_run_lock.locked():
        raise HTTPException(status_code=409, detail="A catalog recognition run is already active.")
    async with _catalog_run_lock:
        result = await asyncio.to_thread(_run_catalog_recognition, scope)
    _audit("catalog_recognition_completed", {"scope_id": scope, "run_id": result["run"]["id"]})
    return result


@app.post("/api/catalog/recognition/run-live")
async def start_live_catalog_recognition(scope_id: str) -> dict[str, Any]:
    """Sample the catalog's items against the live camera feeds repeatedly
    over CATALOG_LIVE_RUN_DURATION_SECONDS instead of a single instant
    snapshot - a slow-moving or briefly-occluded item is more likely to be
    caught. Only matches items enrolled via AI Check-in (same catalog the
    scheduled/instant recognition uses)."""
    scope = _catalog_scope(scope_id)
    existing = _live_catalog_runs.get(scope)
    if existing and existing["status"] == "running":
        return _live_catalog_status_payload(scope)

    started_at = datetime.now(timezone.utc)
    ends_at = started_at + timedelta(seconds=CATALOG_LIVE_RUN_DURATION_SECONDS)
    _live_catalog_runs[scope] = {
        "status": "running",
        "started_at": started_at.isoformat(),
        "ends_at": ends_at.isoformat(),
        "items": {},
    }
    _live_catalog_tasks[scope] = asyncio.create_task(_run_live_catalog_recognition(scope, ends_at))
    return _live_catalog_status_payload(scope)


@app.get("/api/catalog/recognition/run-live/status")
def live_catalog_recognition_status(scope_id: str) -> dict[str, Any]:
    scope = _catalog_scope(scope_id)
    return _live_catalog_status_payload(scope)


def _catalog_export_workbook(scope_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    db = _get_catalog_db()
    run = db.latest_run(scope_id)
    results = db.latest_results(scope_id, detected_only=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detected Items"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A6"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "AI Vision — Detected Item Count"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet["A3"] = "Recognition run"
    sheet["B3"] = str((run or {}).get("completed_at") or "No completed run")
    sheet["D3"] = "Schedule"
    sheet["E3"] = f"Every {_catalog_interval_hours()} hours"
    sheet["A4"] = "Scope"
    sheet["B4"] = scope_id
    sheet["D4"] = "Detected item types"
    sheet["E4"] = len(results)
    headers = [
        "Item",
        "Count",
        "Camera / objects",
        "Confidence",
        "Width (cm)",
        "Height (cm)",
        "Depth (cm)",
        "3D method",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="2563EB")
    for cell in sheet[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for result in results:
        sheet.append(
            [
                result["item_name"],
                int(result["quantity"]),
                ", ".join(
                    f"{entry.get('camera_name')}: {int(entry.get('quantity') or 0)}"
                    for entry in result.get("camera_counts") or []
                    if int(entry.get("quantity") or 0) > 0
                )
                or "Unknown camera",
                float(result["confidence"]),
                round(float(result["width_m"]) * 100, 1) if result.get("width_m") else None,
                round(float(result["height_m"]) * 100, 1) if result.get("height_m") else None,
                round(float(result["depth_m"]) * 100, 1) if result.get("depth_m") else None,
                result.get("measurement_method") or "Not measured",
            ]
        )
    if results:
        table = Table(displayName="DetectedItems", ref=f"A5:H{5 + len(results)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 30
    for row in sheet.iter_rows(min_row=6, max_row=5 + len(results), min_col=2, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for cell in sheet["D"][5:]:
        cell.number_format = "0.0%"
    for column in ("E", "F", "G"):
        for cell in sheet[column][5:]:
            cell.number_format = "0.0"
    sheet.auto_filter.ref = f"A5:H{max(5, 5 + len(results))}"
    sheet.print_title_rows = "1:5"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@app.get("/api/catalog/results/export.xlsx")
def export_catalog_results(scope_id: str) -> Response:
    scope = _catalog_scope(scope_id)
    content = _catalog_export_workbook(scope)
    filename = f"detected-items-{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v2/companies")
def list_companies() -> dict[str, Any]:
    return {"companies": _get_accounts_db().list_companies()}


@app.post("/api/v2/companies")
def create_company(payload: CompanyCreate) -> dict[str, Any]:
    company = _get_accounts_db().create_company(payload.name.strip())
    _audit("company_created", {"company_id": company["id"], "name": company["name"]})
    return company


@app.put("/api/v2/companies/{company_id}")
def rename_company(company_id: str, payload: CompanyRename) -> dict[str, Any]:
    company = _get_accounts_db().rename_company(company_id, payload.name.strip())
    if company is None:
        raise HTTPException(status_code=404, detail="Company not found.")
    return company


@app.delete("/api/v2/companies/{company_id}")
def delete_company(company_id: str) -> dict[str, Any]:
    deleted = _get_accounts_db().delete_company(company_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Company not found.")
    _audit("company_deleted", {"company_id": company_id})
    return {"deleted": True, "company_id": company_id}


@app.put("/api/v2/companies/{company_id}/camera-config")
def update_camera_config(company_id: str, payload: CameraConfigUpdate) -> dict[str, Any]:
    nvrs = payload.cameraConfig.get("nvrs") or []
    if len(nvrs) > 5:
        raise HTTPException(status_code=400, detail="A company may connect at most 5 NVRs.")
    for nvr in nvrs:
        if int(nvr.get("slots") or 0) > 15:
            raise HTTPException(status_code=400, detail="Each NVR may expose at most 15 camera slots.")
    company = _get_accounts_db().set_camera_config(company_id, payload.cameraConfig)
    if company is None:
        raise HTTPException(status_code=404, detail="Company not found.")
    return company


@app.post("/api/v2/companies/{company_id}/roles")
def create_role(company_id: str, payload: RoleCreate) -> dict[str, Any]:
    if _get_accounts_db().get_company(company_id) is None:
        raise HTTPException(status_code=404, detail="Company not found.")
    try:
        role = _get_accounts_db().create_role(
            company_id,
            payload.name.strip(),
            payload.login.strip(),
            payload.password,
            access_camera=payload.access_camera,
            access_analytics=payload.access_analytics,
        )
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    _audit("role_created", {"company_id": company_id, "role_id": role["id"], "login": role["login"]})
    return role


@app.put("/api/v2/roles/{role_id}")
def update_role(role_id: str, payload: RoleUpdate) -> dict[str, Any]:
    if _get_accounts_db().get_role(role_id) is None:
        raise HTTPException(status_code=404, detail="Role not found.")
    try:
        role = _get_accounts_db().update_role(
            role_id,
            name=payload.name.strip() if payload.name is not None else None,
            login=payload.login.strip() if payload.login is not None else None,
            password=payload.password,
            access_camera=payload.access_camera,
            access_analytics=payload.access_analytics,
        )
    except Exception as exc:  # duplicate login, etc.
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return role  # type: ignore[return-value]


@app.delete("/api/v2/roles/{role_id}")
def delete_role(role_id: str) -> dict[str, Any]:
    deleted = _get_accounts_db().delete_role(role_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Role not found.")
    _audit("role_deleted", {"role_id": role_id})
    return {"deleted": True, "role_id": role_id}


@app.get("/api/v2/accounts/{role_id}")
def get_account(role_id: str) -> dict[str, Any]:
    account = _get_accounts_db().get_role_public(role_id)
    if account is None:
        raise HTTPException(status_code=404, detail="Account not found.")
    return account


@app.get("/api/v2/admin/profile")
def get_admin_profile() -> dict[str, Any]:
    return _get_accounts_db().get_profile()


@app.put("/api/v2/admin/profile")
def update_admin_profile(payload: ProfileUpdate) -> dict[str, Any]:
    if payload.avatar and len(payload.avatar) > 3 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Profile picture is too large.")
    avatar_value: str | None = "__unset__"
    if payload.remove_avatar:
        avatar_value = None
    elif payload.avatar is not None:
        avatar_value = payload.avatar
    return _get_accounts_db().update_profile(
        login=payload.login.strip() if payload.login else None,
        password=payload.password,
        avatar=avatar_value,
    )


@app.get("/api/logs/stream")
async def stream_logs():
    async def event_generator():
        last_pos = 0
        while True:
            if LOG_PATH.exists():
                try:
                    with LOG_PATH.open("r", encoding="utf-8", errors="replace") as f:
                        f.seek(last_pos)
                        data = f.read()
                        if data:
                            for line in data.splitlines():
                                yield f"data: {line}\n\n"
                        last_pos = f.tell()
                except Exception:
                    # swallow errors and continue polling
                    pass
            await asyncio.sleep(0.5)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/live_mjpeg")
async def live_mjpeg(slot: int | None = None, camera: str | None = None):
    """Stream frames from the persistent Stream Manager's shared buffer."""

    boundary = "frame"

    async def frame_generator():
        last_sent: bytes | None = None
        while True:
            data = _get_stream_manager().latest_frame_bytes(slot_number=slot, name=camera)
            # Only push a part when the frame actually changed, so the MJPEG
            # stream tracks the Stream Manager's real frame rate instead of
            # re-transmitting the same JPEG on every poll.
            if data is not None and data is not last_sent and data != last_sent:
                try:
                    header = (
                        f"--{boundary}\r\n"
                        "Content-Type: image/jpeg\r\n"
                        f"Content-Length: {len(data)}\r\n\r\n"
                    ).encode("utf-8")
                    yield header + data + b"\r\n"
                    last_sent = data
                except Exception:
                    # ignore read errors
                    pass
            # Poll faster than the source frame rate so new frames are forwarded
            # with minimal latency (Stream Manager publishes up to ~15 fps).
            await asyncio.sleep(0.03)

    return StreamingResponse(frame_generator(), media_type=f"multipart/x-mixed-replace; boundary={boundary}")


@app.websocket("/api/live_ws")
async def live_websocket(websocket: WebSocket):
    """Multiplex every requested camera over one continuous connection.

    Binary messages contain a two-byte unsigned slot number followed by one
    complete JPEG. The Stream Manager remains the sole RTSP decoder owner.
    """

    raw_slots = websocket.query_params.get("slots", "")
    slots = sorted(
        {
            int(value)
            for value in raw_slots.split(",")
            if value.strip().isdigit() and 1 <= int(value) <= MAX_CAMERA_SLOTS
        }
    )
    if not slots:
        await websocket.close(code=1008, reason="At least one camera slot is required.")
        return

    await websocket.accept()
    last_sent: dict[int, bytes] = {}
    target_fps = max(1.0, min(float(os.getenv("LIVE_WEBSOCKET_FPS", "10")), 15.0))
    interval = 1.0 / target_fps
    try:
        while True:
            cycle_started = time.monotonic()
            for slot in slots:
                data = _get_stream_manager().latest_frame_bytes(slot_number=slot)
                if data is None or data == last_sent.get(slot):
                    continue
                await websocket.send_bytes(struct.pack("!H", slot) + data)
                last_sent[slot] = data
            elapsed = time.monotonic() - cycle_started
            await asyncio.sleep(max(0.001, interval - elapsed))
    except (WebSocketDisconnect, RuntimeError):
        return


@app.get("/api/live_frame")
async def live_frame(slot: int | None = None, camera: str | None = None):
    """Return the latest processed JPEG frame for one camera.

    The dashboard grid uses this polling endpoint instead of opening one
    long-lived MJPEG connection per slot. Browsers often cap concurrent
    connections per origin, so 10 simultaneous MJPEG streams can leave some
    screens stuck on "Waiting for frames" even when the backend is healthy.
    """

    data = _get_stream_manager().latest_frame_bytes(slot_number=slot, name=camera)
    if data and data.startswith(b"\xff\xd8") and data.endswith(b"\xff\xd9"):
        return Response(
            content=data,
            media_type="image/jpeg",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "X-AI-Frame-Source": "stream-manager",
            },
        )

    raise HTTPException(status_code=404, detail="No live frame is available yet.")


def _live_feed_path(slot: int | None = None, camera: str | None = None) -> Path:
    if slot is not None:
        return SNAPSHOT_DIR / f"latest_stream_slot_{slot}.jpg"
    if camera:
        safe_name = "".join(ch if ch.isalnum() else "_" for ch in camera).strip("_") or "camera"
        return SNAPSHOT_DIR / f"latest_stream_{safe_name}.jpg"
    return SNAPSHOT_DIR / "latest_stream.jpg"


def _live_feed_paths(slot: int | None = None, camera: str | None = None) -> list[Path]:
    paths = [_live_feed_path(slot=slot, camera=camera)]
    if slot is not None:
        paths.append(SNAPSHOT_DIR / f"latest_slot_{slot}.jpg")
    if camera:
        safe_name = "".join(ch if ch.isalnum() else "_" for ch in camera).strip("_") or "camera"
        paths.append(SNAPSHOT_DIR / f"latest_{safe_name}.jpg")
    if slot is None and not camera:
        paths.append(SNAPSHOT_DIR / "latest.jpg")
    return paths


app.mount("/dashboard-v2/assets", StaticFiles(directory=DASHBOARD_V2_DIR), name="dashboard-v2-assets")
app.mount("/assets", StaticFiles(directory=DASHBOARD_V2_DIR), name="dashboard-assets")
app.mount("/snapshots", StaticFiles(directory=SNAPSHOT_DIR), name="snapshots")
